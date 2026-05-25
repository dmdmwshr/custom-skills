#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from validate_extractions import ALLOWED_VALUES, load_cases, validate_cases

RED_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
NO_FILL = PatternFill(fill_type=None)

FIELD_HEADERS = {
    "unit_name": ("单位名称",),
    "project_no": ("项目编号",),
    "unit_address": ("单位地址",),
    "nominal_producers_text": ("标称生产者",),
    "case_handler": ("立卷人",),
    "inspector": ("检查人",),
    "inspection_month": ("检查时间",),
    "products_text": ("消防产品",),
    "station_or_team": ("是否为微型消防站或快速处置队",),
    "method": ("现场判定或抽样送检",),
    "qualified": ("是否合格",),
    "case_type": ("是否为行案/刑案", "是否为行案刑案"),
    "online_sale": ("是否为网售",),
}

VALIDATION_VALUES = {
    "inspection_month": [f"{i}月" for i in range(1, 13)],
    "station_or_team": ["微型消防站", "快速处置队", "否"],
    "method": ["现场判定", "抽样送检"],
    "qualified": ["合格", "不合格"],
    "case_type": ["刑案", "行案", "否"],
    "online_sale": ["是", "否"],
}

ROLE_VALUES = {"案卷清单", "检查记录表", "其他附件"}
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def parse_month(value: str) -> str:
    value = text(value)
    if not value:
        return ""
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", value)
    if match:
        return f"{int(match.group(2))}月"
    match = re.search(r"\d{4}年(\d{1,2})月\d{1,2}日", value)
    if match:
        return f"{int(match.group(1))}月"
    match = re.search(r"(^|\D)(\d{1,2})月", value)
    if match:
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{month}月"
    return ""


def parse_iso_date(value: str) -> str:
    value = text(value)
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", value)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", value)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return "未知日期"


def list_text(values: Any) -> str:
    if not isinstance(values, list):
        return ""
    return "\n".join(text(item) for item in values if text(item))


def sanitize_filename_part(value: str, fallback: str = "未命名") -> str:
    cleaned = INVALID_FILENAME_CHARS.sub("_", text(value))
    cleaned = re.sub(r"\s+", "", cleaned).strip(" ._")
    if not cleaned:
        cleaned = fallback
    return cleaned[:80]


def header_map(ws: Any) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        value = normalize(cell.value)
        if value:
            mapping[value] = cell.column
    return mapping


def find_header_column(ws: Any, candidates: tuple[str, ...]) -> int:
    mapping = header_map(ws)
    normalized_candidates = [normalize(candidate) for candidate in candidates]
    for candidate in normalized_candidates:
        if candidate in mapping:
            return mapping[candidate]
    for header, column in mapping.items():
        if any(candidate and candidate in header for candidate in normalized_candidates):
            return column
    raise ValueError(f"{ws.title} 缺少表头：{'/'.join(candidates)}")


def field_columns(ws: Any) -> dict[str, int]:
    return {field: find_header_column(ws, headers) for field, headers in FIELD_HEADERS.items()}


def allowed_or_blank(field: str, value: Any, warnings: list[str], project_no: str) -> str:
    candidate = text(value)
    if not candidate:
        return ""
    allowed = ALLOWED_VALUES.get(field)
    if allowed and candidate not in allowed:
        warnings.append(f"{project_no}: {field} 非法值已留空：{candidate}")
        return ""
    return candidate


def find_project_row(ws: Any, project_col: int, project_no: str) -> int | None:
    for row in range(2, ws.max_row + 1):
        if text(ws.cell(row, project_col).value) == project_no:
            return row
    return None


def first_empty_row(ws: Any, columns: dict[str, int]) -> int | None:
    data_cols = [columns[key] for key in FIELD_HEADERS if key != "project_no"]
    for row in range(2, ws.max_row + 1):
        if all(text(ws.cell(row, col).value) == "" for col in data_cols):
            return row
    return None


def copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def extend_data_validations(ws: Any, row: int, columns: dict[str, int]) -> None:
    if not ws.data_validations:
        return
    for dv in ws.data_validations.dataValidation:
        formula = text(dv.formula1)
        for field, values in VALIDATION_VALUES.items():
            if all(value in formula for value in values):
                dv.add(ws.cell(row, columns[field]).coordinate)


def append_row(ws: Any, columns: dict[str, int]) -> int:
    target_row = ws.max_row + 1
    copy_row_style(ws, ws.max_row, target_row)
    previous_serial = ws.cell(target_row - 1, 1).value
    try:
        serial = int(previous_serial) + 1
    except Exception:
        serial = target_row - 1
    ws.cell(target_row, 1).value = serial
    extend_data_validations(ws, target_row, columns)
    return target_row


def target_values(case: dict[str, Any], warnings: list[str]) -> dict[str, str]:
    project_no = text(case.get("project_no"))
    return {
        "unit_name": text(case.get("unit_name")),
        "project_no": project_no,
        "unit_address": text(case.get("unit_address")),
        "nominal_producers_text": list_text(case.get("nominal_producers")),
        "case_handler": text(case.get("case_handler")),
        "inspector": text(case.get("inspector")),
        "inspection_month": parse_month(text(case.get("inspection_date"))),
        "products_text": list_text(case.get("products")),
        "station_or_team": allowed_or_blank("station_or_team", case.get("station_or_team"), warnings, project_no),
        "method": allowed_or_blank("method", case.get("method"), warnings, project_no),
        "qualified": allowed_or_blank("qualified", case.get("qualified"), warnings, project_no),
        "case_type": allowed_or_blank("case_type", case.get("case_type"), warnings, project_no),
        "online_sale": allowed_or_blank("online_sale", case.get("online_sale"), warnings, project_no),
    }


def set_cell_value(cell: Any, value: str) -> None:
    if value:
        cell.value = value
        cell.fill = copy(NO_FILL)
    else:
        cell.value = None
        cell.fill = copy(RED_FILL)


def update_case_row(ws: Any, row: int, columns: dict[str, int], case: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    values = target_values(case, warnings)
    for key, value in values.items():
        set_cell_value(ws.cell(row, columns[key]), value)
    return {
        "project_no": values["project_no"],
        "row": row,
        "unit_name": values["unit_name"],
    }


def unique_target_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 2
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def resolve_source_file(image_dir: Path, file_name: str) -> Path:
    path = Path(file_name)
    if not path.is_absolute():
        path = image_dir / file_name
    return path


def rename_case_images(image_dir: Path, case: dict[str, Any]) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    date_part = parse_iso_date(text(case.get("inspection_date")))
    unit_part = sanitize_filename_part(text(case.get("unit_name")), "未知单位")
    roles = case.get("file_roles") if isinstance(case.get("file_roles"), dict) else {}

    for file_name in case.get("source_files", []):
        source = resolve_source_file(image_dir, file_name)
        role = roles.get(file_name, "其他附件")
        if role not in ROLE_VALUES:
            role = "其他附件"
        if not source.exists():
            results.append({"source": str(source), "status": "missing"})
            continue
        target_name = f"{date_part}_{unit_part}_{role}{source.suffix.lower()}"
        raw_target = source.with_name(target_name)
        if source.resolve() == raw_target.resolve():
            results.append({"source": str(source), "target": str(raw_target), "status": "unchanged"})
            continue
        target = unique_target_path(raw_target)
        source.rename(target)
        results.append({"source": str(source), "target": str(target), "status": "renamed"})
    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="更新产品案卷数据工作簿并原地重命名截图。")
    parser.add_argument("--workbook", required=True, type=Path, help="产品案卷数据.xlsx")
    parser.add_argument("--brigade", required=True, help="工作表/大队名称")
    parser.add_argument("--image-dir", required=True, type=Path, help="截图目录")
    parser.add_argument("--extractions", required=True, type=Path, help="抽取结果 JSON")
    parser.add_argument("--dry-run", action="store_true", help="只校验并模拟写入，不保存工作簿、不重命名截图")
    parser.add_argument("--skip-rename", action="store_true", help="只更新工作簿，不重命名截图")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    warnings: list[str] = []
    cases = load_cases(args.extractions)
    validation = validate_cases(cases, args.image_dir)
    if validation["errors"]:
        print(json.dumps(validation, ensure_ascii=False, indent=2))
        return 1
    warnings.extend(validation["warnings"])

    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)
    if not args.image_dir.exists():
        raise FileNotFoundError(args.image_dir)

    wb = load_workbook(args.workbook)
    if args.brigade not in wb.sheetnames:
        raise ValueError(f"工作簿中未找到工作表：{args.brigade}")
    ws = wb[args.brigade]
    columns = field_columns(ws)

    updated: list[dict[str, Any]] = []
    for case in cases:
        project_no = text(case.get("project_no"))
        row = find_project_row(ws, columns["project_no"], project_no)
        operation = "update"
        if row is None:
            row = first_empty_row(ws, columns)
            operation = "insert"
        if row is None:
            row = append_row(ws, columns)
            operation = "append"
        info = update_case_row(ws, row, columns, case, warnings)
        info["operation"] = operation
        updated.append(info)

    backup_path = None
    renamed: list[dict[str, str]] = []
    if not args.dry_run:
        backup_path = args.workbook.with_name(
            f"{args.workbook.stem}_自动填充备份_{datetime.now():%Y%m%d_%H%M%S}{args.workbook.suffix}"
        )
        shutil.copy2(args.workbook, backup_path)
        wb.save(args.workbook)
        if not args.skip_rename:
            for case in cases:
                renamed.extend(rename_case_images(args.image_dir, case))

    result = {
        "status": "ok",
        "dry_run": args.dry_run,
        "workbook": str(args.workbook),
        "backup": str(backup_path) if backup_path else None,
        "sheet": args.brigade,
        "updated_cases": updated,
        "renamed_files": renamed,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except PermissionError as exc:
        print(json.dumps({"status": "errors_found", "errors": [f"文件可能被 WPS/Excel 占用：{exc}"]}, ensure_ascii=False, indent=2))
        sys.exit(1)
    except Exception as exc:
        print(json.dumps({"status": "errors_found", "errors": [str(exc)]}, ensure_ascii=False, indent=2))
        sys.exit(1)
