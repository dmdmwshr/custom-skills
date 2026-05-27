#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.datavalidation import DataValidationList


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm(value: Any) -> str:
    return "".join(text(value).split())


def header_col(ws: Any, header: str) -> int | None:
    target = norm(header)
    for cell in ws[1]:
        value = norm(cell.value)
        if value == target or target in value:
            return cell.column
    return None


def header_values(ws: Any) -> list[str]:
    return [text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]


def sheet_has_data(ws: Any) -> bool:
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            if text(ws.cell(row, col).value):
                return True
    return False


def copy_cell_format(source: Any, target: Any) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def copy_dimensions(source_ws: Any, target_ws: Any, max_col: int, max_row: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        source_dim = source_ws.column_dimensions[letter]
        target_dim = target_ws.column_dimensions[letter]
        target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden
        target_dim.outlineLevel = source_dim.outlineLevel
        target_dim.bestFit = source_dim.bestFit
    for row in range(1, max_row + 1):
        source_dim = source_ws.row_dimensions[row]
        target_dim = target_ws.row_dimensions[row]
        target_dim.height = source_dim.height
        target_dim.hidden = source_dim.hidden
        target_dim.outlineLevel = source_dim.outlineLevel


def copy_data_validations(source_ws: Any, target_ws: Any) -> None:
    target_ws.data_validations = DataValidationList()
    for dv in source_ws.data_validations.dataValidation:
        target_ws.add_data_validation(copy(dv))
    ensure_phase_validation(target_ws)


def ensure_phase_validation(ws: Any) -> None:
    phase_col = header_col(ws, "初查/复查")
    if not phase_col:
        return
    letter = get_column_letter(phase_col)
    formula = '"初查,复查"'
    for dv in ws.data_validations.dataValidation:
        if text(dv.formula1) == formula:
            dv.add(f"{letter}2:{letter}1048576")
            return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}1048576")


def copy_merged_cells(source_ws: Any, target_ws: Any) -> None:
    for merged_range in list(target_ws.merged_cells.ranges):
        target_ws.unmerge_cells(str(merged_range))
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))


def overwrite_empty_sheet(source_ws: Any, target_ws: Any) -> None:
    max_col = max(source_ws.max_column, target_ws.max_column)
    max_row = max(source_ws.max_row, target_ws.max_row)
    copy_merged_cells(source_ws, target_ws)
    copy_dimensions(source_ws, target_ws, source_ws.max_column, source_ws.max_row)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            target = target_ws.cell(row, col)
            if row <= source_ws.max_row and col <= source_ws.max_column:
                source = source_ws.cell(row, col)
                copy_cell_format(source, target)
                target.value = source.value if row == 1 or col == 1 else None
            else:
                target.value = None
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    copy_data_validations(source_ws, target_ws)


def insert_missing_template_columns(source_ws: Any, target_ws: Any) -> list[str]:
    inserted: list[str] = []
    for source_col, header in enumerate(header_values(source_ws), start=1):
        if not header or header_col(target_ws, header):
            continue
        target_ws.insert_cols(source_col)
        target_ws.column_dimensions[get_column_letter(source_col)].width = source_ws.column_dimensions[get_column_letter(source_col)].width
        for row in range(1, max(source_ws.max_row, target_ws.max_row) + 1):
            source = source_ws.cell(row, source_col)
            target = target_ws.cell(row, source_col)
            copy_cell_format(source, target)
            target.value = header if row == 1 else None
        inserted.append(header)
    return inserted


def sync_non_empty_sheet(source_ws: Any, target_ws: Any) -> bool:
    inserted_headers = insert_missing_template_columns(source_ws, target_ws)
    max_col = source_ws.max_column
    copy_dimensions(source_ws, target_ws, max_col, max(source_ws.max_row, target_ws.max_row))
    for col in range(1, max_col + 1):
        source = source_ws.cell(1, col)
        target = target_ws.cell(1, col)
        copy_cell_format(source, target)
        target.value = source.value
    for row in range(2, target_ws.max_row + 1):
        for col in range(1, max_col + 1):
            copy_cell_format(source_ws.cell(min(row, source_ws.max_row), col), target_ws.cell(row, col))
    copy_data_validations(source_ws, target_ws)
    return bool(inserted_headers)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="把模板工作表列结构、样式和下拉验证同步到其他大队工作表。")
    parser.add_argument("--workbook", required=True, type=Path, help="产品案卷数据.xlsx")
    parser.add_argument("--template", default="江阴大队", help="模板工作表名称")
    parser.add_argument("--targets", nargs="*", help="目标工作表；省略时同步除模板外所有工作表")
    parser.add_argument("--dry-run", action="store_true", help="只检查，不保存")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)
    wb = load_workbook(args.workbook)
    if args.template not in wb.sheetnames:
        raise ValueError(f"未找到模板工作表：{args.template}")
    source_ws = wb[args.template]
    if not header_col(source_ws, "标称生产者"):
        raise ValueError(f"{args.template} 缺少 标称生产者 表头")
    if not header_col(source_ws, "初查/复查"):
        raise ValueError(f"{args.template} 缺少 初查/复查 表头")
    ensure_phase_validation(source_ws)

    targets = args.targets or [name for name in wb.sheetnames if name != args.template and name.endswith("大队")]
    result: list[dict[str, Any]] = []
    for name in targets:
        if name not in wb.sheetnames:
            result.append({"sheet": name, "status": "missing"})
            continue
        target_ws = wb[name]
        empty = not sheet_has_data(target_ws)
        if empty:
            overwrite_empty_sheet(source_ws, target_ws)
            status = "overwritten_empty_template"
        else:
            inserted = sync_non_empty_sheet(source_ws, target_ws)
            status = "inserted_missing_columns" if inserted else "synced_existing"
        result.append({"sheet": name, "status": status})

    backup_path = None
    if not args.dry_run:
        backup_path = args.workbook.with_name(
            f"{args.workbook.stem}_模板同步备份_{datetime.now():%Y%m%d_%H%M%S}{args.workbook.suffix}"
        )
        shutil.copy2(args.workbook, backup_path)
        wb.save(args.workbook)

    print(json.dumps({"status": "ok", "dry_run": args.dry_run, "backup": str(backup_path) if backup_path else None, "sheets": result}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except PermissionError as exc:
        print(json.dumps({"status": "errors_found", "errors": [f"文件可能被 WPS/Excel 占用：{exc}"]}, ensure_ascii=False, indent=2))
        sys.exit(1)
    except Exception as exc:
        print(json.dumps({"status": "errors_found", "errors": [str(exc)]}, ensure_ascii=False, indent=2))
        sys.exit(1)
