import argparse
import hashlib
import json
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries

import monthly_grade_register as grade_register
import monthly_workflow as workflow

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


CONFIG = workflow.load_config()
DEFAULT_TEMPLATE_DIR = workflow.template_root(CONFIG)
DEFAULT_WORK_PLAN = workflow.data_source_path("work_plan", CONFIG)
DEFAULT_CASE_DATA = workflow.data_source_path("case_data", CONFIG)

PENDING_PREFIX = CONFIG["naming"]["pending_prefix"]
RED_RGB = "FFFF0000"
BLACK_RGB = "FF000000"
PENDING_STYLE = CONFIG["pending_rules"]["spreadsheet_pending_style"]
PENDING_FILL_RGB = PENDING_STYLE["fill_color"]
PENDING_FONT_RGB = PENDING_STYLE["font_color"]

ROOT_FILES = workflow.bulletin_root_map(CONFIG)

BRIGADE_ORDER = ["梁溪大队", "锡山大队", "惠山大队", "滨湖大队", "新吴大队", "江阴大队", "宜兴大队", "经开大队"]
BRIGADE_ALIASES = {"轨交大队": "轨道交通大队"}
MONTH_HEADERS = {
    1: "一月",
    2: "二月",
    3: "三月",
    4: "四月",
    5: "五月",
    6: "六月",
    7: "七月",
    8: "八月",
    9: "九月",
    10: "十月",
    11: "十一月",
    12: "十二月",
}


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def is_under(path, root):
    path = Path(path).resolve()
    root = Path(root).resolve()
    return path == root or root in path.parents


def add_blocker(blockers, message, **extra):
    item = {"message": message}
    item.update({key: str(value) for key, value in extra.items()})
    blockers.append(item)


def normalize_brigade(name):
    text = str(name or "").strip()
    text = re.sub(r"\d+人$", "", text)
    return BRIGADE_ALIASES.get(text, text)


def parse_staff_count(name):
    match = re.search(r"(\d+)人", str(name or ""))
    return int(match.group(1)) if match else None


def previous_month_window(year, month):
    start_year, start_month = (year - 1, 12) if month == 1 else (year, month - 1)
    return date(start_year, start_month, 26), date(year, month, 25)


def root_file_name(item, year, month, pending=False):
    name = item["target"].format(year=year, month=month)
    return f"{PENDING_PREFIX}{name}" if pending else name


def root_file_paths(bulletin_dir, year, month):
    bulletin_dir = Path(bulletin_dir)
    return {
        item["key"]: {
            "normal": bulletin_dir / root_file_name(item, year, month, pending=False),
            "pending": bulletin_dir / root_file_name(item, year, month, pending=True),
        }
        for item in ROOT_FILES
    }


def product_stats_pending_value(staff_count):
    template = CONFIG["pending_rules"]["product_stats_required_counts"]["value"]
    return template.format(staff_count=staff_count)


def openpyxl_pending_fill():
    return PatternFill(fill_type="solid", fgColor=PENDING_FILL_RGB)


def openpyxl_pending_font():
    return Font(color=PENDING_FONT_RGB)


def mark_openpyxl_pending(cell):
    cell.fill = openpyxl_pending_fill()
    cell.font = openpyxl_pending_font()


def excel_mark_pending(cell):
    cell.Interior.Color = 255
    cell.Font.Color = 0


def work_report_pending_cells():
    cells = []
    for cell_range in CONFIG["pending_rules"]["work_report_tech_and_product"]["pending_cells"]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cells.append(f"{get_column_letter(col)}{row}")
    return cells


def column_range_letters(column_range):
    start, end = column_range.split(":", 1)
    min_col, _, max_col, _ = range_boundaries(f"{start}1:{end}1")
    return [get_column_letter(col) for col in range(min_col, max_col + 1)]


def skeleton_path(template_dir, item):
    return Path(template_dir) / "X月通报" / item["skeleton"]


def numbered_template_path(template_dir, item):
    if not item.get("numbered"):
        return None
    return Path(template_dir) / "_编号模板库" / item["numbered"]


def read_staff_counts(work_plan):
    book = xlrd.open_workbook(str(work_plan), formatting_info=True)
    sheet = book.sheet_by_name("任务总表")
    counts = {}
    for row in range(1, sheet.nrows):
        brigade = normalize_brigade(sheet.cell_value(row, 0))
        staff = parse_staff_count(sheet.cell_value(row, 0))
        if brigade and staff:
            counts[brigade] = staff
    return counts


def _xlrd_rgb(book, index):
    return book.colour_map.get(index)


def _is_yellow(rgb):
    return rgb in {(255, 255, 0), (255, 255, 153)}


def _is_red(rgb):
    return rgb in {(255, 0, 0), (255, 51, 51), (192, 0, 0)}


def work_plan_header_applies(header, month):
    text = str(header or "").strip()
    if not text:
        return False
    if "常规" in text:
        return True
    start_match = re.search(r"(\d{1,2})月开始", text)
    if start_match:
        return int(start_match.group(1)) <= month
    for index, month_header in MONTH_HEADERS.items():
        if text == month_header:
            return index <= month
    return False


def parse_numbered_tasks(text):
    raw = str(text or "").strip()
    if not raw:
        return []
    tasks = []
    for line in raw.splitlines():
        line = line.strip()
        line = re.sub(r"^\d+[、.．]\s*", "", line)
        if line:
            tasks.append(line)
    return tasks


def extract_red_text_map_xlrd(book, sheet, candidates):
    result = {}
    run_map = getattr(sheet, "rich_text_runlist_map", {}) or {}
    for item in candidates:
        row = item["row"]
        col = item["col"]
        text = str(sheet.cell_value(row, col) or "")
        if not text:
            continue
        red_parts = []
        runs = run_map.get((row, col), [])
        if runs:
            spans = list(runs) + [(len(text), None)]
            for index in range(len(spans) - 1):
                start, font_index = spans[index]
                end = spans[index + 1][0]
                if font_index is None:
                    continue
                font = book.font_list[font_index]
                if _is_red(_xlrd_rgb(book, font.colour_index)):
                    red_parts.append(text[start:end])
        else:
            xf = book.xf_list[sheet.cell_xf_index(row, col)]
            font = book.font_list[xf.font_index]
            if _is_red(_xlrd_rgb(book, font.colour_index)):
                red_parts.append(text)
        red_text = "".join(red_parts).strip()
        if red_text:
            result[item["cell"]] = red_text
    return result


def extract_red_text_map(work_plan, sheet_name, cells, blockers):
    if not cells:
        return {}
    try:
        import win32com.client
    except Exception as exc:
        add_blocker(blockers, "无法读取工作计划富文本红字", error=exc)
        return {}

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    result = {}
    try:
        workbook = excel.Workbooks.Open(str(Path(work_plan).resolve()), ReadOnly=True)
        try:
            sheet = workbook.Worksheets(sheet_name)
            for cell_ref in cells:
                cell = sheet.Range(cell_ref)
                text = str(cell.Value or "")
                red_chars = []
                for index, char in enumerate(text, start=1):
                    try:
                        if cell.Characters(index, 1).Font.Color == 255:
                            red_chars.append(char)
                    except Exception:
                        add_blocker(blockers, "无法读取工作计划单元格富文本红字", cell=cell_ref)
                        red_chars = []
                        break
                red_text = "".join(red_chars).strip()
                if red_text:
                    result[cell_ref] = red_text
        finally:
            workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return result


def read_work_plan_month_tasks(work_plan, month, blockers=None):
    blockers = blockers if blockers is not None else []
    book = xlrd.open_workbook(str(work_plan), formatting_info=True)
    sheet = book.sheet_by_name("任务总表")
    applicable_cols = [col for col in range(1, sheet.ncols) if work_plan_header_applies(sheet.cell_value(0, col), month)]
    if not applicable_cols:
        add_blocker(blockers, "工作计划缺少可识别的累计任务列", month=month)
        return {}

    candidates = {}
    for row in range(1, sheet.nrows):
        brigade = normalize_brigade(sheet.cell_value(row, 0))
        if not brigade:
            continue
        candidates.setdefault(brigade, [])
        for month_col in applicable_cols:
            xf = book.xf_list[sheet.cell_xf_index(row, month_col)]
            rgb = _xlrd_rgb(book, xf.background.pattern_colour_index)
            if not _is_yellow(rgb):
                continue
            cell_ref = f"{xlrd.formula.colname(month_col)}{row + 1}"
            candidates[brigade].append(
                {
                    "row": row,
                    "col": month_col,
                    "cell": cell_ref,
                    "text": str(sheet.cell_value(row, month_col) or ""),
                }
            )

    all_candidates = [item for values in candidates.values() for item in values]
    red_map = extract_red_text_map_xlrd(book, sheet, all_candidates)
    tasks = {}
    for brigade, values in candidates.items():
        brigade_tasks = []
        for value in values:
            red_text = red_map.get(value["cell"], "")
            if red_text:
                brigade_tasks.extend(parse_numbered_tasks(red_text))
        tasks[brigade] = brigade_tasks
    return tasks


def read_case_counts(case_data, year, month):
    start, end = previous_month_window(year, month)
    year_start = date(year, 1, 1)
    workbook = load_workbook(case_data, read_only=True, data_only=True)
    mapping = CONFIG.get("case_data_mapping", {})
    fallback_indexes = {
        "date_column": 8,
        "project_column": 2,
        "micro_column": 10,
        "judge_type_column": 11,
        "qualified_column": 12,
        "case_type_column": 13,
        "review_column": 14,
    }
    result = {}
    for sheet in workbook.worksheets:
        brigade = normalize_brigade(sheet.title)
        if brigade not in BRIGADE_ORDER:
            continue
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header_map = {str(value or "").strip(): index for index, value in enumerate(headers)}
        indexes = {
            key: header_map.get(str(mapping.get(key, "")).strip(), fallback)
            for key, fallback in fallback_indexes.items()
        }
        rows = 0
        unique_projects = set()
        unqualified_projects = set()
        sample_rows = 0
        onsite_rows = 0
        micro_rows = 0
        unqualified_rows = 0
        administrative_case_year = 0
        criminal_case_year = 0
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(values):
                continue
            review_value = value_at(values, indexes["review_column"])
            if is_review_case(review_value):
                continue
            raw_date = value_at(values, indexes["date_column"])
            check_date = normalize_date(raw_date)
            if not check_date:
                continue
            case_type = str(value_at(values, indexes["case_type_column"]) or "")
            if year_start <= check_date <= end:
                if "行案" in case_type:
                    administrative_case_year += 1
                if "刑案" in case_type:
                    criminal_case_year += 1
            if start <= check_date <= end:
                rows += 1
                project_key = str(value_at(values, indexes["project_column"]) or "").strip() or f"{sheet.title}!{row_index}"
                unique_projects.add(project_key)
                if value_at(values, indexes["micro_column"]) and "微型" in str(value_at(values, indexes["micro_column"])):
                    micro_rows += 1
                judge_type = str(value_at(values, indexes["judge_type_column"]) or "")
                if "抽样" in judge_type:
                    sample_rows += 1
                if "现场" in judge_type:
                    onsite_rows += 1
                qualified = str(value_at(values, indexes["qualified_column"]) or "")
                if "不合格" in qualified:
                    unqualified_rows += 1
                    unqualified_projects.add(project_key)
        result[brigade] = {
            "rows": rows,
            "unique_projects": len(unique_projects),
            "unqualified_projects": len(unqualified_projects),
            "sample_rows": sample_rows,
            "onsite_rows": onsite_rows,
            "micro_rows": micro_rows,
            "unqualified_rows": unqualified_rows,
            "administrative_case_year": administrative_case_year,
            "criminal_case_year": criminal_case_year,
        }
    workbook.close()
    return result


def value_at(values, index):
    return values[index] if index is not None and len(values) > index else None


def is_review_case(value):
    text = str(value or "").strip().lower()
    if not text:
        return False
    return "复查" in text or text in {"是", "y", "yes", "true", "1"}


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", value)
        if match:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def parse_stat_count(value):
    match = re.search(r"(\d+)\s*(?:次|起|份|个)", str(value or ""))
    return int(match.group(1)) if match else None


def parse_first_number(value):
    match = re.search(r"(\d+)", str(value or ""))
    return int(match.group(1)) if match else None


def parse_required_count(value):
    match = re.search(r"[（(]\s*(\d+)\s*[）)]", str(value or ""))
    return int(match.group(1)) if match else None


def product_stats_count_for_column(letter, value):
    return parse_stat_count(value)


def read_product_stats(path):
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    stats = {}
    for row in range(sheet.nrows):
        brigade = normalize_brigade(sheet.cell_value(row, 0))
        if brigade in BRIGADE_ORDER:
            row_values = {}
            for col in range(1, min(sheet.ncols, 8)):
                letter = get_column_letter(col + 1)
                value = sheet.cell_value(row, col)
                row_values[letter] = {
                    "value": value,
                    "count": product_stats_count_for_column(letter, value),
                    "required": parse_required_count(value),
                }
            stats[brigade] = {
                "row": row + 1,
                "value": sheet.cell_value(row, 1),
                "count": product_stats_count_for_column("B", sheet.cell_value(row, 1)),
                "columns": row_values,
            }
    return stats


def product_timeliness_text(tasks, required_count=None, actual_count=None):
    issues = list(tasks or [])
    if required_count is not None and actual_count is not None and actual_count < required_count:
        issues.append(f"应完成{required_count}起案件实际完成{actual_count}起")
    if not issues:
        return "\\"
    lines = ["应完成但还未完成的："]
    lines.extend(f"{index}、{item}；" for index, item in enumerate(issues, start=1))
    lines[-1] = lines[-1].rstrip("；") + "。"
    return "\n".join(lines)


def timeliness_count_from_stats(stats, brigade, staff_counts):
    column_b = stats.get(brigade, {}).get("columns", {}).get("B", {})
    required = column_b.get("required") or staff_counts.get(brigade)
    actual = column_b.get("count")
    return required, actual


def audit_product_stats(stats, case_counts):
    blockers = []
    warnings = []
    mapping = CONFIG.get("case_data_mapping", {}).get("product_stats_audit", {})
    manual_columns = CONFIG.get("case_data_mapping", {}).get("manual_review_columns", {})
    for brigade in BRIGADE_ORDER:
        if brigade not in stats:
            add_blocker(blockers, "消防产品监督统计表缺少大队行", brigade=brigade)
            continue
        row_stats = stats[brigade]
        row_cases = case_counts.get(brigade, {})
        for column, case_key in mapping.items():
            actual = row_stats.get("columns", {}).get(column, {}).get("count")
            expected = row_cases.get(case_key)
            value = row_stats.get("columns", {}).get(column, {}).get("value")
            if actual is None:
                warnings.append(
                    {
                        "type": "product_stats_unfilled",
                        "message": "消防产品监督统计表未填写可核对数据",
                        "brigade": brigade,
                        "column": column,
                        "value": value,
                        "case_field": case_key,
                    }
                )
            elif expected is not None and actual != expected:
                warnings.append(
                    {
                        "type": "product_stats_mismatch",
                        "message": "消防产品监督统计表与案卷数据不一致",
                        "brigade": brigade,
                        "column": column,
                        "stat_count": actual,
                        "case_count": expected,
                        "case_field": case_key,
                    }
                )
        for column, reason in manual_columns.items():
            warnings.append({"message": "消防产品监督统计表列需人工核对", "brigade": brigade, "column": column, "reason": reason})
    return {"blockers": blockers, "warnings": warnings}


def pending_copy_roots(bulletin_dir, template_dir, year, month, actions, blockers, apply):
    bulletin_dir = Path(bulletin_dir)
    template_dir = Path(template_dir)
    bulletin_dir.mkdir(parents=True, exist_ok=True) if apply else None
    for item in ROOT_FILES:
        source = skeleton_path(template_dir, item)
        target = bulletin_dir / root_file_name(item, year, month, pending=True)
        normal = bulletin_dir / root_file_name(item, year, month, pending=False)
        if not source.exists():
            add_blocker(blockers, "通报骨架文件不存在", path=source)
            continue
        if target.exists():
            actions.append(
                {
                    "kind": "copy_pending_root_file",
                    "status": "skip_existing_pending_root_file",
                    "src": str(source),
                    "dst": str(target),
                    "sha256": sha256_file(source),
                }
            )
            continue
        actions.append(
            {
                "kind": "copy_pending_root_file",
                "status": "planned" if not apply else "done",
                "src": str(source),
                "dst": str(target),
                "sha256": sha256_file(source),
            }
        )
        if apply:
            shutil.copy2(source, target)
        if normal.exists():
            actions.append({"kind": "remove_unprefixed_root_file", "status": "planned" if not apply else "done", "path": str(normal)})
            if apply:
                normal.unlink()


def sync_work_report_skeleton(template_dir, actions, blockers, apply):
    item = next(entry for entry in ROOT_FILES if entry["key"] == "work_report")
    source = skeleton_path(template_dir, item)
    target = numbered_template_path(template_dir, item)
    if not source.exists():
        add_blocker(blockers, "重点工作骨架文件不存在，无法同步编号模板", path=source)
        return
    if target is None:
        return
    if target.exists() and sha256_file(source) == sha256_file(target):
        actions.append({"kind": "sync_work_report_numbered_template", "status": "skip_same_hash", "src": str(source), "dst": str(target)})
        return
    actions.append({"kind": "sync_work_report_numbered_template", "status": "planned" if not apply else "done", "src": str(source), "dst": str(target)})
    if apply:
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)


def find_score_sources(score_dir, month):
    score_dir = Path(score_dir)
    product_matches = sorted(score_dir.glob(f"*{month}月*产品巡查底册*.docx")) + sorted(score_dir.glob("*产品巡查底册*.docx"))
    product_matches = [path for path in product_matches if not path.name.startswith("~$")]
    network_matches = sorted(score_dir.glob("*联网监测基础信息考评明细表"))
    network_dir = network_matches[0] if network_matches else None
    network_stats = network_dir / "联网监测统计表.xls" if network_dir else None
    base_info = None
    root_base = sorted(score_dir.glob("*基础信息考评截图*（不发）*.xls")) + sorted(score_dir.glob("*基础信息考评截图*.xls"))
    root_base = [path for path in root_base if not path.name.startswith("~$")]
    if root_base:
        base_info = root_base[0]
    elif network_dir:
        nested = sorted(network_dir.glob("*基础信息考评截图*.xls"))
        base_info = nested[0] if nested else None
    return {
        "product_register": product_matches[0] if product_matches else None,
        "network_dir": network_dir,
        "network_stats": network_stats if network_stats and network_stats.exists() else None,
        "base_info": base_info,
    }


def read_score_office_content(score_dir, month, blockers):
    sources = find_score_sources(score_dir, month)
    for label, key in [
        ("产品巡查底册", "product_register"),
        ("联网监测统计表", "network_stats"),
        ("基础信息考评截图", "base_info"),
    ]:
        if not sources.get(key):
            add_blocker(blockers, f"成绩月份源文件不存在，无法填根层科室表{label}", path=score_dir)
    if blockers:
        return sources, [], []
    product_records = grade_register.parse_product_register(sources["product_register"])
    monitor_scores = grade_register.read_monitor_scores(sources["network_stats"])
    monitor_details = grade_register.read_monitor_details(sources["base_info"], monitor_scores)
    return sources, product_records, monitor_details


def row_by_label(sheet, label, label_column=3):
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, label_column).value or "").strip() == label:
            return row
    return None


def fill_office_record_score_rows(path, product_records, monitor_details):
    workbook = load_workbook(path)
    changed = []
    product_by_short = {item["short"]: item for item in product_records}
    for sheet in workbook.worksheets:
        product_row = row_by_label(sheet, "产品案卷核查")
        monitor_row = row_by_label(sheet, "联网系统核查")
        headers = {grade_register.brigade_short(sheet.cell(2, col).value): col for col in range(1, sheet.max_column + 1)}
        for short in grade_register.OFFICE_ORDER:
            col = headers.get(short)
            if not col:
                continue
            if product_row and short in product_by_short:
                cell = sheet.cell(product_row, col)
                cell.value = grade_register.product_office_text(product_by_short[short])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                changed.append({"sheet": sheet.title, "cell": cell.coordinate, "kind": "product_office_text"})
            if monitor_row:
                cell = sheet.cell(monitor_row, col)
                cell.value = grade_register.monitor_office_text(short, monitor_details)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                changed.append({"sheet": sheet.title, "cell": cell.coordinate, "kind": "monitor_office_text"})
    workbook.save(path)
    return changed


def build_timeliness_by_brigade(work_plan, year, month, product_stats, staff_counts, blockers):
    tasks_by_brigade = read_work_plan_month_tasks(work_plan, month, blockers)
    result = {}
    for brigade in BRIGADE_ORDER:
        required_count, actual_count = timeliness_count_from_stats(product_stats, brigade, staff_counts)
        result[brigade] = product_timeliness_text(
            tasks_by_brigade.get(brigade, []),
            required_count=required_count,
            actual_count=actual_count,
        )
    return result


def previous_bulletin_month(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def find_previous_work_report(bulletin_dir, year, month):
    prev_year, prev_month = previous_bulletin_month(year, month)
    if prev_year != year:
        return None
    prev_dir = Path(bulletin_dir).parent / f"{prev_month}月通报"
    item = next(entry for entry in ROOT_FILES if entry["key"] == "work_report")
    candidates = [
        prev_dir / root_file_name(item, prev_year, prev_month, pending=False),
        prev_dir / root_file_name(item, prev_year, prev_month, pending=True),
    ]
    return next((path for path in candidates if path.exists()), None)


def read_previous_tech_need_statuses(bulletin_dir, year, month):
    previous = find_previous_work_report(bulletin_dir, year, month)
    if not previous:
        return {}
    book = xlrd.open_workbook(str(previous))
    sheet = book.sheet_by_index(0)
    statuses = {}
    for row in range(sheet.nrows):
        brigade = normalize_brigade(sheet.cell_value(row, 1))
        if brigade in BRIGADE_ORDER:
            statuses[brigade] = str(sheet.cell_value(row, 10) or "").strip()
    return statuses


def tech_need_statuses(work_plan, bulletin_dir, year, month, blockers):
    inherited = read_previous_tech_need_statuses(bulletin_dir, year, month)
    current_tasks = read_work_plan_month_tasks(work_plan, month, blockers)
    statuses = {}
    for brigade in BRIGADE_ORDER:
        current_text = "\n".join(current_tasks.get(brigade, []))
        if "科技需求" in current_text or "上报一项科技需求" in current_text:
            statuses[brigade] = "未完成"
        elif inherited.get(brigade):
            statuses[brigade] = inherited[brigade]
        else:
            statuses[brigade] = "未到时间"
    return statuses


def work_report_product_values(case_counts):
    values = {}
    warnings = []
    for brigade in BRIGADE_ORDER:
        counts = case_counts.get(brigade, {})
        if not counts:
            continue
        values[brigade] = {
            "O": f"{counts.get('unique_projects', 0)}起（其中现场判定{counts.get('onsite_rows', 0)}起，抽样送检{counts.get('sample_rows', 0)}起，不合格{counts.get('unqualified_rows', 0)}起）",
            "P": f"{counts.get('administrative_case_year', 0)}起",
            "Q": f"{counts.get('criminal_case_year', 0)}起",
            "R": "0起",
        }
        warnings.append({"message": "召回不合格消防产品起数当前按 0 起预填，仍需人工最终确认。", "brigade": brigade, "column": "R"})
    return values, warnings


def mark_pending_office(path, timeliness_by_brigade=None, product_records=None, monitor_details=None):
    workbook = load_workbook(path)
    changed = []
    rule = CONFIG["pending_rules"]["office_product_timeliness"]
    for sheet in workbook.worksheets:
        header = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(rule["header_row"], col).value
            if value:
                header[normalize_brigade(value)] = col
        product_row = row_by_label(sheet, "产品案卷核查", rule["label_column"])
        monitor_row = row_by_label(sheet, "联网系统核查", rule["label_column"])
        product_by_short = {item["short"]: item for item in (product_records or [])}
        for short in grade_register.OFFICE_ORDER:
            col = header.get(f"{short}大队")
            if not col:
                continue
            if product_row and short in product_by_short:
                cell = sheet.cell(product_row, col)
                cell.value = grade_register.product_office_text(product_by_short[short])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                changed.append({"sheet": sheet.title, "cell": cell.coordinate, "kind": "product_office_text"})
            if monitor_row and monitor_details is not None:
                cell = sheet.cell(monitor_row, col)
                cell.value = grade_register.monitor_office_text(short, monitor_details)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                changed.append({"sheet": sheet.title, "cell": cell.coordinate, "kind": "monitor_office_text"})
        target_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, rule["label_column"]).value == rule["label"]:
                target_row = row
                break
        if target_row is None:
            continue
        for brigade in BRIGADE_ORDER:
            col = header.get(brigade)
            if not col:
                continue
            cell = sheet.cell(target_row, col)
            if timeliness_by_brigade is not None:
                cell.value = timeliness_by_brigade.get(brigade, "\\")
            mark_openpyxl_pending(cell)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            changed.append({"sheet": sheet.title, "cell": cell.coordinate, "value": cell.value})
    workbook.save(path)
    return changed


def excel_com():
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    return excel


def ensure_product_stats_required_value(value, staff_count):
    text = str(value or "").strip()
    required = product_stats_pending_value(staff_count)
    if parse_required_count(text):
        return text
    return f"{text}{required}" if text else required


def mark_pending_product_stats(path, staff_counts):
    excel = excel_com()
    changed = []
    rule = CONFIG["pending_rules"]["product_stats_required_counts"]
    pending_letters = column_range_letters(rule.get("pending_columns", "B:B"))
    try:
        workbook = excel.Workbooks.Open(str(Path(path).resolve()))
        try:
            sheet = workbook.Worksheets(1)
            used_rows = sheet.UsedRange.Rows.Count
            for row in range(1, used_rows + 1):
                brigade = normalize_brigade(sheet.Cells(row, 1).Value)
                if brigade not in staff_counts:
                    continue
                for letter in pending_letters:
                    col = coordinate_to_tuple(f"{letter}1")[1]
                    cell = sheet.Cells(row, col)
                    if letter == "B":
                        cell.Value = ensure_product_stats_required_value(cell.Value, staff_counts[brigade])
                    excel_mark_pending(cell)
                    changed.append({"sheet": sheet.Name, "cell": f"{letter}{row}", "value": str(cell.Value or "")})
            workbook.Save()
        finally:
            workbook.Close(SaveChanges=True)
    finally:
        excel.Quit()
    return changed


def mark_pending_work_report(path, tech_status_by_brigade=None, product_values_by_brigade=None):
    excel = excel_com()
    changed = []
    rule = CONFIG["pending_rules"]["work_report_tech_and_product"]
    try:
        workbook = excel.Workbooks.Open(str(Path(path).resolve()))
        try:
            sheet = workbook.Worksheets(1)
            min_col, min_row, max_col, max_row = range_boundaries(rule["completed_cells"])
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.Cells(row, col).Value = rule["completed_value"]
                    sheet.Cells(row, col).Font.Color = 0
                    sheet.Cells(row, col).Interior.Pattern = 0
            for cell_ref in work_report_pending_cells():
                row, col = coordinate_to_tuple(cell_ref)
                cell = sheet.Cells(row, col)
                brigade = normalize_brigade(sheet.Cells(row, 2).Value)
                letter = get_column_letter(col)
                if letter == rule.get("tech_need_column", "K") and tech_status_by_brigade is not None:
                    cell.Value = tech_status_by_brigade.get(brigade, cell.Value)
                elif product_values_by_brigade is not None:
                    value = product_values_by_brigade.get(brigade, {}).get(letter)
                    if value is not None:
                        cell.Value = value
                excel_mark_pending(cell)
                changed.append({"sheet": sheet.Name, "row": row, "col": col, "value": str(cell.Value)})
            workbook.Save()
        finally:
            workbook.Close(SaveChanges=True)
    finally:
        excel.Quit()
    return changed


def mark_pending_cells(
    bulletin_dir,
    year,
    month,
    staff_counts,
    work_plan,
    case_data,
    score_dir,
    score_month,
    actions,
    blockers,
    warnings,
    apply,
):
    paths = root_file_paths(bulletin_dir, year, month)
    timeliness_by_brigade = None
    tech_status_by_brigade = None
    product_values_by_brigade = None
    product_records = None
    monitor_details = None
    product_stats = {}
    if Path(case_data).exists():
        case_counts = read_case_counts(case_data, year, month)
        stats_path = paths["product_stats"]["pending"] if paths["product_stats"]["pending"].exists() else paths["product_stats"]["normal"]
        if stats_path.exists():
            product_stats = read_product_stats(stats_path)
            stats_audit = audit_product_stats(product_stats, case_counts)
            blockers.extend(stats_audit["blockers"])
            warnings.extend(stats_audit["warnings"])
        elif apply:
            add_blocker(blockers, "消防产品监督统计表不存在，无法先核对后生成产品工作实效", path=stats_path)
        timeliness_by_brigade = build_timeliness_by_brigade(work_plan, year, month, product_stats, staff_counts, blockers)
        product_values_by_brigade, product_warnings = work_report_product_values(case_counts)
        warnings.extend(product_warnings)
    else:
        add_blocker(blockers, "产品案卷数据不存在，无法预填产品工作实效和产品统计口径", path=case_data)
    tech_status_by_brigade = tech_need_statuses(work_plan, bulletin_dir, year, month, blockers)
    if score_dir:
        sources, product_records, monitor_details = read_score_office_content(score_dir, score_month or month, blockers)
        actions.append({"kind": "resolve_score_sources_for_root_office", "status": "done", "sources": {key: str(value) if value else None for key, value in sources.items()}})
    else:
        warnings.append({"message": "未提供 --score-dir，根层科室表不会预填产品案卷核查和联网系统核查。"})
    planned = [
        {
            "kind": "mark_pending_office_cells",
            "path": str(paths["office_record"]["pending"]),
            "cells": "产品案卷核查行、联网系统核查行、产品工作实效行 D:K",
        },
        {"kind": "mark_pending_product_stats_cells", "path": str(paths["product_stats"]["pending"]), "cells": "B:H 大队行红底，B列写要求数"},
        {"kind": "mark_pending_work_report_cells", "path": str(paths["work_report"]["pending"]), "cells": "K3:K10,O3:R10 红底并预填可得数据"},
    ]
    if not apply:
        actions.extend({**item, "status": "planned"} for item in planned)
        return
    if blockers:
        return
    for key in paths:
        if not paths[key]["pending"].exists():
            add_blocker(blockers, "待补根层文件不存在，无法标红", path=paths[key]["pending"])
            return
    actions.append(
        {
            "kind": "mark_pending_office_cells",
            "status": "done",
            "path": str(paths["office_record"]["pending"]),
            "changed": mark_pending_office(
                paths["office_record"]["pending"],
                timeliness_by_brigade=timeliness_by_brigade,
                product_records=product_records,
                monitor_details=monitor_details,
            ),
        }
    )
    actions.append({"kind": "mark_pending_product_stats_cells", "status": "done", "path": str(paths["product_stats"]["pending"]), "changed": mark_pending_product_stats(paths["product_stats"]["pending"], staff_counts)})
    actions.append(
        {
            "kind": "mark_pending_work_report_cells",
            "status": "done",
            "path": str(paths["work_report"]["pending"]),
            "changed": mark_pending_work_report(
                paths["work_report"]["pending"],
                tech_status_by_brigade=tech_status_by_brigade,
                product_values_by_brigade=product_values_by_brigade,
            ),
        }
    )


def remove_pending_prefixes(bulletin_dir, year, month, actions, blockers, apply):
    paths = root_file_paths(bulletin_dir, year, month)
    for key, item in paths.items():
        pending = item["pending"]
        normal = item["normal"]
        if not pending.exists():
            continue
        if normal.exists():
            add_blocker(blockers, "定稿文件已存在，无法去掉待补前缀", pending=pending, normal=normal)
            continue
        actions.append({"kind": "remove_pending_prefix", "status": "planned" if not apply else "done", "src": str(pending), "dst": str(normal)})
        if apply:
            pending.rename(normal)


def run(args):
    actions = []
    blockers = []
    warnings = []
    apply = args.apply
    bulletin_dir = Path(args.bulletin_dir)
    template_dir = Path(args.template_dir)
    work_plan = Path(args.work_plan)
    case_data = Path(args.case_data)
    score_dir = Path(args.score_dir) if getattr(args, "score_dir", None) else None
    score_month = getattr(args, "score_month", None)

    for required in [template_dir / "X月通报", work_plan]:
        if not required.exists():
            add_blocker(blockers, "必要路径不存在", path=required)

    staff_counts = read_staff_counts(work_plan) if work_plan.exists() else {}
    sync_work_report_skeleton(template_dir, actions, blockers, apply)

    if args.mode == "pending":
        pending_copy_roots(bulletin_dir, template_dir, args.year, args.month, actions, blockers, apply)
        if not blockers:
            mark_pending_cells(
                bulletin_dir,
                args.year,
                args.month,
                staff_counts,
                work_plan,
                case_data,
                score_dir,
                score_month,
                actions,
                blockers,
                warnings,
                apply,
            )
    else:
        paths = root_file_paths(bulletin_dir, args.year, args.month)
        stats_path = paths["product_stats"]["pending"] if paths["product_stats"]["pending"].exists() else paths["product_stats"]["normal"]
        if not stats_path.exists():
            add_blocker(blockers, "消防产品监督统计表不存在，无法核对", path=stats_path)
        if not case_data.exists():
            add_blocker(blockers, "产品案卷数据不存在，无法核对", path=case_data)
        if not blockers:
            audit = audit_product_stats(read_product_stats(stats_path), read_case_counts(case_data, args.year, args.month))
            blockers.extend(audit["blockers"])
            warnings.extend(audit["warnings"])
        if not blockers:
            remove_pending_prefixes(bulletin_dir, args.year, args.month, actions, blockers, apply)

    return {
        "mode": args.mode,
        "apply": apply,
        "bulletin_dir": str(bulletin_dir),
        "year": args.year,
        "month": args.month,
        "case_window": [str(item) for item in previous_month_window(args.year, args.month)],
        "staff_counts": staff_counts,
        "actions": actions,
        "warnings": warnings,
        "blockers": blockers,
        "ok": not blockers,
    }


def main():
    parser = argparse.ArgumentParser(
        description="处理通报月份根层三张当月表的待补和25号核对。",
        epilog="参考：references/monthly/output_R01_office_record.md、output_R02_product_stats.md、output_R03_work_report.md。",
    )
    parser.add_argument("--bulletin-dir", required=True)
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--work-plan", default=str(DEFAULT_WORK_PLAN))
    parser.add_argument("--case-data", default=str(DEFAULT_CASE_DATA))
    parser.add_argument("--template-dir", default=str(DEFAULT_TEMPLATE_DIR))
    parser.add_argument("--score-dir", help="成绩月份巡查目录，用于预填根层科室月考核情况记录表的产品/联网核查行。")
    parser.add_argument("--score-year", type=int, help="成绩月份年份；默认与通报年份一致。")
    parser.add_argument("--score-month", type=int, help="成绩月份；默认与通报月份一致。")
    parser.add_argument("--mode", choices=["pending", "final-audit"], required=True)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    if args.score_month is None:
        args.score_month = args.month
    if args.score_year is None:
        args.score_year = args.year
    result = run(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if result["blockers"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
