import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import monthly_bulletin_root as root


class MonthlyBulletinRootTests(unittest.TestCase):
    def test_previous_month_window_uses_26_to_25(self):
        self.assertEqual(root.previous_month_window(2026, 6), (date(2026, 5, 26), date(2026, 6, 25)))
        self.assertEqual(root.previous_month_window(2026, 1), (date(2025, 12, 26), date(2026, 1, 25)))

    def test_pending_root_names_use_prefix(self):
        paths = root.root_file_paths(Path(r"C:\work\6月通报"), 2026, 6)
        self.assertEqual(set(paths), {"office_record", "product_stats"})
        self.assertEqual(paths["office_record"]["pending"].name, "【待补】2026年6月科室月考核情况记录表.xlsx")
        self.assertEqual(paths["product_stats"]["pending"].name, "【待补】2026年6月消防产品监督统计表.xls")

    def test_work_report_root_file_is_retired(self):
        root_ids = {item["id"] for item in root.CONFIG["bulletin_root_files"]}
        template_ids = {item["id"] for item in root.CONFIG["templates"]}
        self.assertIn("R03", root_ids)
        self.assertNotIn("T10", template_ids)

    def test_staff_count_and_pending_value(self):
        self.assertEqual(root.normalize_brigade("江阴大队12人"), "江阴大队")
        self.assertEqual(root.parse_staff_count("江阴大队12人"), 12)
        self.assertEqual(root.parse_staff_count("经开大队6人"), 6)
        self.assertEqual(root.product_stats_pending_value(8), "（8）")

    def test_timeliness_text_matches_may_style(self):
        text = root.product_timeliness_text(["6月底前完成1起消防产品行政处罚案件"], required_count=12, actual_count=7)
        self.assertEqual(
            text,
            "应完成但还未完成的：\n1、6月底前完成1起消防产品行政处罚案件；\n2、应完成12起案件实际完成7起。",
        )
        self.assertEqual(root.product_timeliness_text([]), "\\")

    def test_ensure_product_stats_required_value_preserves_reported_count(self):
        self.assertEqual(root.ensure_product_stats_required_value("", 8), "（8）")
        self.assertEqual(root.ensure_product_stats_required_value("7次", 8), "7次（8）")
        self.assertEqual(root.ensure_product_stats_required_value("7次（8）", 8), "7次（8）")

    def test_read_work_plan_month_tasks_uses_cumulative_yellow_red_text_only(self):
        class FakeSheet:
            nrows = 3
            ncols = 5
            rich_text_runlist_map = {(1, 2): [(0, 1)]}

            def __init__(self):
                self.values = {
                    (0, 0): "",
                    (0, 1): "一月",
                    (0, 2): "二月",
                    (0, 3): "三月",
                    (1, 0): "梁溪大队8人",
                    (1, 1): "1、黑字已完成",
                    (1, 2): "1、红字未完成",
                    (1, 3): "1、普通格不登记",
                    (2, 0): "经开大队6人",
                    (2, 1): "1、黄色无红字",
                }

            def cell_value(self, row, col):
                return self.values.get((row, col), "")

            def cell_xf_index(self, row, col):
                return 1 if (row, col) in {(1, 1), (1, 2), (2, 1)} else 0

            def sheet_by_name(self, name):
                return self

        fake_book = SimpleNamespace(
            sheet_by_name=lambda name: FakeSheet(),
            xf_list=[
                SimpleNamespace(background=SimpleNamespace(pattern_colour_index=0), font_index=0),
                SimpleNamespace(background=SimpleNamespace(pattern_colour_index=1), font_index=0),
            ],
            font_list=[
                SimpleNamespace(colour_index=0),
                SimpleNamespace(colour_index=2),
            ],
            colour_map={0: (255, 255, 255), 1: (255, 255, 0), 2: (255, 0, 0)},
        )
        with patch.object(root.xlrd, "open_workbook", return_value=fake_book):
            blockers = []
            tasks = root.read_work_plan_month_tasks(Path("fake.xls"), 3, blockers)

        self.assertEqual(blockers, [])
        self.assertEqual(tasks["梁溪大队"], ["红字未完成"])
        self.assertEqual(tasks["经开大队"], [])

    def test_read_case_counts_excludes_review_and_counts_unqualified_projects(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cases.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "梁溪大队"
            headers = [
                "序号",
                "单位名称",
                "项目编号",
                "初查/复查",
                "单位地址",
                "检查人员",
                "立案人",
                "检查人",
                "检查时间（以产品检查记录表登记时间为准）",
                "检查产品",
                "是否为微型消防站或快速处置队",
                "现场判定或抽样送检",
                "是否合格",
                "是否为行案/刑案",
                "是否为复查",
            ]
            ws.append(headers)
            ws.append([1, "A", "P1", "", "", "", "", "", date(2026, 6, 1), "", "", "现场", "不合格", "行案", ""])
            ws.append([2, "A复查", "P1", "", "", "", "", "", date(2026, 6, 10), "", "", "现场", "不合格", "行案", "复查"])
            ws.append([3, "B", "P2", "", "", "", "", "", date(2026, 6, 12), "", "微型消防站", "抽样", "合格", "", ""])
            wb.save(path)

            counts = root.read_case_counts(path, 2026, 6)

        self.assertEqual(counts["梁溪大队"]["unique_projects"], 2)
        self.assertEqual(counts["梁溪大队"]["unqualified_projects"], 1)
        self.assertEqual(counts["梁溪大队"]["unqualified_rows"], 1)
        self.assertEqual(counts["梁溪大队"]["sample_rows"], 1)
        self.assertEqual(counts["梁溪大队"]["administrative_case_year"], 1)

    def test_audit_product_stats_warns_mapped_column_mismatch(self):
        stats = {}
        cases = {}
        for brigade in root.BRIGADE_ORDER:
            stats[brigade] = {
                "row": 1,
                "columns": {
                    "B": {"value": "1次（8）", "count": 1},
                    "C": {"value": "0份", "count": 0},
                    "D": {"value": "0份", "count": 0},
                    "E": {"value": "0份", "count": 0},
                    "F": {"value": "0次", "count": 0},
                    "G": {"value": "0份", "count": 0},
                },
            }
            cases[brigade] = {
                "unique_projects": 1,
                "unqualified_projects": 0,
                "unqualified_rows": 0,
                "sample_rows": 0,
                "micro_rows": 0,
                "administrative_case_year": 0,
            }
        stats["江阴大队"]["columns"]["B"] = {"value": "6次（12）", "count": 6}
        cases["江阴大队"]["unique_projects"] = 11
        result = root.audit_product_stats(stats, cases)
        warnings = result["warnings"]
        self.assertEqual(result["blockers"], [])
        self.assertTrue(result["warnings"])
        self.assertTrue(any(item.get("type") == "product_stats_mismatch" and item.get("brigade") == "江阴大队" for item in warnings))

    def test_timeliness_uses_product_stats_required_and_actual_count(self):
        stats = {
            "梁溪大队": {"columns": {"B": {"value": "7次（8）", "count": 7, "required": 8}}},
            "经开大队": {"columns": {"B": {"value": "（6）", "count": None, "required": 6}}},
        }
        required, actual = root.timeliness_count_from_stats(stats, "梁溪大队", {"梁溪大队": 9})
        self.assertEqual((required, actual), (8, 7))
        required, actual = root.timeliness_count_from_stats(stats, "经开大队", {"经开大队": 6})
        self.assertEqual((required, actual), (6, None))

    def test_mark_pending_office_marks_product_timeliness_row_red(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "office.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "1"
            ws.cell(2, 4).value = "梁溪大队"
            ws.cell(2, 11).value = "经开大队"
            ws.cell(15, 3).value = "产品工作实效"
            wb.save(path)

            changed = root.mark_pending_office(
                path,
                timeliness_by_brigade={
                    "梁溪大队": "应完成但还未完成的：\n1、应完成8起案件实际完成7起。",
                    "经开大队": "\\",
                },
            )
            wb2 = load_workbook(path)
            ws2 = wb2["1"]

            self.assertIn("应完成8起案件实际完成7起", ws2.cell(15, 4).value)
            self.assertEqual(ws2.cell(15, 11).value, "\\")
            self.assertEqual(ws2.cell(15, 4).fill.fgColor.rgb, "FFFF0000")
            self.assertEqual(ws2.cell(15, 4).font.color.rgb, "FF000000")
            self.assertEqual(len(changed), 2)


if __name__ == "__main__":
    unittest.main()
