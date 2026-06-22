import argparse
import json
import sys
from pathlib import Path

import xlrd
from openpyxl import load_workbook

import monthly_grade_register as grade_register
import monthly_workflow as workflow
import template_resolver


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def first_existing(paths):
    for path in paths:
        path = Path(path)
        if path.exists():
            return path
    return None


def find_product_register(score_dir, year, month):
    score_dir = Path(score_dir)
    candidates = sorted(score_dir.glob(f"*{month}月*产品巡查底册*.docx")) + sorted(score_dir.glob("*产品巡查底册*.docx"))
    return candidates[0] if candidates else None


def find_network_dir(score_dir, year, month):
    score_dir = Path(score_dir)
    candidates = sorted(score_dir.glob("*联网监测基础信息考评明细表"))
    return candidates[0] if candidates else None


def find_base_info(score_dir, network_dir, year, month):
    score_dir = Path(score_dir)
    candidates = [
        score_dir / f"{year}年{month}月基础信息考评截图（不发）.xls",
        score_dir / f"{month}月基础信息考评截图（不发）.xls",
        score_dir / f"【待补】{month}月基础信息考评截图（不发）.xls",
    ]
    if network_dir:
        candidates.append(Path(network_dir) / f"{month}月基础信息考评截图.xls")
    found = first_existing(candidates)
    if found:
        return found
    nested = sorted(score_dir.glob("**/*基础信息考评截图*.xls"))
    return nested[0] if nested else None


def build_grade_outputs(score_dir, year, month, config):
    score_dir = Path(score_dir)
    outputs = {}
    for item in workflow.grade_outputs(config):
        if item.get("requires_flag"):
            continue
        outputs[item["key"]] = str(score_dir / item["target"].format(year=year, month=month))
    return outputs


def existing_status(path):
    path = Path(path)
    return {"path": str(path), "exists": path.exists(), "is_dir": path.is_dir() if path.exists() else False}


def scan_pending_text(path, marker="【待补】"):
    path = Path(path)
    if not path.exists() or path.is_dir():
        return []
    hits = []
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, data_only=False, read_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if marker in str(cell.value or ""):
                        hits.append({"path": str(path), "sheet": sheet.title, "cell": cell.coordinate, "value": str(cell.value)})
    elif path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(str(path))
        for sheet in book.sheets():
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)
                    if marker in str(value or ""):
                        hits.append({"path": str(path), "sheet": sheet.name, "row": row + 1, "column": col + 1, "value": str(value)})
    return hits


def adopted_template_path(template_result, key):
    for status in template_result["statuses"]:
        if status["key"] == key and status.get("adopted_path"):
            return Path(status["adopted_path"])
    return None


def run(args):
    config = workflow.load_config()
    bulletin_dir = Path(args.bulletin_dir)
    score_dir = bulletin_dir / workflow.score_dir_name(args.score_month, config)
    template_result = template_resolver.resolve_templates(
        template_dir=args.template_dir,
        include_reserved=True,
        allow_snapshot_fallback=True,
        config=config,
    )

    product_register = find_product_register(score_dir, args.score_year, args.score_month)
    network_dir = find_network_dir(score_dir, args.score_year, args.score_month)
    network_stats = Path(network_dir) / "联网监测统计表.xls" if network_dir else None
    base_info = find_base_info(score_dir, network_dir, args.score_year, args.score_month)
    work_plan = Path(args.work_plan) if args.work_plan else workflow.data_source_path("work_plan", config)
    case_data = Path(args.case_data) if args.case_data else workflow.data_source_path("case_data", config)

    root_files = {}
    pending_text_hits = []
    for item in workflow.bulletin_root_map(config):
        normal = bulletin_dir / workflow.root_file_name(item, args.bulletin_year, args.bulletin_month, pending=False, config=config)
        pending = bulletin_dir / workflow.root_file_name(item, args.bulletin_year, args.bulletin_month, pending=True, config=config)
        root_files[item["key"]] = {
            "id": item["id"],
            "normal": existing_status(normal),
            "pending": existing_status(pending),
        }
        for path in [normal, pending]:
            if path.exists():
                pending_text_hits.extend(scan_pending_text(path))

    grade_outputs = {
        key: existing_status(path)
        for key, path in build_grade_outputs(score_dir, args.score_year, args.score_month, config).items()
    }

    blockers = []
    warnings = list(template_result["warnings"])
    warnings.extend(
        {
            "message": "外部模板与 skill 快照不一致，生成会采用外部模板。",
            "template_id": status["id"],
            "file": status["file"],
            "external_path": status["external_path"],
            "snapshot_path": status["snapshot_path"],
        }
        for status in template_result["statuses"]
        if status["external_exists"] and status["snapshot_exists"] and not status["match"]
    )
    for label, path in [
        ("通报月份目录", bulletin_dir),
        ("成绩月份巡查目录", score_dir),
        ("产品巡查底册", product_register),
        ("联网监测明细目录", network_dir),
        ("联网监测统计表", network_stats),
        ("基础信息考评截图", base_info),
        ("工作计划", work_plan),
        ("产品案卷数据", case_data),
    ]:
        if not path or not Path(path).exists():
            blockers.append({"message": f"{label}不存在", "path": str(path) if path else ""})

    for hit in pending_text_hits:
        warnings.append({"message": "根层表仍包含单元格文本【待补】，应改为红底标记并保留/预填业务值。", **hit})

    product_detail_leaks = []
    person_match_issues = []
    if product_register and Path(product_register).exists():
        try:
            product_records = grade_register.parse_product_register(product_register)
            product_detail_leaks = grade_register.product_detail_leak_issues_for_records(product_records)
            if network_stats and Path(network_stats).exists() and base_info and Path(base_info).exists():
                personal_template = adopted_template_path(template_result, "personal_stats")
                if personal_template and personal_template.exists():
                    monitor_scores = grade_register.read_monitor_scores(network_stats)
                    monitor_details = grade_register.read_monitor_details(base_info, monitor_scores)
                    person_match_issues = grade_register.collect_person_match_issues(personal_template, product_records, monitor_details)
        except Exception as exc:
            warnings.append({"message": "审计产品/个人统计风险时失败", "error": str(exc)})
    warnings.extend(product_detail_leaks)
    warnings.extend(person_match_issues)

    all_blockers = blockers + template_result["blockers"]
    return {
        "ok": not all_blockers,
        "bulletin": {
            "dir": str(bulletin_dir),
            "year": args.bulletin_year,
            "month": args.bulletin_month,
            "root_files": root_files,
            "pending_text_hits": pending_text_hits,
        },
        "score": {
            "dir": str(score_dir),
            "year": args.score_year,
            "month": args.score_month,
            "input_files": {
                "product_register": str(product_register) if product_register else None,
                "network_dir": str(network_dir) if network_dir else None,
                "network_stats": str(network_stats) if network_stats else None,
                "base_info": str(base_info) if base_info else None,
            },
            "outputs": grade_outputs,
        },
        "data_sources": {
            "work_plan": existing_status(work_plan),
            "case_data": existing_status(case_data),
        },
        "templates": {
            "policy": "external_template_source_is_authoritative; skill_snapshot_is_for_hash_verification",
            "statuses": template_result["statuses"],
        },
        "warnings": warnings,
        "blockers": all_blockers,
    }


def main():
    parser = argparse.ArgumentParser(
        description="只读审计月度通报目录、数据源、模板版本和目标成品状态。",
        epilog="参考：references/monthly/00_workflow_router.md；审计规则见 references/monthly/validation_and_audit.md。",
    )
    parser.add_argument("--bulletin-dir", required=True)
    parser.add_argument("--bulletin-year", type=int, required=True)
    parser.add_argument("--bulletin-month", type=int, required=True)
    parser.add_argument("--score-year", type=int, required=True)
    parser.add_argument("--score-month", type=int, required=True)
    parser.add_argument("--template-dir", default=str(workflow.template_root()))
    parser.add_argument("--work-plan")
    parser.add_argument("--case-data")
    args = parser.parse_args()
    result = run(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if result["blockers"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
