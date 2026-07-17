import argparse
import copy
import hashlib
import json
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

from docx import Document
from docx.enum.text import WD_COLOR_INDEX
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
ROOT_DIR = SKILL_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import writer
import monthly_workflow
import template_resolver

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


PRODUCT_ORDER = ["江阴", "宜兴", "梁溪", "锡山", "惠山", "滨湖", "新吴", "经开"]
OFFICE_ORDER = ["梁溪", "锡山", "惠山", "滨湖", "新吴", "江阴", "宜兴", "经开"]
CASE_ORDER = ["江阴", "宜兴", "梁溪", "锡山", "惠山", "滨湖", "新吴", "经开"]
PERSON_SCORE_COLUMNS = list(range(28, 36))  # AB:AI
REPORT_TITLE_FONT = "方正黑体_GBK"
REPORT_BODY_FONT = "方正楷体_GBK"
MONTHLY_TEMPLATE_MANIFEST = SKILL_DIR / "resources" / "monthly_templates" / "manifest.json"
MONITOR_HISTORY_PATH = SKILL_DIR / "resources" / "history" / "monitor_report_history.json"
REVIEW_RULES_JS = SKILL_DIR / "resources" / "review_rules" / "消防产品专项监督抽查卷评查规则.js"
EXTERNAL_MONTHLY_TEMPLATE_KEYS = {
    "personal_stats",
    "case_scores",
    "monthly_report",
}
BUILTIN_MONTHLY_TEMPLATE_KEYS = {
    "product_archive_detail",
    "product_summary",
}
OPTIONAL_BUILTIN_TEMPLATE_KEYS = {
    "office_record",
}
MONTHLY_TEMPLATE_KEYS = EXTERNAL_MONTHLY_TEMPLATE_KEYS | BUILTIN_MONTHLY_TEMPLATE_KEYS | OPTIONAL_BUILTIN_TEMPLATE_KEYS
ALLOWED_DEDUCTION_VALUES = {0.1, 0.2, 0.5, 1.0}
DESCRIPTION_WARN_LENGTH = 30
NO_UNQUALIFIED_PRODUCT_CASE_TEXT = "本月未完成不合格消防产品案卷"
INVALID_MONITOR_CONTACTS = {"消防机构", "机构联系人", "联系人"}
YEAR_MONTH_PATTERN = re.compile(r"(\d{4})年(\d{1,2})月")
MONTH_ONLY_PATTERN = re.compile(r"(\d{1,2})月")
WORKFLOW_CONFIG = monthly_workflow.load_config()
PERSONAL_RULES = WORKFLOW_CONFIG.get("personal_stats_rules", {})
PENDING_FILL_RGB = PERSONAL_RULES.get("mismatch_fill_color", "FFFF0000")
PENDING_FONT_RGB = PERSONAL_RULES.get("mismatch_font_color", "FF000000")
PRODUCT_SCORE_NUMBER_FORMAT = "0.0"
MONITOR_SCORE_NUMBER_FORMAT = "0.00"
CASE_SCORE_MONITOR_NUMBER_FORMAT = "0.0"


class HumanReviewRequired(RuntimeError):
    def __init__(self, issues):
        self.issues = issues
        super().__init__("人员信息需要人工核对")


class ProductScoreReviewRequired(RuntimeError):
    def __init__(self, guard):
        self.guard = guard
        super().__init__("产品底册评分需要人工核对")


class HistoryMonthConflict(RuntimeError):
    def __init__(self, issues):
        self.issues = issues
        super().__init__("历史台账月份存在冲突")


def norm_text(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def brigade_short(value):
    return norm_text(value).replace("大队", "").replace("轨道交通", "轨交")


def score_text(value):
    if value is None or value == "":
        return ""
    num = float(value)
    return str(round(num, 2)).rstrip("0").rstrip(".")


def write_openpyxl_score(cell, value, number_format):
    cell.value = value
    cell.number_format = number_format


def write_com_score(cell, value, number_format):
    cell.Value = value
    cell.NumberFormat = number_format


def paragraph_has_yellow_highlight(paragraph):
    return any(run.font.highlight_color == WD_COLOR_INDEX.YELLOW for run in paragraph.runs)


def require_path(path, label):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"{label}不存在：{path}")
    return path


def find_one(base, patterns, label):
    matches = []
    for pattern in patterns:
        matches.extend(Path(base).glob(pattern))
    matches = [p for p in matches if not p.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"未找到{label}：{patterns}")
    if len(matches) > 1:
        matches = sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0]


def find_optional_one(base, patterns):
    matches = []
    for pattern in patterns:
        matches.extend(Path(base).glob(pattern))
    matches = [p for p in matches if not p.name.startswith("~$")]
    if not matches:
        return None
    return sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def find_base_info(month_dir, network_dir):
    root_match = find_optional_one(
        month_dir,
        [
            "*基础信息考评截图*（不发）*.xls",
            "*基础信息考评截图*.xls",
        ],
    )
    if root_match:
        return root_match
    return find_one(network_dir, ["*基础信息考评截图*.xls"], "基础信息考评截图")


def copy_output(src, dst, force=False):
    dst = Path(dst)
    if dst.exists() and not force:
        raise FileExistsError(f"目标文件已存在，使用 --force 覆盖生成文件：{dst}")
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst


def sha256_text(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def utc_now_text():
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def make_month_key(year, month):
    return f"{int(year):04d}-{int(month):02d}"


def previous_month(year, month):
    year = int(year)
    month = int(month)
    if month == 1:
        return year - 1, 12
    return year, month - 1


def build_score_month_info(registration_year, registration_month, source):
    return {
        "registration_year": int(registration_year),
        "registration_month": int(registration_month),
        "registration_month_key": make_month_key(registration_year, registration_month),
        "score_year": int(registration_year),
        "score_month": int(registration_month),
        "score_month_key": make_month_key(registration_year, registration_month),
        "source": source,
    }


def extract_registration_month_from_text(text, default_year=None):
    text = str(text or "")
    match = YEAR_MONTH_PATTERN.search(text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = MONTH_ONLY_PATTERN.search(text)
    if match and default_year is not None:
        return int(default_year), int(match.group(1))
    return None


def resolve_registration_month(path_like=None, default_year=None, fallback_date=None, allow_runtime_fallback=True):
    if path_like:
        path = Path(path_like)
        resolved = extract_registration_month_from_text(path.name, default_year=default_year)
        if resolved:
            return {
                "registration_year": resolved[0],
                "registration_month": resolved[1],
                "source": "filename",
            }
        for part in reversed(path.parts[:-1]):
            resolved = extract_registration_month_from_text(part, default_year=default_year)
            if resolved:
                return {
                    "registration_year": resolved[0],
                    "registration_month": resolved[1],
                    "source": "folder",
                }
    if not allow_runtime_fallback:
        raise ValueError(f"无法从路径识别登记月份：{path_like}")
    fallback_date = fallback_date or date.today()
    return {
        "registration_year": int(fallback_date.year),
        "registration_month": int(fallback_date.month),
        "source": "runtime",
    }


def resolve_score_month(path_like=None, default_year=None, fallback_date=None, allow_runtime_fallback=True):
    registration = resolve_registration_month(
        path_like,
        default_year=default_year,
        fallback_date=fallback_date,
        allow_runtime_fallback=allow_runtime_fallback,
    )
    return build_score_month_info(
        registration["registration_year"],
        registration["registration_month"],
        registration["source"],
    )


def load_builtin_monthly_templates(required_keys):
    templates = {}
    by_key = monthly_workflow.internal_templates_by_key(WORKFLOW_CONFIG)
    for key in required_keys:
        if key not in by_key:
            raise FileNotFoundError(f"monthly_workflow.json 缺少内置模板配置：{key}")
        path = monthly_workflow.internal_template_path(by_key[key], WORKFLOW_CONFIG)
        require_path(path, f"内置月度模板 {key}")
        templates[key] = path
    return templates


def load_monthly_template_manifest(manifest_path=MONTHLY_TEMPLATE_MANIFEST, include_score_office_record=False):
    manifest_path = Path(manifest_path)
    if not manifest_path.exists():
        raise FileNotFoundError(f"月度模板 manifest 不存在：{manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    base_dir = manifest_path.parent
    templates = {}
    required_external = set(EXTERNAL_MONTHLY_TEMPLATE_KEYS)
    for item in manifest.get("templates", []):
        key = item.get("key")
        if key not in required_external:
            continue
        if item.get("reserved") or not item.get("used_in_monthly_register"):
            continue
        path = base_dir / item["file"]
        require_path(path, f"月度模板 {key}")
        templates[key] = path
    missing = sorted(required_external - set(templates))
    if missing:
        raise FileNotFoundError("月度模板 manifest 缺少必要模板：" + "、".join(missing))
    required_builtin = set(BUILTIN_MONTHLY_TEMPLATE_KEYS)
    if include_score_office_record:
        required_builtin |= OPTIONAL_BUILTIN_TEMPLATE_KEYS
    templates.update(load_builtin_monthly_templates(required_builtin))
    return templates, str(manifest_path)


def load_external_templates(template_dir):
    return load_monthly_templates(
        template_dir=template_dir,
        allow_snapshot_fallback=False,
        include_score_office_record=True,
    )


def load_monthly_templates(template_dir=None, allow_snapshot_fallback=False, include_score_office_record=False):
    resolved = template_resolver.resolve_templates(
        template_dir=template_dir,
        include_reserved=False,
        allow_snapshot_fallback=allow_snapshot_fallback,
    )
    if resolved["blockers"]:
        messages = [
            f"{item['template_id']} {item['file']}: {item['message']} ({item['external_path']})"
            for item in resolved["blockers"]
        ]
        raise FileNotFoundError("月度模板外部事实源不可用：" + "；".join(messages))
    templates = {key: path for key, path in resolved["templates"].items() if key in EXTERNAL_MONTHLY_TEMPLATE_KEYS}
    missing = sorted(EXTERNAL_MONTHLY_TEMPLATE_KEYS - set(templates))
    if missing:
        raise FileNotFoundError("月度模板缺少必要模板：" + "、".join(missing))
    required_builtin = set(BUILTIN_MONTHLY_TEMPLATE_KEYS)
    if include_score_office_record:
        required_builtin |= OPTIONAL_BUILTIN_TEMPLATE_KEYS
    templates.update(load_builtin_monthly_templates(required_builtin))
    return templates, {
        "mode": "external_plus_builtin",
        "template_dir": str(Path(template_dir)) if template_dir else None,
        "policy": "external_templates_follow_optimized_x_month_tree; builtin_generation_assets_stay_in_skill_snapshot",
        "warnings": resolved["warnings"],
        "statuses": resolved["statuses"],
    }


def parse_deduction_value(line):
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", line)]
    small = [x for x in nums if x <= 1.0]
    return small[-1] if small else None


def parse_general_index(line):
    match = re.search(r"3\s*\.\s*(\d+)", line)
    return int(match.group(1)) if match else None


def has_detail_marker(text):
    text = str(text or "")
    return any(marker in text for marker in ("（", "(", "ps：", "PS：", "ps:", "PS:"))


def strip_detail_text(text):
    text = re.sub(r"（[^）]*）", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.split(r"\bps\s*[:：]", text, maxsplit=1, flags=re.IGNORECASE)[0]
    return text.strip(" ；;，,、")


def public_product_description(text):
    raw = clean_error_line(text)
    public = clean_error_line(strip_detail_text(raw))
    return public or raw


def match_product_doc_labels(compact):
    labels = []
    doc_markers = [
        ("责令限期改正通知书", "责令限期改正通知书"),
        ("限期改正通知书", "责令限期改正通知书"),
        ("限改", "责令限期改正通知书"),
        ("现场检查判定不合格通知书", "消防产品现场检查判定不合格通知书"),
        ("现场检查判定不合格", "消防产品现场检查判定不合格通知书"),
        ("检验结果通知书", "消防产品质量检验结果通知书"),
        ("质量监督抽查抽样单", "消防产品质量监督抽查抽样单"),
        ("抽样单", "消防产品质量监督抽查抽样单"),
        ("监督检查记录", "消防产品监督检查记录"),
        ("通报函", "通报函"),
    ]
    seen = set()
    for marker, label in doc_markers:
        if marker in compact and label not in seen:
            labels.append(label)
            seen.add(label)
    return labels


def normalize_broad_description(text):
    return public_product_description(text)


def classify_product_description(text):
    raw = clean_error_line(text)
    broad = strip_detail_text(raw)
    compact = norm_text(raw)
    doc_labels = match_product_doc_labels(compact)

    if any(keyword in compact for keyword in ("审批表", "呈请", "办案部门意见")):
        if len(doc_labels) == 1:
            return f"{doc_labels[0]}审批材料缺失"
        return "审批材料缺失"

    if "消防产品监督检查记录" in compact or "监督检查记录" in compact:
        if any(keyword in compact for keyword in ("不规范", "缺失", "备注", "空白", "市场准入", "产品所在部位")):
            return "消防产品监督检查记录填写不规范"
    if "现场检查判定不合格通知书" in compact or "消防现场检查不合格通知书" in compact:
        return "消防产品现场检查判定不合格通知书填写不规范"
    if "质量监督抽查抽样单" in compact or "抽样单" in compact:
        return "消防产品质量监督抽查抽样单填写不规范"
    if ("责令限期改正通知书" in compact or "限期改正通知书" in compact or "限改" in compact) and any(
        keyword in compact for keyword in ("具体问题", "规格型号", "生产企业", "编号", "填写不完整", "填写不规范", "记载不规范", "违法行为")
    ):
        return "责令限期改正通知书填写不规范"
    if any(keyword in compact for keyword in ("授权委托书", "法人证明", "营业执照", "型试检验报告", "型式检验报告")):
        return "附件材料缺失"
    if "照片" in compact:
        return "照片证据不完整"
    if "送达" in compact:
        return "送达材料填写不规范"
    if any(keyword in compact for keyword in ("上传", "系统", "附件错传", "材料上传错误")):
        return "系统上传材料不规范"
    return broad or raw


def unique_join(parts):
    result = []
    seen = set()
    for part in parts:
        text = str(part or "").strip(" ；;，,")
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return "；".join(result)


def unique_text_list(parts):
    result = []
    seen = set()
    for part in parts:
        text = str(part or "").strip(" ；;，,")
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def clean_error_line(line):
    text = str(line or "").strip()
    text = re.sub(r"^\d+\s*[、.．]\s*", "", text)
    return text.strip(" ；;，,")


def is_blank_product_template_item(description, deduction_line=""):
    description_text = norm_text(description).replace("(", "（").replace(")", "）")
    deduction_text = norm_text(deduction_line)
    return description_text in {"（）", ""} and not parse_deduction_value(deduction_text) and not parse_general_index(deduction_text)


def monitor_contact_warnings(monitor_details):
    warnings = []
    for detail in monitor_details:
        if detail.get("联系人"):
            continue
        note = str(detail.get("note") or "")
        if "联系人" not in note:
            continue
        warnings.append(
            {
                "type": "monitor_contact_label",
                "message": "联网备注包含联系人类标签但未提取到人名，已作为问题描述处理，不写入个人执法统计表。",
                "大队": detail.get("大队"),
                "单位": detail.get("单位"),
                "path": detail.get("source_path"),
                "sheet": detail.get("sheet"),
                "row": detail.get("row"),
                "column": detail.get("note_column"),
                "note": note,
            }
        )
    return warnings


def parse_product_register(path):
    document = Document(str(path))
    texts = [
        {
            "paragraph": index,
            "text": p.text.strip(),
            "yellow": paragraph_has_yellow_highlight(p),
        }
        for index, p in enumerate(document.paragraphs, 1)
    ]
    blocks = []
    current = None
    for item in texts:
        paragraph = item["paragraph"]
        text = item["text"]
        if not text:
            continue
        if text.startswith("大队："):
            if current:
                blocks.append(current)
            current = {"大队": text.split("：", 1)[1].strip(), "lines": []}
        elif current:
            current["lines"].append(item)
    if current:
        blocks.append(current)

    records = []
    for block in blocks:
        record = {
            "大队": block["大队"],
            "short": brigade_short(block["大队"]),
            "题名": "",
            "编号": "",
            "立卷人": "",
            "检查人": "",
            "errors": [],
            "archive_errors": [],
            "deductions": [],
            "ignored_yellow_errors": [],
            "parse_issues": [],
            "score": None,
            "no_case": False,
        }
        pending_error = None
        for entry in block["lines"]:
            line = entry["text"]
            paragraph = entry["paragraph"]
            is_yellow = entry.get("yellow", False)
            if line.startswith("题名："):
                record["题名"] = line.split("：", 1)[1].strip()
            elif line.startswith("编号："):
                record["编号"] = line.split("：", 1)[1].strip()
            elif line.startswith("立卷人："):
                record["立卷人"] = line.split("：", 1)[1].strip()
            elif line.startswith("检查人："):
                record["检查人"] = line.split("：", 1)[1].strip()
            elif line == "错误" or re.fullmatch(r"\d+\s*[、.．]?", line):
                continue
            elif line.startswith("扣分"):
                value = parse_deduction_value(line)
                general_index = parse_general_index(line)
                if pending_error:
                    pending_text = pending_error["text"]
                    if pending_error.get("yellow") or is_yellow:
                        if not is_blank_product_template_item(pending_text, line):
                            record["ignored_yellow_errors"].append(
                                {
                                    "description": pending_text,
                                    "line": line,
                                    "paragraph": pending_error["paragraph"],
                                    "deduction_paragraph": paragraph,
                                }
                            )
                        pending_error = None
                        continue
                    if is_blank_product_template_item(pending_text, line):
                        pending_error = None
                        continue
                    broad_description = normalize_broad_description(pending_text)
                    record["errors"].append(pending_text)
                    record["archive_errors"].append(broad_description)
                    if value is not None and general_index is not None:
                        record["deductions"].append(
                            {
                                "general_index": general_index,
                                "value": value,
                                "line": line,
                                "paragraph": pending_error["paragraph"],
                                "deduction_paragraph": paragraph,
                                "description": pending_text,
                                "detail_description": pending_text,
                                "broad_description": broad_description,
                            }
                        )
                    else:
                        record["parse_issues"].append(
                            {
                                "description": pending_text,
                                "paragraph": pending_error["paragraph"],
                                "line": line,
                                "deduction_paragraph": paragraph,
                                "message": "扣分行缺少可解析的 3.x 条款或分值",
                            }
                        )
                pending_error = None
            else:
                cleaned = clean_error_line(line)
                if cleaned:
                    pending_error = {"text": cleaned, "paragraph": paragraph, "yellow": is_yellow}
        if pending_error:
            if not is_blank_product_template_item(pending_error["text"]):
                if pending_error.get("yellow"):
                    record["ignored_yellow_errors"].append(
                        {
                            "description": pending_error["text"],
                            "line": "",
                            "paragraph": pending_error["paragraph"],
                            "deduction_paragraph": None,
                        }
                    )
                else:
                    record["errors"].append(pending_error["text"])
                    record["archive_errors"].append(normalize_broad_description(pending_error["text"]))

        if not record["题名"] and not record["编号"] and not record["立卷人"]:
            record["no_case"] = True
            record["score"] = 0.0
        else:
            total = sum(item["value"] for item in record["deductions"])
            record["score"] = round(10.0 - total, 1)
        records.append(record)

    by_short = {item["short"]: item for item in records}
    return [by_short[key] for key in PRODUCT_ORDER if key in by_short]


def load_product_template_record():
    template_path = SKILL_DIR / "resources" / "template.json"
    data = json.loads(template_path.read_text(encoding="utf-8"))
    return data[0]


def general_key(record, index):
    prefix = f"{index}_"
    for key in record["一般要素"].keys():
        if key.startswith(prefix):
            return key
    raise KeyError(f"template.json 中找不到一般要素 {index}")


def build_product_batch(product_records, year, month):
    template = load_product_template_record()
    batch = []
    for item in product_records:
        record = copy.deepcopy(template)
        record["fields"].update(
            {
                "被检查单位": item["大队"],
                "评查日期": f"{year}年{month}月",
                "题名": item["题名"],
                "编号": item["编号"],
                "立卷人": item["立卷人"],
                "检查人": item["检查人"],
            }
        )
        if item["no_case"]:
            record["no_case"] = True
            record["score_override"] = 0
        else:
            for deduction in item["deductions"]:
                key = general_key(record, deduction["general_index"])
                current_score = record["一般要素"][key].get("扣分情况", "")
                current_desc = record["一般要素"][key].get("扣分说明", "")
                new_score = deduction["value"] + (float(current_score) if current_score else 0.0)
                descriptions = [current_desc, deduction.get("broad_description") or deduction["description"]]
                record["一般要素"][key]["扣分情况"] = score_text(new_score)
                record["一般要素"][key]["扣分说明"] = unique_join(descriptions)
        batch.append(record)
    return batch


def load_review_rule_limits(path=REVIEW_RULES_JS):
    path = require_path(path, "消防产品专项监督抽查卷评查规则 JS")
    text = path.read_text(encoding="utf-8-sig")
    match = re.search(r"const\s+\w+\s*=\s*(\{.*?\})\s*;\s*module\.exports", text, flags=re.S)
    if not match:
        raise ValueError(f"无法解析评查规则 JS：{path}")
    data = json.loads(match.group(1))
    limits = {}

    def visit(node):
        if isinstance(node, dict):
            code = str(node.get("code", ""))
            if re.fullmatch(r"3\.\d+", code) and node.get("maxScore") is not None:
                limits[int(code.split(".", 1)[1])] = float(node["maxScore"])
            for child in node.get("children", []):
                visit(child)
        elif isinstance(node, list):
            for item in node:
                visit(item)

    visit(data.get("sections", []))
    missing = [idx for idx in range(1, 16) if idx not in limits]
    if missing:
        raise ValueError("评查规则 JS 缺少一般要素上限：" + "、".join(f"3.{idx}" for idx in missing))
    return limits


def validate_product_score_guard(product_records):
    limits = load_review_rule_limits()
    blocking = []
    warnings = []
    for record in product_records:
        if record["no_case"]:
            continue
        short = record["short"]
        totals = {}
        for issue in record.get("parse_issues", []):
            blocking.append(
                {
                    "大队": record["大队"],
                    "题名": record["题名"],
                    "type": "parse_error",
                    "message": issue["message"],
                    "description": issue["description"],
                    "paragraph": issue.get("paragraph"),
                    "line": issue["line"],
                    "deduction_paragraph": issue.get("deduction_paragraph"),
                }
            )
        for deduction in record["deductions"]:
            index = deduction["general_index"]
            value = float(deduction["value"])
            totals[index] = totals.get(index, 0.0) + value
            if index not in limits:
                blocking.append(
                    {
                        "大队": record["大队"],
                        "题名": record["题名"],
                        "type": "unknown_general_index",
                        "message": f"扣分条款 3.{index} 不存在，无法落格",
                        "description": deduction["description"],
                    }
                )
            if round(value, 2) not in ALLOWED_DEDUCTION_VALUES:
                warnings.append(
                    {
                        "大队": record["大队"],
                        "题名": record["题名"],
                        "type": "non_standard_single_value",
                        "message": f"单项扣分 {score_text(value)} 不是 0.1/0.2/0.5/1，仅提示人工复核",
                        "description": deduction["description"],
                    }
                )
            if not has_detail_marker(deduction["detail_description"]):
                warnings.append(
                    {
                        "大队": record["大队"],
                        "题名": record["题名"],
                        "type": "missing_detail_marker",
                        "message": "错误描述未使用括号或 ps 标注具体问题",
                        "description": deduction["description"],
                    }
                )
            broad = deduction.get("broad_description", "")
            if len(broad) > DESCRIPTION_WARN_LENGTH or "，" in broad or "、" in broad:
                warnings.append(
                    {
                        "大队": record["大队"],
                        "题名": record["题名"],
                        "type": "broad_description_too_specific",
                        "message": "宽泛描述偏长或包含字段堆叠，仅提示人工复核",
                        "broad_description": broad,
                        "description": deduction["description"],
                    }
                )
        for index, total in totals.items():
            if index in limits and total > limits[index] + 1e-9:
                blocking.append(
                    {
                        "大队": record["大队"],
                        "题名": record["题名"],
                        "type": "general_limit_exceeded",
                        "message": f"3.{index} 扣分合计 {score_text(total)} 超过原规则上限 {score_text(limits[index])}",
                    }
                )
        general_total = sum(float(item["value"]) for item in record["deductions"])
        if general_total > 5 + 1e-9:
            blocking.append(
                {
                    "大队": record["大队"],
                    "题名": record["题名"],
                    "type": "general_total_exceeded",
                    "message": f"一般要素扣分合计 {score_text(general_total)} 超过 5 分上限",
                }
            )
    return {"blocking_issues": blocking, "warnings": warnings}


def find_soffice():
    found = shutil.which("soffice")
    if found:
        return found
    fallback = Path(r"D:\Program_Files\LibreOffice\program\soffice.com")
    if fallback.exists():
        return str(fallback)
    raise FileNotFoundError("未找到 LibreOffice soffice，用于读取 .xls")


def load_xls_as_workbook(path, data_only=True):
    soffice = find_soffice()
    temp_dir = tempfile.TemporaryDirectory(prefix="xf_grade_xls_")
    out_dir = Path(temp_dir.name)
    convert_dir = out_dir / "converted"
    profile_dir = out_dir / "lo-profile"
    convert_dir.mkdir()
    profile_dir.mkdir()
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(convert_dir),
                str(path),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        converted = next(convert_dir.glob("*.xlsx"))
    except Exception as soffice_error:
        try:
            converted = convert_xls_with_excel(path, convert_dir)
        except Exception as excel_error:
            raise RuntimeError(
                f"无法转换 .xls：{path}；LibreOffice 错误：{soffice_error}；Excel 兜底错误：{excel_error}"
            ) from excel_error
    workbook = load_workbook(converted, data_only=data_only)
    workbook._xf_temp_dir = temp_dir
    return workbook


def convert_xls_with_excel(path, out_dir):
    import win32com.client

    path = Path(path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    converted = out_dir / f"{path.stem}.xlsx"
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        try:
            workbook.SaveAs(str(converted.resolve()), FileFormat=51)
        finally:
            workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return converted


def read_monitor_scores(stats_path):
    workbook = load_xls_as_workbook(stats_path, data_only=True)
    ws = workbook.active
    scores = {}
    warnings = []
    for row in range(3, ws.max_row + 1):
        short = brigade_short(ws.cell(row, 1).value)
        if short not in PRODUCT_ORDER:
            continue
        row_scores = [float(ws.cell(row, col).value) for col in range(2, 12)]
        raw_avg = ws.cell(row, 12).value
        computed_avg = sum(row_scores) / len(row_scores)
        if raw_avg is None:
            avg = computed_avg
        else:
            avg = float(raw_avg)
            if avg > 10:
                warnings.append(
                    {
                        "type": "monitor_avg_gt_10",
                        "path": str(Path(stats_path)),
                        "sheet": ws.title,
                        "row": row,
                        "column": 12,
                        "cell": f"L{row}",
                        "brigade": short,
                        "raw_avg": raw_avg,
                        "computed_avg": round(computed_avg, 2),
                        "message": "联网监测均分列大于 10，脚本按 10 个案卷分数重算均分，未修改源表。",
                    }
                )
                avg = computed_avg
        scores[short] = {"scores": row_scores, "avg": round(float(avg), 2)}
    read_monitor_scores.last_warnings = warnings
    return scores


read_monitor_scores.last_warnings = []


def normalize_monitor_issues(raw):
    text = str(raw or "").strip(" ；;，,、")
    parts = [p.strip() for p in re.split(r"[、,，]+", text) if p.strip()]
    normalized = []
    for part in parts:
        low = part.lower()
        if part in {"人员信息", "未上传人员信息"}:
            normalized.append("缺人员信息")
        elif re.fullmatch(r"缺?\s*pdf(?:图)?", low) or part in {"火灾防控图", "缺火灾防控图"}:
            normalized.append("缺火灾防控图")
        elif re.fullmatch(r"缺?\s*cad(?:图)?", low) or part in {"点位图", "缺点位图"}:
            normalized.append("缺点位图")
        elif part == "主机日期":
            normalized.append("消控主机生产日期未录入")
        elif part in {"未登录", "账号长期登录"}:
            normalized.append("账号长期未登录")
        elif part == "脱岗":
            normalized.append("消控室脱岗")
        elif part in {"消控室电话", "消控电话不是固话", "消控非固话", "没有消控电话"}:
            normalized.append("消控室电话不规范")
        elif part == "监督编码":
            normalized.append("监督编码缺失")
        elif part == "设备离线":
            normalized.append("设备离线")
        elif "维保合同维保记录" in part:
            normalized.append(part.replace("维保合同维保记录", "缺维保合同、缺维保记录"))
        elif part == "维保记录":
            normalized.append("缺维保记录")
        elif part == "维保合同":
            normalized.append("缺维保合同")
        elif part == "缺维保合同":
            normalized.append("缺维保合同")
        else:
            normalized.append(part.replace("维保合同维保记录", "缺维保合同、缺维保记录"))
    return "、".join(normalized)


def parse_monitor_note(note):
    note = str(note or "").strip()
    match = re.search(r"\d{3,4}\s*[，,、]?\s*([\u4e00-\u9fff]{2,4})", note)
    person = match.group(1) if match else ""
    if person in INVALID_MONITOR_CONTACTS or "机构" in person or "联系人" in person:
        person = ""
    if match:
        issues = note[match.end() :]
    else:
        issues = note
    issues = issues.strip(" /\\，,、")
    return person, issues


def read_monitor_details(base_info_path, monitor_scores):
    workbook = load_xls_as_workbook(base_info_path, data_only=False)
    details = []
    for ws in workbook.worksheets:
        short = brigade_short(ws.title)
        if short not in monitor_scores:
            continue
        scores = monitor_scores[short]["scores"]
        for row in range(1, min(ws.max_row, len(scores)) + 1):
            unit = ws.cell(row, 1).value
            non_empty = [
                (col, ws.cell(row, col).value)
                for col in range(1, ws.max_column + 1)
                if ws.cell(row, col).value not in (None, "")
            ]
            note_col, note_value = non_empty[-1] if non_empty else (None, "")
            note = str(note_value).strip() if non_empty else ""
            person, issues = parse_monitor_note(note)
            details.append(
                {
                    "short": short,
                    "大队": f"{short}大队",
                    "index": row,
                    "单位": str(unit or "").strip(),
                    "联系人": person,
                    "score": float(scores[row - 1]),
                    "issues_raw": issues,
                    "issues_report": normalize_monitor_issues(issues),
                    "note": note,
                    "sheet": ws.title,
                    "row": row,
                    "note_column": note_col,
                    "source_path": str(Path(base_info_path)),
                }
            )
    return details


def collect_person_match_issues(template_path, product_records, monitor_details):
    workbook = load_workbook(template_path)
    ws = workbook.active
    person_index = build_person_index(ws)
    issues = []

    for item in product_records:
        if item["no_case"] or not item["立卷人"]:
            continue
        key = (item["short"], norm_text(item["立卷人"]))
        if key not in person_index:
            issues.append(
                {
                    "type": "product",
                    "大队": item["大队"],
                    "姓名": item["立卷人"],
                    "案卷": item["题名"],
                    "message": f"个人统计表未找到产品立卷人：{item['大队']} {item['立卷人']}",
                }
            )

    for detail in monitor_details:
        name = norm_text(detail["联系人"])
        if not name:
            continue
        key = (detail["short"], name)
        if key not in person_index:
            issues.append(
                {
                    "type": "monitor",
                    "大队": detail["大队"],
                    "姓名": detail["联系人"],
                    "单位": detail["单位"],
                    "path": detail.get("source_path"),
                    "sheet": detail.get("sheet"),
                    "row": detail.get("row"),
                    "column": detail.get("note_column"),
                    "note": detail.get("note"),
                    "message": f"个人统计表未找到联网联系人：{detail['大队']} {detail['联系人']}",
                }
            )

    return issues


def validate_person_matches(template_path, product_records, monitor_details):
    issues = collect_person_match_issues(template_path, product_records, monitor_details)
    if issues:
        raise HumanReviewRequired(issues)
    return True


def write_product_summary(template_path, output_path, product_records, force):
    copy_output(template_path, output_path, force=force)
    workbook = load_workbook(output_path)
    ws = workbook.active
    scores = {item["short"]: item["score"] for item in product_records}
    for row in range(2, ws.max_row + 1):
        short = brigade_short(ws.cell(row, 1).value)
        if short in scores:
            write_openpyxl_score(ws.cell(row, 2), scores[short], PRODUCT_SCORE_NUMBER_FORMAT)
    workbook.save(output_path)


def build_person_index(ws):
    index = {}
    current_brigade = ""
    for row in range(6, ws.max_row + 1):
        if ws.cell(row, 1).value not in (None, ""):
            current_brigade = brigade_short(ws.cell(row, 1).value)
        name = norm_text(ws.cell(row, 2).value)
        if current_brigade and name:
            index[(current_brigade, name)] = row
    return index


def build_brigade_rows(ws):
    rows = {}
    for row in range(6, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value not in (None, ""):
            short = brigade_short(value)
            if short:
                rows[short] = row
    return rows


def mark_personal_product_mismatches(ws, issues):
    product_issues = [item for item in issues if item.get("type") == "product"]
    if not product_issues:
        return []
    brigade_rows = build_brigade_rows(ws)
    changed = []
    by_short = {}
    for issue in product_issues:
        by_short.setdefault(brigade_short(issue.get("大队")), []).append(issue)
    fill = PatternFill(fill_type="solid", fgColor=PENDING_FILL_RGB)
    font = Font(color=PENDING_FONT_RGB)
    for short, items in by_short.items():
        row = brigade_rows.get(short)
        if not row:
            continue
        cell = ws.cell(row, 27)
        names = unique_text_list(item.get("姓名") for item in items)
        names_text = "、".join(names)
        cell.value = f"{PERSONAL_RULES.get('mismatch_marker', '待核对')}：{names_text}"
        cell.fill = fill
        cell.font = font
        comment_lines = [
            PERSONAL_RULES.get("mismatch_comment_prefix", "产品底册立卷人未在个人执法统计表中找到"),
            *[f"{item.get('大队')} {item.get('姓名')}：{item.get('案卷')}" for item in items],
        ]
        cell.comment = Comment("\n".join(comment_lines), "xf-report-filler")
        changed.append({"大队": f"{short}大队", "cell": cell.coordinate, "names": names})
    return changed


def write_personal_stats(template_path, output_path, product_records, monitor_details, month, force):
    match_issues = collect_person_match_issues(template_path, product_records, monitor_details)
    copy_output(template_path, output_path, force=force)
    workbook = load_workbook(output_path)
    ws = workbook.active
    for col in range(27, 36):
        ws.cell(5, col).value = f"{month}月"
    for row in range(6, ws.max_row + 1):
        for col in range(27, 36):
            ws.cell(row, col).value = None

    person_index = build_person_index(ws)
    warnings = []
    for item in product_records:
        if item["no_case"] or not item["立卷人"]:
            continue
        key = (item["short"], norm_text(item["立卷人"]))
        row = person_index.get(key)
        if row:
            write_openpyxl_score(ws.cell(row, 27), item["score"], PRODUCT_SCORE_NUMBER_FORMAT)
        else:
            warnings.append(f"个人统计表未找到产品立卷人：{item['大队']} {item['立卷人']}")
    mismatch_marks = mark_personal_product_mismatches(ws, match_issues)
    for item in mismatch_marks:
        warnings.append(f"已在个人执法统计表标记待核对：{item['大队']} {item['cell']} {'、'.join(item['names'])}")

    next_col_by_person = {}
    for detail in monitor_details:
        name = norm_text(detail["联系人"])
        if not name:
            warnings.append(f"联网备注未提取到联系人，未写入个人执法统计表：{detail['大队']} {detail['单位']}")
            continue
        key = (detail["short"], name)
        row = person_index.get(key)
        if not row:
            warnings.append(f"个人统计表未找到联网联系人：{detail['大队']} {detail['联系人']}")
            continue
        col = next_col_by_person.get(key, PERSON_SCORE_COLUMNS[0])
        if col > PERSON_SCORE_COLUMNS[-1]:
            warnings.append(f"联网联系人超过 AB:AI 可填列：{detail['大队']} {detail['联系人']}")
            continue
        write_openpyxl_score(ws.cell(row, col), detail["score"], MONITOR_SCORE_NUMBER_FORMAT)
        next_col_by_person[key] = col + 1
    workbook.save(output_path)
    return warnings


def product_office_text(item):
    if item["no_case"]:
        return f"{item['大队']}{NO_UNQUALIFIED_PRODUCT_CASE_TEXT}。"
    broad_issues = unique_text_list(item.get("archive_errors", []))
    if not broad_issues:
        broad_issues = ["案卷材料规范性问题"]
    lines = [f"{item['题名']}产品监督检查中存在以下问题"]
    for idx, issue in enumerate(broad_issues, start=1):
        lines.append(f"{idx}、{issue}")
    return "\n".join(lines)


def product_report_sentence(item):
    if item["no_case"]:
        return f"{item['大队']}{NO_UNQUALIFIED_PRODUCT_CASE_TEXT}"
    broad_issues = unique_text_list(item.get("archive_errors", []))
    if not broad_issues:
        broad_issues = ["案卷材料规范性问题"]
    issues_text = "、".join(broad_issues)
    suffix = "" if len(broad_issues) == 1 else "等问题"
    return f"{item['大队']}{item['立卷人']}承办的{item['题名']}存在{issues_text}{suffix}"


def product_detail_fragments(text):
    fragments = []
    for match in re.finditer(r"[（(]([^（）()]{2,})[）)]", str(text or "")):
        fragment = match.group(1).strip()
        if fragment:
            fragments.append(fragment)
    ps_part = re.split(r"\bps\s*[:：]", str(text or ""), maxsplit=1, flags=re.IGNORECASE)
    if len(ps_part) > 1 and ps_part[1].strip():
        fragments.append(ps_part[1].strip())
    return fragments


def product_detail_leak_issues_for_records(product_records):
    issues = []
    for item in product_records:
        output_texts = {
            "office_record_text": product_office_text(item),
            "report_sentence": product_report_sentence(item),
            "archive_errors": "\n".join(item.get("archive_errors", [])),
        }
        for raw_error in item.get("errors", []):
            expected_raw_public = public_product_description(raw_error)
            for fragment in product_detail_fragments(raw_error):
                if fragment == expected_raw_public:
                    continue
                for field, output_text in output_texts.items():
                    if fragment and fragment in output_text:
                        issues.append(
                            {
                                "type": "product_detail_leak",
                                "大队": item.get("大队"),
                                "案卷": item.get("题名"),
                                "field": field,
                                "fragment": fragment,
                                "message": "产品底册括号内或 ps 细节进入了对外输出文本",
                            }
                        )
        expected_public = unique_text_list(public_product_description(raw_error) for raw_error in item.get("errors", []))
        actual_public = unique_text_list(item.get("archive_errors", []))
        if expected_public != actual_public:
            issues.append(
                {
                    "type": "product_public_description_mismatch",
                    "大队": item.get("大队"),
                    "案卷": item.get("题名"),
                    "expected": expected_public,
                    "actual": actual_public,
                    "message": "产品公开描述必须严格使用底册括号前文本，不允许用括号内细节做语义归类",
                }
            )
    return issues


def monitor_office_text(short, details):
    rows = [d for d in details if d["short"] == short]
    lines = []
    for idx, detail in enumerate(rows, start=1):
        issues = normalize_monitor_issues(detail["issues_raw"])
        lines.append(f"{idx}、{detail['单位']}{issues}")
    return "\n".join(lines)


def write_office_record(template_path, output_path, product_records, monitor_details, force):
    copy_output(template_path, output_path, force=force)
    workbook = load_workbook(output_path)
    ws = workbook.active
    product_by_short = {item["short"]: item for item in product_records}
    headers = {brigade_short(ws.cell(2, col).value): col for col in range(4, 15)}
    for short in OFFICE_ORDER:
        col = headers.get(short)
        if not col:
            continue
        if short in product_by_short:
            ws.cell(14, col).value = product_office_text(product_by_short[short])
            ws.cell(14, col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(16, col).value = monitor_office_text(short, monitor_details)
        ws.cell(16, col).alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(output_path)


def write_case_scores(template_path, output_path, product_records, monitor_scores, force):
    copy_output(template_path, output_path, force=force)
    import win32com.client

    product_scores = {item["short"]: item["score"] for item in product_records}
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(str(Path(output_path).resolve()))
    try:
        ws = workbook.Worksheets(1)
        for row in range(3, 11):
            short = brigade_short(ws.Cells(row, 1).Value)
            if short in product_scores:
                write_com_score(ws.Cells(row, 2), product_scores[short], PRODUCT_SCORE_NUMBER_FORMAT)
            if short in monitor_scores:
                write_com_score(ws.Cells(row, 11), monitor_scores[short]["avg"], CASE_SCORE_MONITOR_NUMBER_FORMAT)
        workbook.Save()
    finally:
        workbook.Close(False)
        excel.Quit()


def decode_legacy_text(raw):
    for encoding in ("gb18030", "utf-8", "utf-16"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("gb18030", errors="replace")


def extract_monitor_report_section(report_path):
    out_dir = Path(tempfile.mkdtemp(prefix="xf_report_history_"))
    try:
        subprocess.run(
            [find_soffice(), "--headless", "--convert-to", "txt:Text", "--outdir", str(out_dir), str(report_path)],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        txt_files = list(out_dir.glob("*.txt"))
        if not txt_files:
            return ""
        text = decode_legacy_text(txt_files[0].read_bytes())
        start = text.find("（五）联网监测")
        end = text.find("三、下一步工作要求")
        if start < 0:
            return ""
        return text[start:end if end > start else len(text)]
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def parse_report_month(report_path, default_year=None, fallback_date=None):
    info = resolve_score_month(report_path, default_year=default_year, fallback_date=fallback_date)
    return info["score_year"], info["score_month"]


def extract_monitor_brigades(section):
    selected = []
    used = set()
    for match in re.finditer(r"([\u4e00-\u9fff]{2})大队", section):
        short = match.group(1)
        if short not in CASE_ORDER or short in used:
            continue
        selected.append({"short": short, "大队": f"{short}大队"})
        used.add(short)
    return selected


def migrate_history_record_month(record):
    migrated = copy.deepcopy(record)
    info = resolve_score_month(
        record.get("report_file"),
        default_year=record.get("year"),
        allow_runtime_fallback=False,
    )
    migrated["year"] = info["score_year"]
    migrated["month"] = info["score_month"]
    migrated["month_key"] = info["score_month_key"]
    return migrated, info


def migrate_monitor_history_records(history):
    version = history.get("version", 1)
    migrated_records = []
    actions = []
    seen = {}
    for index, record in enumerate(history.get("records", [])):
        migrated, info = migrate_history_record_month(record)
        month_key = migrated["month_key"]
        if month_key in seen:
            raise HistoryMonthConflict(
                [
                    {
                        "month_key": month_key,
                        "existing_report_file": seen[month_key].get("report_file"),
                        "conflicting_report_file": migrated.get("report_file"),
                    }
                ]
            )
        seen[month_key] = migrated
        if (
            record.get("year") != migrated["year"]
            or record.get("month") != migrated["month"]
            or record.get("month_key") != migrated["month_key"]
        ):
            actions.append(
                {
                    "index": index,
                    "report_file": migrated.get("report_file"),
                    "from": {
                        "year": record.get("year"),
                        "month": record.get("month"),
                        "month_key": record.get("month_key"),
                    },
                    "to": {
                        "year": migrated["year"],
                        "month": migrated["month"],
                        "month_key": migrated["month_key"],
                    },
                    "source": info["source"],
                }
            )
        migrated_records.append(migrated)
    migrated_records.sort(key=lambda item: item.get("month_key", ""))
    return {"version": version, "records": migrated_records}, actions


def load_monitor_history(path=MONITOR_HISTORY_PATH):
    path = Path(path)
    if not path.exists():
        return {"version": 1, "records": []}, []
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    data.setdefault("version", 1)
    data.setdefault("records", [])
    return migrate_monitor_history_records(data)


def normalize_history_record(record):
    normalized = copy.deepcopy(record)
    normalized.pop("updated_at", None)
    return normalized


def history_record_action(history, record):
    for existing in history.get("records", []):
        if existing.get("month_key") == record.get("month_key"):
            return "unchanged" if normalize_history_record(existing) == normalize_history_record(record) else "replace"
    return "add"


def upsert_history_record(history, record):
    records = history.setdefault("records", [])
    for index, existing in enumerate(records):
        if existing.get("month_key") == record.get("month_key"):
            if normalize_history_record(existing) == normalize_history_record(record):
                return "unchanged"
            records[index] = record
            return "replace"
    records.append(record)
    records.sort(key=lambda item: item.get("month_key", ""))
    return "add"


def save_monitor_history_if_changed(history, path=MONITOR_HISTORY_PATH):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    new_text = json.dumps(history, ensure_ascii=False, indent=2) + "\n"
    if path.exists() and path.read_text(encoding="utf-8-sig") == new_text:
        return False
    path.write_text(new_text, encoding="utf-8")
    return True


def build_scanned_history_records(workspace_root, current_report_path=None, default_year=None, fallback_date=None):
    workspace_root = Path(workspace_root)
    current_resolved = str(Path(current_report_path).resolve()).lower() if current_report_path else ""
    records = {}
    for report_path in workspace_root.glob("*月/*通报.doc"):
        if report_path.name.startswith("~$"):
            continue
        if current_resolved and str(report_path.resolve()).lower() == current_resolved:
            continue
        month_info = resolve_score_month(report_path, default_year=default_year, fallback_date=fallback_date)
        year = month_info["score_year"]
        month = month_info["score_month"]
        section = extract_monitor_report_section(report_path)
        selected = extract_monitor_brigades(section)
        if not selected:
            continue
        month_key = make_month_key(year, month)
        if month_key in records:
            raise HistoryMonthConflict(
                [
                    {
                        "month_key": month_key,
                        "existing_report_file": records[month_key].get("report_file"),
                        "conflicting_report_file": str(report_path),
                    }
                ]
            )
        records[month_key] = {
            "year": year,
            "month": month,
            "month_key": month_key,
            "report_file": str(report_path),
            "report_section_hash": sha256_text(section.strip()),
            "selected_cases": [
                {
                    "rank": index,
                    "brigade": item["short"],
                    "brigade_name": item["大队"],
                }
                for index, item in enumerate(selected, 1)
            ],
            "source": "scanned_existing",
        }
    return records


def effective_history_records(history, scanned_records):
    records = {item.get("month_key"): item for item in history.get("records", []) if item.get("month_key")}
    for month_key, record in scanned_records.items():
        records.setdefault(month_key, record)
    return records


def monitor_history_counts(records_by_month, exclude_month_key=None):
    counts = {}
    for month_key, record in records_by_month.items():
        if month_key == exclude_month_key:
            continue
        for item in record.get("selected_cases", []):
            short = item.get("brigade") or brigade_short(item.get("brigade_name"))
            if short:
                counts[short] = counts.get(short, 0) + 1
    return counts


def build_generated_history_record(registration_year, registration_month, report_output, monitor_text, selected_cases, monitor_scores):
    month_info = build_score_month_info(registration_year, registration_month, "argument")
    return {
        "year": month_info["score_year"],
        "month": month_info["score_month"],
        "month_key": month_info["score_month_key"],
        "report_file": str(Path(report_output)),
        "report_section_hash": sha256_text(monitor_text.strip()),
        "selected_cases": [
            {
                "rank": index,
                "brigade": item["short"],
                "brigade_name": item["大队"],
                "contact": item["联系人"],
                "unit": item["单位"],
                "case_score": item["score"],
                "brigade_avg": monitor_scores.get(item["short"], {}).get("avg"),
            }
            for index, item in enumerate(selected_cases, 1)
        ],
        "source": "generated",
        "updated_at": utc_now_text(),
    }


def prepare_monitor_history(workspace_root, default_year, current_month_info, current_report_path=None, fallback_date=None):
    history, migration_actions = load_monitor_history()
    scanned = build_scanned_history_records(
        workspace_root,
        current_report_path,
        default_year=default_year,
        fallback_date=fallback_date,
    )
    records = effective_history_records(history, scanned)
    current_key = current_month_info["score_month_key"]
    return history, scanned, monitor_history_counts(records, exclude_month_key=current_key), migration_actions


def update_monitor_history(history, scanned_records, current_record):
    changed_actions = []
    existing_keys = {item.get("month_key") for item in history.get("records", [])}
    for month_key in sorted(scanned_records):
        if month_key in existing_keys:
            continue
        record = copy.deepcopy(scanned_records[month_key])
        record["updated_at"] = utc_now_text()
        action = upsert_history_record(history, record)
        if action != "unchanged":
            changed_actions.append({"month_key": month_key, "action": action, "source": "scanned_existing"})
    current_action = upsert_history_record(history, current_record)
    if current_action != "unchanged":
        changed_actions.append({"month_key": current_record["month_key"], "action": current_action, "source": "generated"})
    written = save_monitor_history_if_changed(history)
    return {"written": written, "actions": changed_actions, "current_action": current_action}


def select_monitor_report_cases(monitor_details, monitor_scores=None, history_counts=None):
    order_index = {name: idx for idx, name in enumerate(CASE_ORDER)}
    monitor_scores = monitor_scores or {}
    history_counts = history_counts or {}
    selected = []
    used = set()

    def sort_key(detail):
        avg = monitor_scores.get(detail["short"], {}).get("avg", 999)
        history = history_counts.get(detail["short"], 0)
        return (
            detail["score"],
            avg,
            history,
            order_index.get(detail["short"], 999),
            detail["index"],
        )

    for detail in sorted(
        monitor_details,
        key=sort_key,
    ):
        if detail["short"] in used:
            continue
        selected.append(detail)
        used.add(detail["short"])
        if len(selected) == 3:
            break
    return selected


def validate_text_quality(label, text):
    issues = []
    if re.search(r"[，,、；;。]{2,}", text):
        issues.append(f"{label}存在连续标点")
    if text.count("（") != text.count("）"):
        issues.append(f"{label}中文括号不闭合")
    if text.count("(") != text.count(")"):
        issues.append(f"{label}英文括号不闭合")
    bad_patterns = ["，。", "；。", "、。", "（)", "（）", "扣分：扣分：", "，，", "、、", "；；"]
    for pattern in bad_patterns:
        if pattern in text:
            issues.append(f"{label}存在异常片段：{pattern}")
    if re.search(r"[，,；;]\s*[。；;]", text):
        issues.append(f"{label}存在空错误项或空分句")
    return issues


def assert_report_text_quality(product_text, monitor_text):
    issues = []
    issues.extend(validate_text_quality("消防产品通报", product_text))
    issues.extend(validate_text_quality("联网监测通报", monitor_text))
    if issues:
        raise ValueError("通报文案自检未通过：" + "；".join(issues))


def build_report_sections(product_records, monitor_details, monitor_scores=None, history_counts=None):
    product_sentences = []
    for item in product_records:
        product_sentences.append(product_report_sentence(item))
    product_text = "；".join(product_sentences) + "。"

    monitor_sentences = []
    for detail in select_monitor_report_cases(monitor_details, monitor_scores, history_counts):
        monitor_sentences.append(
            f"{detail['大队']}{detail['联系人']}联系的{detail['单位']}"
            f"{detail['issues_report']}"
        )
    monitor_text = "。".join(monitor_sentences) + "。"
    assert_report_text_quality(product_text, monitor_text)
    return product_text, monitor_text


def ensure_word_fonts(word_app, fonts):
    available = {str(word_app.FontNames.Item(i)) for i in range(1, word_app.FontNames.Count + 1)}
    missing = [font for font in fonts if font not in available]
    if missing:
        raise RuntimeError("通报字体未安装：" + "、".join(missing))


def apply_report_fonts(doc):
    for paragraph in doc.Paragraphs:
        text = paragraph.Range.Text.replace("\r", "").replace("\x07", "").strip()
        if not text:
            continue
        is_title = text.startswith(("二、", "三、")) or text in {"（四）消防产品", "（五）联网监测"}
        font_name = REPORT_TITLE_FONT if is_title else REPORT_BODY_FONT
        paragraph.Range.Font.Name = font_name
        paragraph.Range.Font.NameFarEast = font_name
        paragraph.Range.Font.Bold = -1 if is_title else 0
        if is_title:
            paragraph.Format.FirstLineIndent = 0
            paragraph.Format.CharacterUnitFirstLineIndent = 0
        else:
            paragraph.Format.CharacterUnitFirstLineIndent = 2


def write_report_doc(template_path, output_path, product_records, monitor_details, force, monitor_scores=None, history_counts=None):
    if output_path.exists() and not force:
        raise FileExistsError(f"目标文件已存在，使用 --force 覆盖生成文件：{output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    import win32com.client

    product_text, monitor_text = build_report_sections(product_records, monitor_details, monitor_scores, history_counts)
    word = win32com.client.DispatchEx("Word.Application")
    word.Visible = False
    word.DisplayAlerts = False
    ensure_word_fonts(word, [REPORT_TITLE_FONT, REPORT_BODY_FONT])
    doc = word.Documents.Open(str(Path(template_path).resolve()), ReadOnly=True)
    try:
        raw = doc.Content.Text
        start_product = raw.find("（四）消防产品")
        start_monitor = raw.find("（五）联网监测")
        start_next = raw.find("三、下一步工作要求")
        if min(start_product, start_monitor, start_next) < 0:
            raise ValueError("通报模板缺少必要章节标记")
        before = raw[:start_product]
        after = raw[start_next:]
        new_text = (
            before
            + "（四）消防产品\r"
            + product_text
            + "\r"
            + "（五）联网监测\r"
            + monitor_text
            + "\r"
            + after
        )
        doc.Content.Text = new_text
        apply_report_fonts(doc)
        doc.SaveAs2(str(Path(output_path).resolve()), FileFormat=0)
    finally:
        doc.Close(False)
        word.Quit()


def generate_product_docs(product_records, product_output_dir, year, month, force, template_path=None):
    if product_output_dir.exists() and force:
        for path in product_output_dir.glob("*产品监督档案.doc"):
            path.unlink()
    product_output_dir.mkdir(parents=True, exist_ok=True)
    batch = build_product_batch(product_records, year, month)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as handle:
        json.dump(batch, handle, ensure_ascii=False, indent=2)
        batch_path = Path(handle.name)
    try:
        writer.batch_process(
            template_path or (SKILL_DIR / "resources" / "空表.doc"),
            batch_path,
            product_output_dir,
            month=month,
            export_pdf=False,
        )
    finally:
        try:
            batch_path.unlink()
        except OSError:
            pass


def build_output_files(month_dir, year, month, include_score_office_record=False):
    month_dir = Path(month_dir)
    product_dir = month_dir / f"{year}年{month}月消防产品监督成绩"
    output_files = {
        "product_dir": str(product_dir),
        "product_summary": str(product_dir / "产品监督成绩总表.xlsx"),
        "personal_stats": str(month_dir / f"{year}年{month}月个人执法统计表.xlsx"),
        "case_scores": str(month_dir / f"{year}年{month}月消防监督管理系统消防执法质量（个案成绩）.xls"),
        "monthly_report": str(month_dir / f"{year}年{month}月通报.doc"),
    }
    if include_score_office_record:
        output_files["office_record"] = str(month_dir / f"{year}年{month}月科室月考核情况记录表.xlsx")
    return output_files


def run(args):
    month_dir = require_path(args.month_dir, "月份目录")
    include_score_office_record = getattr(args, "include_score_office_record", False)
    current_month_info = build_score_month_info(args.year, args.month, "argument")
    templates, template_info = load_monthly_templates(
        args.template_dir,
        allow_snapshot_fallback=args.dry_run,
        include_score_office_record=include_score_office_record,
    )
    product_register = find_one(month_dir, [f"*{args.month}月*产品巡查底册*.docx", "*产品巡查底册*.docx"], "产品巡查底册")
    network_dir = find_one(month_dir, ["*联网监测基础信息考评明细表"], "联网监测明细目录")
    network_stats = require_path(network_dir / "联网监测统计表.xls", "联网监测统计表")
    base_info = find_base_info(month_dir, network_dir)
    personal_template = templates["personal_stats"]
    report_output = month_dir / f"{args.year}年{args.month}月通报.doc"
    product_dir = month_dir / f"{args.year}年{args.month}月消防产品监督成绩"
    personal_output = month_dir / f"{args.year}年{args.month}月个人执法统计表.xlsx"
    office_output = month_dir / f"{args.year}年{args.month}月科室月考核情况记录表.xlsx"
    case_scores_output = month_dir / f"{args.year}年{args.month}月消防监督管理系统消防执法质量（个案成绩）.xls"
    output_files = build_output_files(month_dir, args.year, args.month, include_score_office_record)

    product_records = parse_product_register(product_register)
    product_score_guard = validate_product_score_guard(product_records)
    if product_score_guard["blocking_issues"] and not args.dry_run:
        raise ProductScoreReviewRequired(product_score_guard)
    monitor_scores = read_monitor_scores(network_stats)
    monitor_details = read_monitor_details(base_info, monitor_scores)
    contact_warnings = monitor_contact_warnings(monitor_details)
    person_match_issues = collect_person_match_issues(personal_template, product_records, monitor_details)
    product_detail_leak_issues = product_detail_leak_issues_for_records(product_records)
    if product_detail_leak_issues and not args.dry_run:
        raise HumanReviewRequired(product_detail_leak_issues)
    history, scanned_history, history_counts, history_migration = prepare_monitor_history(
        month_dir.parent,
        args.year,
        current_month_info,
        current_report_path=report_output,
    )
    selected_cases = select_monitor_report_cases(monitor_details, monitor_scores, history_counts)

    if args.dry_run:
        report_text_error = None
        try:
            product_text, monitor_text = build_report_sections(product_records, monitor_details, monitor_scores, history_counts)
        except Exception as exc:
            product_text, monitor_text = "", ""
            report_text_error = str(exc)
        current_record = build_generated_history_record(
            args.year,
            args.month,
            report_output,
            monitor_text,
            selected_cases,
            monitor_scores,
        )
        history_action = "disabled" if args.no_history_update else history_record_action(history, current_record)
        backfill_months = sorted(key for key in scanned_history if key not in {item.get("month_key") for item in history.get("records", [])})
        print(
            json.dumps(
                {
                    "template_source": template_info,
                    "input_files": {
                        "month_dir": str(month_dir),
                        "product_register": str(product_register),
                        "network_dir": str(network_dir),
                        "network_stats": str(network_stats),
                        "base_info": str(base_info),
                    },
                    "output_files": output_files,
                    "include_score_office_record": include_score_office_record,
                    "registration_month": {
                        "year": current_month_info["registration_year"],
                        "month": current_month_info["registration_month"],
                        "month_key": current_month_info["registration_month_key"],
                    },
                    "score_month": {
                        "year": current_month_info["score_year"],
                        "month": current_month_info["score_month"],
                        "month_key": current_month_info["score_month_key"],
                        "source": current_month_info["source"],
                    },
                    "products": [
                        {
                            "大队": item["大队"],
                            "题名": item["题名"],
                            "立卷人": item["立卷人"],
                            "score": item["score"],
                            "no_case": item["no_case"],
                            "errors": item["errors"],
                            "archive_errors": item["archive_errors"],
                            "ignored_yellow_errors": item.get("ignored_yellow_errors", []),
                        }
                        for item in product_records
                    ],
                    "product_score_guard": product_score_guard,
                    "product_detail_leak_issues": product_detail_leak_issues,
                    "person_match_issues": person_match_issues,
                    "monitor_contact_warnings": contact_warnings,
                    "monitor_score_warnings": read_monitor_scores.last_warnings,
                    "monitor_avg": {k: v["avg"] for k, v in monitor_scores.items()},
                    "monitor_report_history_counts": history_counts,
                    "monitor_report_history_source": {
                        "path": str(MONITOR_HISTORY_PATH),
                        "migrated_records": history_migration,
                        "scanned_missing_months": backfill_months,
                    },
                    "history_record_would_update": history_action,
                    "monitor_report_cases": selected_cases,
                    "product_result_preview": [
                        {
                            "大队": item["大队"],
                            "broad_issues": unique_text_list(item.get("archive_errors", [])),
                            "ignored_yellow_errors": item.get("ignored_yellow_errors", []),
                            "office_record_text": product_office_text(item),
                            "report_sentence": product_report_sentence(item),
                        }
                        for item in product_records
                    ],
                    "report_text_preview": {
                        "product": product_text,
                        "monitor": monitor_text,
                    },
                    "report_text_quality": {
                        "product": validate_text_quality("消防产品通报", product_text),
                        "monitor": validate_text_quality("联网监测通报", monitor_text),
                    },
                    "report_text_error": report_text_error,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    if personal_template.resolve() == personal_output.resolve():
        raise FileNotFoundError("月份目录缺少个人执法统计表模板，不能用已生成成品覆盖自身")
    generate_product_docs(product_records, product_dir, args.year, args.month, args.force, templates["product_archive_detail"])
    write_product_summary(
        templates["product_summary"],
        product_dir / "产品监督成绩总表.xlsx",
        product_records,
        args.force,
    )
    personal_warnings = write_personal_stats(
        personal_template,
        personal_output,
        product_records,
        monitor_details,
        args.month,
        args.force,
    )
    if include_score_office_record:
        write_office_record(
            templates["office_record"],
            office_output,
            product_records,
            monitor_details,
            args.force,
        )
    write_case_scores(
        templates["case_scores"],
        case_scores_output,
        product_records,
        monitor_scores,
        args.force,
    )
    write_report_doc(
        templates["monthly_report"],
        report_output,
        product_records,
        monitor_details,
        args.force,
        monitor_scores,
        history_counts,
    )
    product_text, monitor_text = build_report_sections(product_records, monitor_details, monitor_scores, history_counts)
    current_record = build_generated_history_record(
        args.year,
        args.month,
        report_output,
        monitor_text,
        selected_cases,
        monitor_scores,
    )
    history_update = {"written": False, "actions": [], "current_action": "disabled"}
    if not args.no_history_update:
        history_update = update_monitor_history(history, scanned_history, current_record)

    summary = {
        "template_source": template_info,
        "input_files": {
            "month_dir": str(month_dir),
            "product_register": str(product_register),
            "network_dir": str(network_dir),
            "network_stats": str(network_stats),
            "base_info": str(base_info),
        },
        "output_files": output_files,
        "include_score_office_record": include_score_office_record,
        "registration_month": {
            "year": current_month_info["registration_year"],
            "month": current_month_info["registration_month"],
            "month_key": current_month_info["registration_month_key"],
        },
        "score_month": {
            "year": current_month_info["score_year"],
            "month": current_month_info["score_month"],
            "month_key": current_month_info["score_month_key"],
            "source": current_month_info["source"],
        },
        "product_dir": str(product_dir),
        "product_scores": {item["short"]: item["score"] for item in product_records},
        "product_score_guard": product_score_guard,
        "monitor_contact_warnings": contact_warnings,
        "monitor_score_warnings": read_monitor_scores.last_warnings,
        "monitor_avg": {k: v["avg"] for k, v in monitor_scores.items()},
        "personal_warnings": personal_warnings,
        "monitor_report_cases": [
            {
                "大队": item["大队"],
                "联系人": item["联系人"],
                "单位": item["单位"],
                "score": item["score"],
            }
            for item in selected_cases
        ],
        "monitor_report_history_counts": history_counts,
        "monitor_report_history_migration": history_migration,
        "monitor_report_history_update": history_update,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(
        description="按月生成产品与联网监测考评成绩登记文件。",
        epilog="参考：references/monthly/source_product_register.md、source_monitor_base_info.md、source_monitor_stats.md 和 G01-G06 对象文档。",
    )
    parser.add_argument("--month-dir", required=True, help="月份目录，例如 C:\\...\\5月")
    parser.add_argument("--year", type=int, required=True, help="年份，例如 2026")
    parser.add_argument("--month", type=int, required=True, help="月份，例如 5")
    parser.add_argument("--force", action="store_true", help="覆盖已生成的同名成品文件")
    parser.add_argument("--dry-run", action="store_true", help="只解析并输出数据，不写入文件")
    parser.add_argument("--template-dir", help="临时覆盖月度模板根目录；默认使用 monthly_workflow.json 的外部事实源")
    parser.add_argument("--no-history-update", action="store_true", help="生成文件但不更新联网通报历史台账")
    parser.add_argument(
        "--include-score-office-record",
        action="store_true",
        help="兼容旧流程：在巡查/成绩目录内额外生成成绩月份科室月考核情况记录表",
    )
    args = parser.parse_args()
    try:
        run(args)
    except ProductScoreReviewRequired as exc:
        print(
            json.dumps(
                {
                    "status": "product_score_review_required",
                    "message": "产品底册扣分超过原版评查规则上限或无法落格，已暂停任务，请人工核对。",
                    "product_score_guard": exc.guard,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        raise SystemExit(2)
    except HumanReviewRequired as exc:
        print(
            json.dumps(
                {
                    "status": "human_review_required",
                    "message": "个人成绩登记涉及未在表格出现的人员，已暂停任务，请人工核对。",
                    "issues": exc.issues,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        raise SystemExit(2)
    except HistoryMonthConflict as exc:
        print(
            json.dumps(
                {
                    "status": "history_month_conflict",
                    "message": "联网通报历史月份发生冲突，已暂停任务，请人工核对。",
                    "issues": exc.issues,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        raise SystemExit(2)


if __name__ == "__main__":
    main()
