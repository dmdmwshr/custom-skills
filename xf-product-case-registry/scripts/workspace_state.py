from __future__ import annotations

import json
import os
import re
import stat
import tomllib
import unicodedata
import uuid
from collections.abc import Mapping
from contextlib import contextmanager, suppress
from copy import deepcopy
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from time import monotonic, sleep
from typing import Any

WORKSPACE_SCHEMA = "CaseWaterlineV1"
WORKSPACE_CONFIG_NAME = "workspace.toml"
WATERLINE_JSON_NAME = "案卷水位记录.json"
WATERLINE_XLSX_NAME = "案卷水位记录表.xlsx"
PROJECT_NO = re.compile(r"^\d{8}[A-Z]\d{9}$")
SHA256 = re.compile(r"^sha256:[a-f0-9]{64}$")
SAFE_NAME = re.compile(r"^[^<>:\"/\\|?*\x00-\x1f]{1,120}$")
ARCHIVE_GENERATION = re.compile(r"^\d{8}T\d{6}Z-[a-f0-9]{12}$")
TERMINAL_BATCH_STATES = {
    "COMPLETE",
    "COMPLETED",
    "FINALIZED",
    "READY_FOR_COMPOSE",
    "READY_FOR_ORGANIZATION",
    "ACCEPTANCE_COMPLETE",
    "STABLE",
}
WATERLINE_LOCK_TIMEOUT_SECONDS = 10.0
WATERLINE_LOCK_POLL_SECONDS = 0.05


class WorkspaceStateError(RuntimeError):
    """A workspace boundary, state, or persistence operation was rejected."""


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _timestamp_slug(value: str) -> str:
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as error:
        raise WorkspaceStateError("核验时间必须是 ISO 8601 时间") from error
    if parsed.tzinfo is None:
        raise WorkspaceStateError("核验时间必须包含时区")
    return parsed.astimezone(UTC).strftime("%Y%m%dT%H%M%SZ")


def _absolute_path(value: str | os.PathLike[str], label: str) -> Path:
    path = Path(value)
    if not path.is_absolute():
        raise WorkspaceStateError(f"{label}必须是绝对路径")
    return Path(os.path.abspath(path))


def _is_reparse(path: Path) -> bool:
    try:
        metadata = path.lstat()
    except OSError as error:
        raise WorkspaceStateError(f"无法检查路径：{path}") from error
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    return path.is_symlink() or bool(getattr(metadata, "st_file_attributes", 0) & reparse_flag)


def _reject_reparse_chain(path: Path, label: str) -> None:
    for candidate in (path, *path.parents):
        if not (candidate.exists() or candidate.is_symlink()):
            continue
        if _is_reparse(candidate):
            raise WorkspaceStateError(f"{label}不允许符号链接、联接或其他重解析点：{candidate}")


def _inside_or_equal(child: Path, parent: Path) -> bool:
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False


def _discover_repository(path: Path) -> Path | None:
    for candidate in (path, *path.parents):
        if (candidate / ".git").exists():
            return candidate
    return None


def secure_workspace_root(value: str | os.PathLike[str]) -> Path:
    """Validate a business workspace root without following reparse points."""

    root = _absolute_path(value, "工作根目录")
    _reject_reparse_chain(root, "工作根目录")
    if root.exists() and not root.is_dir():
        raise WorkspaceStateError("工作根目录必须是目录")

    skill_root = Path(__file__).absolute().parents[1]
    repository_root = _discover_repository(skill_root)
    forbidden = {skill_root}
    if repository_root is not None:
        forbidden.add(repository_root)
    for item in forbidden:
        absolute = Path(os.path.abspath(item))
        if _inside_or_equal(root, absolute) or _inside_or_equal(absolute, root):
            raise WorkspaceStateError("工作根目录不得位于或包含代码/Skill 仓库")

    containing_repo = _discover_repository(root if root.exists() else root.parent)
    if containing_repo is not None and _inside_or_equal(root, containing_repo):
        raise WorkspaceStateError(f"工作根目录不得位于代码仓库：{containing_repo}")
    return root


def secure_download_dir(value: str | os.PathLike[str]) -> Path:
    path = _absolute_path(value, "下载目录")
    _reject_reparse_chain(path, "下载目录")
    if not path.is_dir():
        raise WorkspaceStateError(f"下载目录不存在或不是目录：{path}")
    return path


def default_workspace_config_path() -> Path:
    local_app_data = os.environ.get("LOCALAPPDATA")
    root = Path(local_app_data) if local_app_data else Path.home() / "AppData" / "Local"
    root = _absolute_path(root, "LOCALAPPDATA")
    return root / "xf-product-case-registry" / WORKSPACE_CONFIG_NAME


@dataclass(frozen=True, slots=True)
class WorkspaceConfig:
    work_root: Path
    download_dir: Path
    config_path: Path


@dataclass(frozen=True, slots=True)
class BusinessLayout:
    root: Path
    raw_root: Path
    capture_screenshots: Path
    pending_cases: Path
    completed_cases: Path
    work_root: Path
    capture_batches: Path
    verification_records: Path
    history_workspaces: Path
    templates: Path
    waterline_json: Path
    waterline_xlsx: Path

    @classmethod
    def from_root(cls, root: str | os.PathLike[str]) -> BusinessLayout:
        root_path = secure_workspace_root(root)
        raw_root = root_path / "原始案卷"
        work_root = root_path / "工作区"
        return cls(
            root=root_path,
            raw_root=raw_root,
            capture_screenshots=raw_root / "案卷目录截图",
            pending_cases=raw_root / "待处理案卷",
            completed_cases=raw_root / "已处理案卷",
            work_root=work_root,
            capture_batches=work_root / "采集批次",
            verification_records=work_root / "核验记录",
            history_workspaces=work_root / "历史工作区",
            templates=root_path / "资料模板",
            waterline_json=root_path / WATERLINE_JSON_NAME,
            waterline_xlsx=root_path / WATERLINE_XLSX_NAME,
        )

    def batch_dir(self, batch_id: str) -> Path:
        return self.capture_batches / _safe_component(batch_id, "批次编号")

    def pending_case_dir(self, project_no: str) -> Path:
        return self.pending_cases / _project_no(project_no)

    def completed_case_dir(self, project_no: str) -> Path:
        return self.completed_cases / _project_no(project_no)

    def work_case_dir(self, project_no: str) -> Path:
        return self.work_root / _project_no(project_no)


def _safe_component(value: str, label: str) -> str:
    normalized = unicodedata.normalize("NFC", value.strip())
    if normalized in {"", ".", ".."} or not SAFE_NAME.fullmatch(normalized):
        raise WorkspaceStateError(f"{label}不安全")
    return normalized


def _project_no(value: str) -> str:
    normalized = value.strip().upper()
    if not PROJECT_NO.fullmatch(normalized):
        raise WorkspaceStateError("项目编号格式不正确")
    return normalized


def _required_directories(layout: BusinessLayout) -> tuple[Path, ...]:
    return (
        layout.raw_root,
        layout.capture_screenshots,
        layout.pending_cases,
        layout.completed_cases,
        layout.work_root,
        layout.capture_batches,
        layout.verification_records,
        layout.history_workspaces,
        layout.templates,
    )


def ensure_workspace_layout(layout: BusinessLayout) -> BusinessLayout:
    layout.root.mkdir(parents=True, exist_ok=True)
    _reject_reparse_chain(layout.root, "工作根目录")
    for path in _required_directories(layout):
        if path.exists() and (_is_reparse(path) or not path.is_dir()):
            raise WorkspaceStateError(f"工作目录必须是非重解析点目录：{path}")
        path.mkdir(parents=True, exist_ok=True)
        if _is_reparse(path):
            raise WorkspaceStateError(f"工作目录不允许重解析点：{path}")
    return layout


def _config_path(value: str | os.PathLike[str] | None) -> Path:
    path = (
        default_workspace_config_path() if value is None else _absolute_path(value, "工作配置路径")
    )
    _reject_reparse_chain(path, "工作配置路径")
    if path.exists() and not path.is_file():
        raise WorkspaceStateError("工作配置必须是普通文件")
    return path


def _toml_string(value: Path) -> str:
    return json.dumps(str(value), ensure_ascii=False)


def _atomic_bytes(path: Path, content: bytes, *, overwrite: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _reject_reparse_chain(path.parent, "写入目录")
    if path.exists() and _is_reparse(path):
        raise WorkspaceStateError(f"不允许写入重解析点：{path}")
    if not overwrite and path.exists():
        raise WorkspaceStateError(f"目标已存在，拒绝覆盖：{path}")
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("xb") as stream:
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        if overwrite:
            os.replace(temporary, path)
        else:
            try:
                if os.name == "nt":
                    # Windows 同卷 rename 不覆盖现有目标，且可用于不支持硬链接的 exFAT。
                    os.rename(temporary, path)
                else:
                    os.link(temporary, path)
            except FileExistsError as error:
                raise WorkspaceStateError(f"目标已存在，拒绝覆盖：{path}") from error
    except WorkspaceStateError:
        raise
    except OSError as error:
        raise WorkspaceStateError(f"无法安全写入：{path}") from error
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)


def _read_raw_config(path: Path) -> dict[str, Any]:
    try:
        value = tomllib.loads(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError as error:
        raise WorkspaceStateError(f"工作配置不存在：{path}") from error
    except (OSError, UnicodeError, tomllib.TOMLDecodeError) as error:
        raise WorkspaceStateError(f"工作配置无法读取或 TOML 格式错误：{path}") from error
    section = value.get("workspace")
    if not isinstance(section, dict):
        raise WorkspaceStateError("工作配置必须包含 [workspace] 段")
    return section


def read_workspace_config(config_path: str | os.PathLike[str] | None = None) -> WorkspaceConfig:
    path = _config_path(config_path)
    section = _read_raw_config(path)
    root_value, download_value = section.get("root"), section.get("download_dir")
    if not isinstance(root_value, str) or not root_value.strip():
        raise WorkspaceStateError("workspace.root 不能为空")
    if not isinstance(download_value, str) or not download_value.strip():
        raise WorkspaceStateError("workspace.download_dir 不能为空")
    return WorkspaceConfig(
        secure_workspace_root(root_value), secure_download_dir(download_value), path
    )


def _has_active_batches(layout: BusinessLayout) -> bool:
    if not layout.capture_batches.is_dir():
        return False
    for child in layout.capture_batches.iterdir():
        if _is_reparse(child):
            raise WorkspaceStateError(f"采集批次不允许重解析点：{child}")
        if not child.is_dir():
            return True
        state: str | None = None
        for name in ("browser-capture.json", "capture.json", "BrowserCaptureV1.json"):
            candidate = child / name
            if not candidate.is_file():
                continue
            try:
                payload = json.loads(candidate.read_text(encoding="utf-8-sig"))
                raw_state = payload.get("status") or payload.get("state")
                state = raw_state.upper() if isinstance(raw_state, str) else None
            except (OSError, UnicodeError, json.JSONDecodeError):
                state = None
            break
        if state not in TERMINAL_BATCH_STATES:
            return True
    return False


def configure_workspace(
    *,
    work_root: str | os.PathLike[str],
    download_dir: str | os.PathLike[str] | None = None,
    config_path: str | os.PathLike[str] | None = None,
) -> WorkspaceConfig:
    if work_root is None:
        raise WorkspaceStateError("配置工作根时必须显式提供 work_root")
    path = _config_path(config_path)
    root = secure_workspace_root(work_root)
    if download_dir is None:
        download_dir = Path.home() / "Downloads"
    downloads = secure_download_dir(download_dir)

    if path.exists():
        previous = read_workspace_config(path)
        if previous.work_root != root and _has_active_batches(
            BusinessLayout.from_root(previous.work_root)
        ):
            raise WorkspaceStateError("旧工作根仍有未完成采集批次，不得跨根续跑")

    layout = ensure_workspace_layout(BusinessLayout.from_root(root))
    initialize_legacy_waterline(layout)
    content = (
        "# 仅保存本地业务目录；不得写入账号、密码、Cookie 或令牌。\n"
        "[workspace]\n"
        f"root = {_toml_string(root)}\n"
        f"download_dir = {_toml_string(downloads)}\n"
    )
    _atomic_bytes(path, content.encode("utf-8"))
    return WorkspaceConfig(root, downloads, path)


def resolve_workspace(
    *,
    work_root: str | os.PathLike[str] | None = None,
    download_dir: str | os.PathLike[str] | None = None,
    config_path: str | os.PathLike[str] | None = None,
    create_layout: bool = False,
) -> tuple[WorkspaceConfig, BusinessLayout]:
    path = _config_path(config_path)
    stored: WorkspaceConfig | None = None
    if path.exists():
        stored = read_workspace_config(path)
    if work_root is None and stored is None:
        raise WorkspaceStateError("未配置工作根；请先执行 workspace configure")
    root = secure_workspace_root(work_root if work_root is not None else stored.work_root)
    raw_download = (
        download_dir
        if download_dir is not None
        else (stored.download_dir if stored is not None else Path.home() / "Downloads")
    )
    downloads = secure_download_dir(raw_download)
    layout = BusinessLayout.from_root(root)
    if create_layout:
        ensure_workspace_layout(layout)
        initialize_legacy_waterline(layout)
    config = WorkspaceConfig(root, downloads, path)
    return config, layout


def _new_waterline(layout: BusinessLayout) -> dict[str, Any]:
    return {
        "schemaVersion": WORKSPACE_SCHEMA,
        "workspaceRoot": str(layout.root),
        "updatedAt": _utc_now(),
        "cases": {},
    }


def _validate_waterline(layout: BusinessLayout, value: dict[str, Any]) -> dict[str, Any]:
    if value.get("schemaVersion") != WORKSPACE_SCHEMA:
        raise WorkspaceStateError("案卷水位 JSON 版本不支持")
    if value.get("workspaceRoot") != str(layout.root):
        raise WorkspaceStateError("案卷水位 JSON 不属于当前工作根")
    cases = value.get("cases")
    if not isinstance(cases, dict):
        raise WorkspaceStateError("案卷水位 cases 必须是对象")
    for key, record in cases.items():
        if _project_no(key) != key or not isinstance(record, dict):
            raise WorkspaceStateError("案卷水位包含非法项目记录")
        if record.get("projectNo") != key:
            raise WorkspaceStateError(f"案卷水位项目编号不一致：{key}")
        if not isinstance(record.get("state"), str) or not record["state"]:
            raise WorkspaceStateError(f"案卷水位缺少状态：{key}")
    return value


@contextmanager
def _waterline_lock(
    layout: BusinessLayout,
    timeout_seconds: float | None = None,
):
    timeout_seconds = WATERLINE_LOCK_TIMEOUT_SECONDS if timeout_seconds is None else timeout_seconds
    root = secure_workspace_root(layout.root)
    if not root.is_dir():
        raise WorkspaceStateError("工作根不存在，无法锁定案卷水位")
    if timeout_seconds < 0:
        raise WorkspaceStateError("案卷水位锁超时时间不得为负数")
    lock_path = root / ".case-waterline.lock"
    _reject_reparse_chain(lock_path, "案卷水位锁")
    if lock_path.exists() and (not lock_path.is_file() or _is_reparse(lock_path)):
        raise WorkspaceStateError("案卷水位锁必须是非重解析点普通文件")

    flags = os.O_RDWR | os.O_CREAT
    flags |= getattr(os, "O_BINARY", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(lock_path, flags, 0o600)
    except OSError as error:
        raise WorkspaceStateError("无法打开案卷水位锁") from error

    stream = os.fdopen(descriptor, "r+b", buffering=0)
    acquired = False
    try:
        if _is_reparse(lock_path) or not lock_path.is_file():
            raise WorkspaceStateError("案卷水位锁在打开期间变为重解析点")
        if os.fstat(stream.fileno()).st_size < 1:
            stream.write(b"\0")
            stream.flush()
            os.fsync(stream.fileno())
        deadline = monotonic() + timeout_seconds
        while True:
            try:
                stream.seek(0)
                if os.name == "nt":
                    import msvcrt

                    msvcrt.locking(stream.fileno(), msvcrt.LK_NBLCK, 1)
                else:
                    import fcntl

                    fcntl.flock(stream.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                acquired = True
                break
            except OSError as error:
                if monotonic() >= deadline:
                    raise WorkspaceStateError("等待案卷水位锁超时") from error
                sleep(WATERLINE_LOCK_POLL_SECONDS)
        yield
    finally:
        if acquired:
            stream.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(stream.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(stream.fileno(), fcntl.LOCK_UN)
        stream.close()


def load_waterline(layout: BusinessLayout) -> dict[str, Any]:
    _validated_workspace_path(
        layout.waterline_json,
        layout.root,
        "案卷水位 JSON",
        must_exist=layout.waterline_json.exists(),
    )
    if not layout.waterline_json.exists():
        return _new_waterline(layout)
    if _is_reparse(layout.waterline_json):
        raise WorkspaceStateError("案卷水位 JSON 不允许重解析点")
    try:
        value = json.loads(layout.waterline_json.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise WorkspaceStateError("案卷水位 JSON 无法读取或格式错误") from error
    if not isinstance(value, dict):
        raise WorkspaceStateError("案卷水位 JSON 顶层必须是对象")
    return _validate_waterline(layout, value)


def _save_waterline_unlocked(layout: BusinessLayout, value: Mapping[str, Any]) -> Path:
    _validated_workspace_path(
        layout.waterline_json,
        layout.root,
        "案卷水位 JSON",
        must_exist=layout.waterline_json.exists(),
    )
    data = deepcopy(dict(value))
    data["updatedAt"] = _utc_now()
    _validate_waterline(layout, data)
    encoded = (json.dumps(data, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    _atomic_bytes(layout.waterline_json, encoded)
    return layout.waterline_json


def save_waterline(layout: BusinessLayout, value: Mapping[str, Any]) -> Path:
    with _waterline_lock(layout):
        return _save_waterline_unlocked(layout, value)


def _deep_merge(target: dict[str, Any], changes: Mapping[str, Any]) -> None:
    for key, value in changes.items():
        if isinstance(value, Mapping) and isinstance(target.get(key), dict):
            _deep_merge(target[key], value)
        else:
            target[key] = deepcopy(value)


def upsert_case(
    layout: BusinessLayout,
    project_no: str,
    *,
    fields: Mapping[str, Any] | None = None,
    **changes: Any,
) -> dict[str, Any]:
    project = _project_no(project_no)
    merged = dict(fields or {})
    merged.update(changes)
    if "projectNo" in merged and merged["projectNo"] != project:
        raise WorkspaceStateError("不允许修改案卷水位的项目编号")
    with _waterline_lock(layout):
        data = load_waterline(layout)
        now = _utc_now()
        record = data["cases"].get(project)
        if record is None:
            record = {
                "projectNo": project,
                "state": "DISCOVERED",
                "firstSeenAt": now,
                "lastSeenAt": now,
                "source": {"status": "DISCOVERED"},
                "local": {"status": "NOT_STARTED"},
                "upload": {"status": "NOT_STARTED"},
                "nasVerification": {"status": "NOT_STARTED"},
                "completedAt": None,
                "errorSummary": None,
            }
            data["cases"][project] = record
        _deep_merge(record, merged)
        new_state = merged.get("state")
        if isinstance(new_state, str) and new_state:
            for section in ("source", "local", "upload", "nasVerification"):
                section_change = merged.get(section)
                if isinstance(section_change, Mapping) and "status" not in section_change:
                    record[section]["status"] = new_state
        source_change = merged.get("source")
        if isinstance(source_change, Mapping) and "tags" in source_change and "tags" not in merged:
            record["tags"] = deepcopy(source_change["tags"])
        record["projectNo"] = project
        record["lastSeenAt"] = now
        _save_waterline_unlocked(layout, data)
        return deepcopy(record)


def initialize_legacy_waterline(layout: BusinessLayout) -> dict[str, Any]:
    with _waterline_lock(layout):
        data = load_waterline(layout)
        now = _utc_now()
        changed = not layout.waterline_json.exists()
        if layout.pending_cases.is_dir():
            for child in sorted(layout.pending_cases.iterdir(), key=lambda item: item.name):
                if _is_reparse(child):
                    raise WorkspaceStateError(f"待处理案卷不允许重解析点：{child}")
                if not child.is_dir() or not PROJECT_NO.fullmatch(child.name):
                    continue
                if child.name in data["cases"]:
                    continue
                data["cases"][child.name] = {
                    "projectNo": child.name,
                    "state": "LEGACY_NEEDS_REINDEX",
                    "firstSeenAt": now,
                    "lastSeenAt": now,
                    "source": {"status": "LEGACY_NEEDS_REINDEX"},
                    "local": {"status": "LEGACY_NEEDS_REINDEX"},
                    "upload": {"status": "NOT_STARTED"},
                    "nasVerification": {"status": "NOT_STARTED"},
                    "completedAt": None,
                    "errorSummary": "历史案卷待重新清点",
                }
                changed = True
        if changed:
            _save_waterline_unlocked(layout, data)
        return data


def _status(record: Mapping[str, Any], section: str) -> str:
    nested = record.get(section)
    if isinstance(nested, Mapping) and isinstance(nested.get("status"), str):
        return str(nested["status"])
    fallback = record.get(f"{section}Status")
    return str(fallback) if isinstance(fallback, str) else ""


def export_waterline_xlsx(layout: BusinessLayout) -> Path:
    """Render the JSON fact source to Excel; never mutates the JSON on failure."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as error:
        raise WorkspaceStateError("缺少 openpyxl，无法生成案卷水位记录表") from error

    data = load_waterline(layout)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "案卷水位"
    headers = [
        "项目编号",
        "单位名称",
        "所属大队",
        "标签",
        "总体状态",
        "来源采集",
        "本地整理",
        "系统上传",
        "飞牛核验",
        "首次发现时间",
        "最后发现时间",
        "完成时间",
        "错误摘要",
    ]
    sheet.append(headers)
    for project in sorted(data["cases"]):
        record = data["cases"][project]
        tags = record.get("tags", [])
        tags_text = (
            "、".join(str(item) for item in tags) if isinstance(tags, list) else str(tags or "")
        )
        sheet.append(
            [
                project,
                record.get("unitName", ""),
                record.get("brigadeCode") or record.get("brigadeName", ""),
                tags_text,
                record.get("state", ""),
                _status(record, "source"),
                _status(record, "local"),
                _status(record, "upload"),
                _status(record, "nasVerification"),
                record.get("firstSeenAt", ""),
                record.get("lastSeenAt", ""),
                record.get("completedAt", ""),
                record.get("errorSummary", ""),
            ]
        )

    navy, blue, white = "17365D", "D9EAF7", "FFFFFF"
    thin = Side(style="thin", color="B8C6D1")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    state_fills = {
        "COMPLETED": "C6EFCE",
        "VERIFIED": "C6EFCE",
        "LEGACY_NEEDS_REINDEX": "FFF2CC",
        "NEEDS_MANUAL_REVIEW": "FCE4D6",
        "CAPTURE_FAILED": "F4CCCC",
    }
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                # 来源页面文本可以以 =/+/\-/@ 开头；强制按字符串写入，
                # 防止 Excel 将其重解释为公式或外部链接。
                cell.data_type = "s"
            cell.font = Font(name="Arial", size=10, color="1F1F1F")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9E1F2"))
        fill = state_fills.get(str(row[4].value), blue if row[0].row % 2 == 0 else None)
        if fill:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [24, 30, 16, 24, 24, 20, 20, 20, 20, 22, 22, 22, 45]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False
    if sheet.max_row >= 2:
        table = Table(displayName="CaseWaterlineTable", ref=f"A1:M{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    info = workbook.create_sheet("使用说明")
    info.append(["案卷水位记录表"])
    info.append(["数据来源", WATERLINE_JSON_NAME])
    info.append(["使用方式", "本表仅供查看；请勿直接编辑，需要时由 JSON 重新生成。"])
    info.append(["工作根", str(layout.root)])
    info.append(["导出时间", _utc_now()])
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    info.sheet_view.showGridLines = False
    for row in info.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"
            cell.font = Font(name="Arial", size=10, bold=cell.column == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    info["A1"].font = Font(name="Arial", size=16, bold=True, color=navy)

    target = layout.waterline_xlsx
    _validated_workspace_path(
        target,
        layout.root,
        "案卷水位记录表",
        must_exist=target.exists(),
    )
    if target.exists() and _is_reparse(target):
        raise WorkspaceStateError("案卷水位记录表不允许重解析点")
    temporary = target.with_name(f".{target.name}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, target)
    except PermissionError as error:
        raise WorkspaceStateError(
            f"案卷水位 JSON 已保留；请关闭正在打开的 {target.name} 后重新导出"
        ) from error
    except OSError as error:
        raise WorkspaceStateError(f"无法生成案卷水位记录表：{target}") from error
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)
    return target


def _validated_workspace_path(
    path: Path,
    boundary: Path,
    label: str,
    *,
    must_exist: bool,
) -> Path:
    root = secure_workspace_root(boundary)
    if not root.is_dir():
        raise WorkspaceStateError("归档工作根不存在")
    absolute = _absolute_path(path, label)
    if not _inside_or_equal(absolute, root):
        raise WorkspaceStateError(f"{label}超出工作根")
    _reject_reparse_chain(absolute, label)
    if must_exist and not absolute.exists():
        raise WorkspaceStateError(f"{label}不存在：{absolute}")
    if not must_exist and not absolute.parent.is_dir():
        raise WorkspaceStateError(f"{label}的父目录不存在：{absolute.parent}")
    try:
        resolved_root = root.resolve(strict=True)
        resolved = absolute.resolve(strict=must_exist)
    except OSError as error:
        raise WorkspaceStateError(f"无法解析{label}：{absolute}") from error
    if not _inside_or_equal(resolved, resolved_root):
        raise WorkspaceStateError(f"{label}解析后超出工作根")
    return absolute


def _assert_safe_tree(path: Path, boundary: Path, label: str) -> Path:
    path = _validated_workspace_path(path, boundary, label, must_exist=True)
    if not path.is_dir():
        raise WorkspaceStateError(f"{label}必须是目录：{path}")
    if _is_reparse(path):
        raise WorkspaceStateError(f"{label}不允许重解析点：{path}")
    for child in path.rglob("*"):
        if _is_reparse(child):
            raise WorkspaceStateError(f"{label}包含重解析点：{child}")
    return path


def archive_verified_case(
    layout: BusinessLayout,
    project_no: str,
    *,
    upload_status: str,
    verification: Mapping[str, Any],
    manifest_sha256: str,
    package_sha256: str,
    verified_at: str | None = None,
) -> dict[str, Any]:
    """Archive exactly one case after both upload state and FnOS evidence are VERIFIED."""

    project = _project_no(project_no)
    case_id = verification.get("caseId")
    files_verified = verification.get("filesVerified")
    if (
        upload_status != "VERIFIED"
        or verification.get("status") != "VERIFIED"
        or not isinstance(case_id, str)
        or not case_id.strip()
        or type(files_verified) is not int
        or files_verified < 0
    ):
        raise WorkspaceStateError(
            "仅服务端上传和飞牛落盘核验均为 VERIFIED，且核验摘要完整时才能归档"
        )
    if not SHA256.fullmatch(manifest_sha256) or not SHA256.fullmatch(package_sha256):
        raise WorkspaceStateError("归档需要完整的 manifest 和案卷包 SHA-256")
    verified_at = verified_at or _utc_now()
    stamp = _timestamp_slug(verified_at)
    manifest_short = manifest_sha256.removeprefix("sha256:")[:12]
    package_short = package_sha256.removeprefix("sha256:")[:12]
    generation = f"{stamp}-{package_short}"

    pending_source = layout.pending_case_dir(project)
    work_source = layout.work_case_dir(project)
    completed_parent = layout.completed_case_dir(project)
    completed_target = completed_parent / generation
    history_target = layout.history_workspaces / f"{project}-{stamp}-{package_short}"
    evidence_target = layout.verification_records / f"{project}-{stamp}-{manifest_short}.json"
    with _waterline_lock(layout):
        for source, label in (
            (pending_source, "待处理原始案卷"),
            (work_source, "活动项目工作区"),
        ):
            _assert_safe_tree(source, layout.root, label)
        if completed_parent.exists():
            _assert_safe_tree(completed_parent, layout.root, "已处理案卷项目目录")
            legacy_entries = [
                child.name
                for child in completed_parent.iterdir()
                if not child.is_dir() or not ARCHIVE_GENERATION.fullmatch(child.name)
            ]
            if legacy_entries:
                raise WorkspaceStateError("已处理案卷存在旧版平铺内容，需先人工核对迁移")
        else:
            _validated_workspace_path(
                completed_parent,
                layout.root,
                "已处理案卷项目目录",
                must_exist=False,
            )
        for target, label in (
            (history_target, "历史工作区目标"),
            (evidence_target, "核验记录目标"),
        ):
            _validated_workspace_path(target, layout.root, label, must_exist=False)
            if target.exists() or target.is_symlink():
                raise WorkspaceStateError(f"归档目标已存在，拒绝覆盖：{target}")

        moved: list[tuple[Path, Path, str]] = []
        evidence_written = False
        completed_parent_created = False
        data = load_waterline(layout)
        previous_data = deepcopy(data)
        try:
            if not completed_parent.exists():
                _validated_workspace_path(
                    completed_parent,
                    layout.root,
                    "已处理案卷项目目录",
                    must_exist=False,
                )
                completed_parent.mkdir()
                completed_parent_created = True
                _assert_safe_tree(completed_parent, layout.root, "已处理案卷项目目录")
            _validated_workspace_path(
                completed_target, layout.root, "已处理案卷代际目标", must_exist=False
            )
            if completed_target.exists() or completed_target.is_symlink():
                raise WorkspaceStateError(f"归档目标已存在，拒绝覆盖：{completed_target}")
            _assert_safe_tree(pending_source, layout.root, "待处理原始案卷")
            _validated_workspace_path(
                completed_target, layout.root, "已处理案卷代际目标", must_exist=False
            )
            pending_source.replace(completed_target)
            moved.append((completed_target, pending_source, "已处理案卷"))
            _assert_safe_tree(completed_target, layout.root, "已处理案卷目标")

            _assert_safe_tree(work_source, layout.root, "活动项目工作区")
            _validated_workspace_path(
                history_target, layout.root, "历史工作区目标", must_exist=False
            )
            work_source.replace(history_target)
            moved.append((history_target, work_source, "历史工作区"))
            _assert_safe_tree(history_target, layout.root, "历史工作区目标")

            verified_evidence = deepcopy(dict(verification))
            verified_evidence["status"] = "VERIFIED"
            evidence = {
                "recordVersion": 1,
                "projectNo": project,
                "verifiedAt": verified_at,
                "manifestSha256": manifest_sha256,
                "packageSha256": package_sha256,
                "verification": verified_evidence,
                "archivedOriginal": str(completed_target),
                "archivedWorkspace": str(history_target),
            }
            _validated_workspace_path(
                evidence_target, layout.root, "核验记录目标", must_exist=False
            )
            _atomic_bytes(
                evidence_target,
                (json.dumps(evidence, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
                overwrite=False,
            )
            evidence_written = True
            safe_evidence = _validated_workspace_path(
                evidence_target, layout.root, "核验记录目标", must_exist=True
            )
            if not safe_evidence.is_file() or _is_reparse(safe_evidence):
                raise WorkspaceStateError("核验记录目标必须是非重解析点普通文件")
            record = data["cases"].setdefault(
                project,
                {"projectNo": project, "firstSeenAt": verified_at, "lastSeenAt": verified_at},
            )
            history = record.setdefault("history", {})
            if not isinstance(history, dict):
                raise WorkspaceStateError("案卷水位 history 必须是对象")
            completions = history.setdefault("completions", [])
            if not isinstance(completions, list):
                raise WorkspaceStateError("案卷水位 history.completions 必须是数组")
            completions.append(
                {
                    "generation": generation,
                    "verifiedAt": verified_at,
                    "manifestSha256": manifest_sha256,
                    "packageSha256": package_sha256,
                    "originalPath": str(completed_target),
                    "workspacePath": str(history_target),
                    "verificationRecord": str(evidence_target),
                }
            )
            _deep_merge(
                record,
                {
                    "state": "COMPLETED",
                    "local": {"status": "ARCHIVED"},
                    "upload": {"status": "VERIFIED"},
                    "nasVerification": verified_evidence,
                    "completedAt": verified_at,
                    "lastSeenAt": verified_at,
                    "errorSummary": None,
                    "archive": {
                        "generation": generation,
                        "originalPath": str(completed_target),
                        "workspacePath": str(history_target),
                        "verificationRecord": str(evidence_target),
                    },
                },
            )
            _save_waterline_unlocked(layout, data)
        except Exception:
            if evidence_written:
                with suppress(WorkspaceStateError):
                    safe_evidence = _validated_workspace_path(
                        evidence_target,
                        layout.root,
                        "核验记录回滚目标",
                        must_exist=True,
                    )
                    if safe_evidence.is_file():
                        safe_evidence.unlink()
            for archived, original, label in reversed(moved):
                if archived.exists() and not original.exists():
                    _assert_safe_tree(archived, layout.root, f"{label}回滚来源")
                    _validated_workspace_path(
                        original, layout.root, f"{label}回滚目标", must_exist=False
                    )
                    archived.replace(original)
            if completed_parent_created and completed_parent.is_dir():
                with suppress(OSError):
                    completed_parent.rmdir()
            if layout.waterline_json.exists() and previous_data != data:
                with suppress(WorkspaceStateError):
                    _save_waterline_unlocked(layout, previous_data)
            raise

        result = {
            "status": "COMPLETED",
            "projectNo": project,
            "verificationRecord": str(evidence_target),
            "archivedOriginal": str(completed_target),
            "archivedWorkspace": str(history_target),
            "waterlineJson": str(layout.waterline_json),
        }

    excel_error: str | None = None
    try:
        export_waterline_xlsx(layout)
    except WorkspaceStateError as error:
        excel_error = str(error)
    result["waterlineXlsx"] = str(layout.waterline_xlsx) if excel_error is None else None
    result["waterlineXlsxError"] = excel_error
    return result


def doctor_workspace(
    *,
    work_root: str | os.PathLike[str] | None = None,
    download_dir: str | os.PathLike[str] | None = None,
    config_path: str | os.PathLike[str] | None = None,
    create_layout: bool = False,
) -> dict[str, Any]:
    config, layout = resolve_workspace(
        work_root=work_root,
        download_dir=download_dir,
        config_path=config_path,
        create_layout=create_layout,
    )
    missing = [str(path) for path in _required_directories(layout) if not path.is_dir()]
    if create_layout and not missing:
        initialize_legacy_waterline(layout)
    data = load_waterline(layout) if layout.waterline_json.exists() else _new_waterline(layout)
    return {
        "status": "READY" if not missing else "INCOMPLETE",
        "configPath": str(config.config_path),
        "workRoot": str(layout.root),
        "downloadDir": str(config.download_dir),
        "missingDirectories": missing,
        "waterlineSchema": data["schemaVersion"],
        "caseCount": len(data["cases"]),
        "legacyNeedsReindex": sum(
            1 for record in data["cases"].values() if record.get("state") == "LEGACY_NEEDS_REINDEX"
        ),
    }


__all__ = [
    "BusinessLayout",
    "WATERLINE_JSON_NAME",
    "WATERLINE_XLSX_NAME",
    "WORKSPACE_SCHEMA",
    "WorkspaceConfig",
    "WorkspaceStateError",
    "archive_verified_case",
    "configure_workspace",
    "default_workspace_config_path",
    "doctor_workspace",
    "ensure_workspace_layout",
    "export_waterline_xlsx",
    "initialize_legacy_waterline",
    "load_waterline",
    "read_workspace_config",
    "resolve_workspace",
    "save_waterline",
    "secure_download_dir",
    "secure_workspace_root",
    "upsert_case",
]
