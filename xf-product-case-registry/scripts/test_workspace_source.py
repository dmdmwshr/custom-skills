from __future__ import annotations

import io
import json
import multiprocessing
import os
import zipfile
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

import scripts.source_intake as source
import scripts.workspace_state as workspace

PROJECT_A = "99999999T209900001"
PROJECT_B = "99999999T209900002"
FIXED_NOW = "2099-08-21T10:00:00+08:00"
SOURCE_URL = "https://source.example/cases?runId=ephemeral&name=fixture"


@pytest.fixture
def layout(tmp_path: Path) -> workspace.BusinessLayout:
    value = workspace.BusinessLayout.from_root(tmp_path / "business-workspace")
    return workspace.ensure_workspace_layout(value)


def _read_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _record(rwid: str, project_no: str = PROJECT_A, **fields: object) -> dict[str, object]:
    return {
        "RWID": rwid,
        "项目编号": project_no,
        "单位名称": "测试单位",
        **fields,
    }


def _verified_summary(**changes: object) -> dict[str, object]:
    value: dict[str, object] = {
        "status": "VERIFIED",
        "caseId": "fixture-case-id",
        "filesVerified": 1,
        "nasVerifiedAt": FIXED_NOW,
    }
    value.update(changes)
    return value


def _begin_and_stabilize(
    layout: workspace.BusinessLayout,
    items: list[dict[str, object]],
    *,
    batch_id: str = "fixture-batch",
) -> dict[str, object]:
    source.begin_capture(
        layout,
        {"year": 2099, "brigade": "all", "documentType": "消防产品监督检查记录"},
        batch_id=batch_id,
        origin=SOURCE_URL,
        now=FIXED_NOW,
    )
    state: dict[str, object] = {}
    for round_no in (1, 2):
        source.add_page(
            layout,
            batch_id,
            1,
            items,
            len(items),
            1,
            round_no=round_no,
            observed_at=f"2099-08-21T10:0{round_no}:00+08:00",
        )
        state = source.finalize_capture(
            layout,
            batch_id,
            now=f"2099-08-21T10:1{round_no}:00+08:00",
        )
    return state


def _write_zip(
    path: Path, members: dict[str, bytes], *, compression: int = zipfile.ZIP_STORED
) -> None:
    with zipfile.ZipFile(path, "w", compression=compression) as archive:
        for name, content in members.items():
            archive.writestr(name, content)


def _bound_download_baseline(
    layout: workspace.BusinessLayout,
    batch_id: str,
    rwid: str,
    download_dir: Path,
    *,
    observed_at: str = FIXED_NOW,
) -> dict[str, object]:
    capture = _read_json(layout.batch_dir(batch_id) / "browser-capture.json")
    record = capture["records"][rwid]  # type: ignore[index]
    if not record.get("detail"):  # type: ignore[union-attr]
        detail_screenshot = layout.batch_dir(batch_id) / "fixture-detail.png"
        detail_screenshot.write_bytes(b"fixture detail screenshot")
        project_no = record.get("projectNo") or PROJECT_A  # type: ignore[union-attr]
        source.add_detail(
            layout,
            batch_id,
            rwid,
            {"项目编号": project_no, "单位名称": "测试单位", "大队代码": "FIXTURE"},
            f"https://source.example/#/detail?RWID={rwid}",
            detail_screenshot,
            captured_at=observed_at,
        )
    return source.record_download_baseline(
        layout,
        batch_id,
        rwid,
        download_dir,
        observed_at=observed_at,
    )


def _concurrent_upsert_worker(root: str, project_no: str, start: object) -> None:
    child_layout = workspace.BusinessLayout.from_root(root)
    start.wait(10)  # type: ignore[attr-defined]
    workspace.upsert_case(
        child_layout,
        project_no,
        state="DETAIL_CAPTURED",
        unitName=f"并发测试单位-{project_no[-1]}",
    )


def test_workspace_config_precedence_and_dynamic_root(tmp_path: Path) -> None:
    configured_root = tmp_path / "configured-root"
    explicit_root = tmp_path / "explicit-root"
    configured_downloads = tmp_path / "configured-downloads"
    explicit_downloads = tmp_path / "explicit-downloads"
    configured_downloads.mkdir()
    explicit_downloads.mkdir()
    config_path = tmp_path / "config" / "workspace.toml"

    saved = workspace.configure_workspace(
        work_root=configured_root,
        download_dir=configured_downloads,
        config_path=config_path,
    )
    assert saved.work_root == configured_root.absolute()
    assert "password" not in config_path.read_text(encoding="utf-8").casefold()

    from_config, config_layout = workspace.resolve_workspace(config_path=config_path)
    assert from_config.work_root == configured_root.absolute()
    assert from_config.download_dir == configured_downloads.absolute()
    assert config_layout.root == configured_root.absolute()

    overridden, explicit_layout = workspace.resolve_workspace(
        work_root=explicit_root,
        download_dir=explicit_downloads,
        config_path=config_path,
        create_layout=True,
    )
    assert overridden.work_root == explicit_root.absolute()
    assert overridden.download_dir == explicit_downloads.absolute()
    assert explicit_layout.root == explicit_root.absolute()
    assert explicit_layout.pending_cases.is_dir()
    assert workspace.read_workspace_config(config_path).work_root == configured_root.absolute()


@pytest.mark.parametrize("terminal_state", ["READY_FOR_COMPOSE", "READY_FOR_ORGANIZATION"])
def test_workspace_switch_rejects_active_batch(tmp_path: Path, terminal_state: str) -> None:
    first_root = tmp_path / "first-root"
    second_root = tmp_path / "second-root"
    downloads = tmp_path / "downloads"
    downloads.mkdir()
    config_path = tmp_path / "config" / "workspace.toml"
    workspace.configure_workspace(
        work_root=first_root,
        download_dir=downloads,
        config_path=config_path,
    )
    first_layout = workspace.BusinessLayout.from_root(first_root)
    active = first_layout.batch_dir("active-batch")
    active.mkdir(parents=True)
    (active / "browser-capture.json").write_text(
        json.dumps({"schemaVersion": "BrowserCaptureV1", "status": "COLLECTING_LIST"}),
        encoding="utf-8",
    )

    for blocking_state in ("COLLECTING_LIST", "NEEDS_MANUAL_REVIEW", "CAPTURE_FAILED"):
        (active / "browser-capture.json").write_text(
            json.dumps({"schemaVersion": "BrowserCaptureV1", "status": blocking_state}),
            encoding="utf-8",
        )
        with pytest.raises(workspace.WorkspaceStateError, match="未完成采集批次"):
            workspace.configure_workspace(
                work_root=second_root,
                download_dir=downloads,
                config_path=config_path,
            )

    (active / "browser-capture.json").write_text(
        json.dumps({"schemaVersion": "BrowserCaptureV1", "status": terminal_state}),
        encoding="utf-8",
    )
    switched = workspace.configure_workspace(
        work_root=second_root,
        download_dir=downloads,
        config_path=config_path,
    )
    assert switched.work_root == second_root.absolute()


def test_workspace_rejects_repository_relative_and_reparse_roots(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    with pytest.raises(workspace.WorkspaceStateError, match="代码/Skill 仓库"):
        workspace.secure_workspace_root(Path(workspace.__file__).resolve().parents[1] / "cases")
    with pytest.raises(workspace.WorkspaceStateError, match="绝对路径"):
        workspace.secure_workspace_root("relative-cases")

    reparse_root = tmp_path / "reparse-root"
    reparse_root.mkdir()
    original = workspace._is_reparse
    monkeypatch.setattr(
        workspace,
        "_is_reparse",
        lambda path: Path(path) == reparse_root or original(Path(path)),
    )
    with pytest.raises(workspace.WorkspaceStateError, match="重解析点"):
        workspace.secure_workspace_root(reparse_root)


def test_legacy_scan_and_json_excel_readback(layout: workspace.BusinessLayout) -> None:
    legacy = layout.pending_case_dir(PROJECT_A)
    legacy.mkdir(parents=True)
    (legacy / "原始证据.txt").write_text("fixture", encoding="utf-8")
    (layout.pending_cases / "not-a-project").mkdir()

    initialized = workspace.initialize_legacy_waterline(layout)
    assert initialized["schemaVersion"] == "CaseWaterlineV1"
    assert initialized["cases"][PROJECT_A]["state"] == "LEGACY_NEEDS_REINDEX"
    assert "not-a-project" not in initialized["cases"]

    workspace.upsert_case(
        layout,
        PROJECT_A,
        state="DETAIL_CAPTURED",
        unitName="测试单位",
        brigadeCode="FIXTURE",
        tags=["现场检查", "不合格"],
        source={"status": "DETAIL_CAPTURED"},
    )
    json_value = workspace.load_waterline(layout)
    assert json_value["workspaceRoot"] == str(layout.root)
    assert json_value["cases"][PROJECT_A]["unitName"] == "测试单位"

    target = workspace.export_waterline_xlsx(layout)
    workbook = load_workbook(target, read_only=True, data_only=True)
    try:
        sheet = workbook["案卷水位"]
        assert [cell.value for cell in sheet[1]][:5] == [
            "项目编号",
            "单位名称",
            "所属大队",
            "标签",
            "总体状态",
        ]
        assert [cell.value for cell in sheet[2]][:5] == [
            PROJECT_A,
            "测试单位",
            "FIXTURE",
            "现场检查、不合格",
            "DETAIL_CAPTURED",
        ]
        assert workbook["使用说明"]["B2"].value == workspace.WATERLINE_JSON_NAME
    finally:
        workbook.close()
    assert workspace.load_waterline(layout) == json_value


def test_excel_export_forces_source_text_to_strings_without_mutating_json(
    layout: workspace.BusinessLayout,
) -> None:
    workspace.upsert_case(
        layout,
        PROJECT_A,
        unitName="=SUM(1,1)",
        brigadeCode="+FIXTURE",
        tags=["-fixture-tag"],
        errorSummary="@fixture-link",
    )
    before = workspace.load_waterline(layout)
    target = workspace.export_waterline_xlsx(layout)
    workbook = load_workbook(target, read_only=False, data_only=False)
    try:
        row = workbook["案卷水位"][2]
        assert [row[index].value for index in (1, 2, 3, 12)] == [
            "=SUM(1,1)",
            "+FIXTURE",
            "-fixture-tag",
            "@fixture-link",
        ]
        assert all(row[index].data_type == "s" for index in (1, 2, 3, 12))
        formula_cells = [
            cell
            for sheet in workbook.worksheets
            for row_cells in sheet.iter_rows()
            for cell in row_cells
            if cell.data_type == "f"
        ]
        assert formula_cells == []
        assert workbook.calculation is not None
    finally:
        workbook.close()
    assert workspace.load_waterline(layout) == before


def test_concurrent_case_upserts_preserve_both_records(layout: workspace.BusinessLayout) -> None:
    context = multiprocessing.get_context("spawn")
    start = context.Event()
    processes = [
        context.Process(
            target=_concurrent_upsert_worker,
            args=(str(layout.root), project_no, start),
        )
        for project_no in (PROJECT_A, PROJECT_B)
    ]
    try:
        for process in processes:
            process.start()
        start.set()
        for process in processes:
            process.join(15)
        assert [process.exitcode for process in processes] == [0, 0]
    finally:
        for process in processes:
            if process.is_alive():
                process.terminate()
                process.join(5)

    cases = workspace.load_waterline(layout)["cases"]
    assert set(cases) == {PROJECT_A, PROJECT_B}
    assert cases[PROJECT_A]["unitName"] == "并发测试单位-1"
    assert cases[PROJECT_B]["unitName"] == "并发测试单位-2"


def test_atomic_no_overwrite_rejects_publish_race(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "exclusive.json"

    def competing_publish(_temporary: Path, destination: Path) -> None:
        Path(destination).write_bytes(b"competitor")
        raise FileExistsError(destination)

    publish_name = "rename" if os.name == "nt" else "link"
    monkeypatch.setattr(workspace.os, publish_name, competing_publish)
    with pytest.raises(workspace.WorkspaceStateError, match="拒绝覆盖"):
        workspace._atomic_bytes(target, b"ours", overwrite=False)
    assert target.read_bytes() == b"competitor"
    assert not list(tmp_path.glob(".exclusive.json.*.tmp"))


@pytest.mark.skipif(os.name != "nt", reason="Windows/exFAT 发布路径专项")
def test_atomic_no_overwrite_does_not_require_hard_links(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "exclusive-exfat.json"

    def unsupported_link(*_args: object, **_kwargs: object) -> None:
        raise OSError("hard links are not supported")

    monkeypatch.setattr(workspace.os, "link", unsupported_link)
    workspace._atomic_bytes(target, b"portable", overwrite=False)
    assert target.read_bytes() == b"portable"
    with pytest.raises(workspace.WorkspaceStateError, match="拒绝覆盖"):
        workspace._atomic_bytes(target, b"replacement", overwrite=False)


def test_capture_requires_two_stable_rounds_and_deduplicates_rwid(
    layout: workspace.BusinessLayout,
) -> None:
    duplicated = [_record("fixture-rwid-1"), _record("fixture-rwid-1")]
    state = _begin_and_stabilize(layout, duplicated)

    assert state["listResult"] == "STABLE"
    assert state["status"] == "COLLECTING_DETAILS"
    assert state["stableRounds"] == 2
    assert list(state["records"]) == ["fixture-rwid-1"]
    persisted = _read_json(layout.batch_dir("fixture-batch") / "browser-capture.json")
    assert persisted["sourceOrigin"] == "https://source.example/cases?name=fixture"
    assert "runId" not in json.dumps(persisted, ensure_ascii=False)


def test_capture_preserves_document_waterline_and_marks_initial_recheck(
    layout: workspace.BusinessLayout,
) -> None:
    items = [
        _record(
            "fixture-rwid-repeat",
            caseName="测试经营部",
            documentName="消防产品监督检查记录（初查）",
            createdAt="2099-01-01",
            sourcePage=2,
            sourceRow=20,
            sourceOrder=40,
        ),
        _record(
            "fixture-rwid-repeat",
            caseName="测试经营部（个体工商户）",
            documentName="消防产品监督检查记录（复查）",
            createdAt="2099-02-01",
            sourcePage=1,
            sourceRow=1,
            sourceOrder=1,
        ),
    ]

    state = _begin_and_stabilize(layout, items, batch_id="document-waterline")

    assert state["sourceDocumentCount"] == 2
    assert state["uniqueRwidCount"] == 1
    record = state["records"]["fixture-rwid-repeat"]
    assert record["inspectionStages"] == ["INITIAL", "RECHECK"]
    assert [item["inspectionStage"] for item in record["sourceAppearances"]] == [
        "INITIAL",
        "RECHECK",
    ]
    assert record["sourceAppearances"][0]["documentName"].endswith("（初查）")
    assert {item["caseName"] for item in record["sourceAppearances"]} == {
        "测试经营部",
        "测试经营部（个体工商户）",
    }


def test_third_inspection_record_is_nonblocking_anomaly(
    layout: workspace.BusinessLayout,
) -> None:
    items = [
        _record(
            f"fixture-repeat-{index}",
            caseName="同一案卷名称",
            documentName=f"消防产品监督检查记录-{index}",
            sourceOrder=4 - index,
        )
        for index in (1, 2, 3)
    ]

    state = _begin_and_stabilize(layout, items, batch_id="three-inspections")

    assert state["listResult"] == "STABLE"
    assert state["status"] == "COLLECTING_DETAILS"
    assert state["conflicts"] == []
    assert state["sourceDocumentCount"] == 3
    assert state["actionRwids"] == [
        "fixture-repeat-1",
        "fixture-repeat-2",
        "fixture-repeat-3",
    ]
    assert [state["records"][rwid]["inspectionStages"][0] for rwid in state["actionRwids"]] == [
        "INITIAL",
        "RECHECK",
        "ANOMALY",
    ]
    assert state["anomalies"] == [
        {
            "type": "INSPECTION_RECORD_COUNT_EXCEEDED",
            "caseName": "同一案卷名称",
            "recordCount": 3,
            "expectedMaximum": 2,
            "rwids": ["fixture-repeat-1", "fixture-repeat-2", "fixture-repeat-3"],
            "blocking": False,
        }
    ]


def test_acceptance_sample_is_isolated_from_formal_queue_and_waterline(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    batch_id = "acceptance-sample"
    rwid = "fixture-acceptance-rwid"
    source.begin_capture(
        layout,
        {
            "year": 2099,
            "acceptanceMode": "SINGLE_CASE_DOWNLOAD_PROOF",
            "liveTotalCount": 37,
            "sampleCount": 1,
        },
        batch_id=batch_id,
        origin=SOURCE_URL,
        now=FIXED_NOW,
        scope="acceptance",
    )
    state: dict[str, object] = {}
    for round_no in (1, 2):
        source.add_page(
            layout,
            batch_id,
            1,
            [_record(rwid)],
            1,
            1,
            round_no=round_no,
            observed_at=FIXED_NOW,
        )
        state = source.finalize_capture(layout, batch_id, now=FIXED_NOW)
    assert state["scope"] == "acceptance"
    assert state["listContract"] == "SAMPLE_ONLY"
    assert state["updatesGlobalWaterline"] is False
    assert state["listResult"] == "SAMPLE_STABLE"
    assert state["status"] == "ACCEPTANCE_COLLECTING_DETAILS"
    assert workspace.load_waterline(layout)["cases"] == {}

    screenshot = tmp_path / "acceptance.png"
    screenshot.write_bytes(b"acceptance screenshot")
    source.add_detail(
        layout,
        batch_id,
        rwid,
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结果": "合格"},
        "https://source.example/#/detail?runId=discard&RWID=fixture-acceptance-rwid",
        screenshot,
        captured_at=FIXED_NOW,
    )
    downloads = tmp_path / "downloads"
    downloads.mkdir()
    baseline = _bound_download_baseline(layout, batch_id, rwid, downloads)
    _write_zip(downloads / "acceptance.zip", {"document.pdf": b"acceptance package"})
    completed = source.attach_package(
        layout,
        batch_id,
        rwid,
        downloads,
        download_baseline=baseline,
        allowed_download_dir=downloads,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
    )

    assert completed["status"] == "ACCEPTANCE_COMPLETE"
    assert workspace.load_waterline(layout)["cases"] == {}
    assert not layout.pending_case_dir(PROJECT_A).exists()
    acceptance_dir = layout.batch_dir(batch_id) / "验收样本" / PROJECT_A
    assert (acceptance_dir / "source-evidence.json").is_file()
    assert len(list(acceptance_dir.glob("*_案卷包_*.zip"))) == 1


def test_acceptance_sample_rejects_reported_count_above_declared_sample(
    layout: workspace.BusinessLayout,
) -> None:
    batch_id = "acceptance-count-mismatch"
    source.begin_capture(
        layout,
        {
            "year": 2099,
            "acceptanceMode": "SINGLE_CASE_DOWNLOAD_PROOF",
            "liveTotalCount": 37,
            "sampleCount": 1,
        },
        batch_id=batch_id,
        origin=SOURCE_URL,
        now=FIXED_NOW,
        scope="acceptance",
    )
    source.add_page(
        layout,
        batch_id,
        1,
        [_record("fixture-rwid-1"), _record("fixture-rwid-2", PROJECT_B)],
        2,
        1,
        round_no=1,
        observed_at=FIXED_NOW,
    )

    state = source.finalize_capture(layout, batch_id, now=FIXED_NOW)

    assert state["status"] == "NEEDS_MANUAL_REVIEW"
    assert state["listResult"] == "SAMPLE_INVALID"
    assert state["conflicts"] == [
        {
            "type": "ACCEPTANCE_SAMPLE_COUNT_MISMATCH",
            "declaredSampleCount": 1,
            "reportedTotal": 2,
        }
    ]
    assert workspace.load_waterline(layout)["cases"] == {}


def test_acceptance_sample_requires_exactly_one_declared_case(
    layout: workspace.BusinessLayout,
) -> None:
    with pytest.raises(source.SourceIntakeError, match="样本数设为 1"):
        source.begin_capture(
            layout,
            {
                "year": 2099,
                "acceptanceMode": "SINGLE_CASE_DOWNLOAD_PROOF",
                "liveTotalCount": 37,
                "sampleCount": 2,
            },
            batch_id="acceptance-two-cases",
            origin=SOURCE_URL,
            now=FIXED_NOW,
            scope="acceptance",
        )


def test_exclusive_source_json_wraps_publish_os_error_and_cleans_temp(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "exclusive-source.json"

    def failed_publish(*_args: object, **_kwargs: object) -> None:
        raise OSError("fixture publish failure")

    publish_name = "rename" if os.name == "nt" else "link"
    monkeypatch.setattr(source.os, publish_name, failed_publish)

    with pytest.raises(source.SourceIntakeError, match="无法安全发布不可覆盖文件"):
        source._write_json_exclusive(target, {"fixture": True})

    assert not target.exists()
    assert not list(tmp_path.glob("exclusive-source.json.*.tmp"))


def test_exclusive_source_json_wraps_temporary_creation_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "exclusive-source.json"

    def failed_temporary_creation(*_args: object, **_kwargs: object) -> None:
        raise OSError("fixture temporary creation failure")

    monkeypatch.setattr(source.tempfile, "mkstemp", failed_temporary_creation)

    with pytest.raises(source.SourceIntakeError, match="无法安全发布不可覆盖文件"):
        source._write_json_exclusive(target, {"fixture": True})

    assert not target.exists()


def test_capture_defaults_follow_shanghai_new_year_boundary(
    layout: workspace.BusinessLayout,
) -> None:
    before = source.begin_capture(
        layout,
        {},
        batch_id="before-new-year",
        origin=SOURCE_URL,
        now="2099-12-31T23:59:59+08:00",
    )
    after = source.begin_capture(
        layout,
        {},
        batch_id="after-new-year",
        origin=SOURCE_URL,
        now="2100-01-01T00:00:00+08:00",
    )

    assert before["filters"] == {
        "year": 2099,
        "startDate": "2099-01-01",
        "endDate": "2099-12-31",
        "jurisdiction": "全部管辖单位(含派出所)",
        "brigadeScope": "ALL",
        "documentType": "消防产品监督检查记录",
        "documentTypePage": 2,
        "timezone": "Asia/Shanghai",
    }
    assert after["filters"]["year"] == 2100
    assert after["filters"]["startDate"] == "2100-01-01"
    assert after["filters"]["endDate"] == "2100-01-01"
    assert before["workspaceRoot"] == after["workspaceRoot"] == str(layout.root)


@pytest.mark.parametrize(
    "override",
    [
        {"timezone": "UTC"},
        {"startDate": "2099-01-02"},
        {"endDate": "2099-08-20"},
        {"jurisdiction": "全部消防机构"},
        {"documentType": "消防产品监督检查记录表"},
        {"documentTypePage": 1},
    ],
)
def test_capture_rejects_changes_to_fixed_source_filters(
    layout: workspace.BusinessLayout,
    override: dict[str, object],
) -> None:
    with pytest.raises(source.SourceIntakeError, match="采集筛选"):
        source.begin_capture(
            layout,
            override,
            batch_id="invalid-fixed-filter",
            origin=SOURCE_URL,
            now=FIXED_NOW,
        )


@pytest.mark.parametrize(
    "brigade_code",
    ["JIANGYIN", "YIXING", "LIANGXI", "XISHAN", "HUISHAN", "BINHU", "XINWU", "JINGKAI"],
)
def test_capture_accepts_only_standard_single_brigade_codes(
    layout: workspace.BusinessLayout,
    brigade_code: str,
) -> None:
    state = source.begin_capture(
        layout,
        {"brigadeScope": "SINGLE", "brigadeCode": brigade_code},
        batch_id=f"single-{brigade_code.lower()}",
        origin=SOURCE_URL,
        now=FIXED_NOW,
    )
    assert state["filters"]["brigadeScope"] == "SINGLE"
    assert state["filters"]["brigadeCode"] == brigade_code

    with pytest.raises(source.SourceIntakeError, match="8 个标准大队代码"):
        source.begin_capture(
            layout,
            {"brigadeScope": "SINGLE", "brigadeCode": "NOT_A_BRIGADE"},
            batch_id=f"invalid-{brigade_code.lower()}",
            origin=SOURCE_URL,
            now=FIXED_NOW,
        )


def test_capture_checkpoint_rejects_cross_root_resume(tmp_path: Path) -> None:
    first = workspace.ensure_workspace_layout(
        workspace.BusinessLayout.from_root(tmp_path / "first-workspace")
    )
    second = workspace.ensure_workspace_layout(
        workspace.BusinessLayout.from_root(tmp_path / "second-workspace")
    )
    source.begin_capture(
        first,
        {},
        batch_id="cross-root",
        origin=SOURCE_URL,
        now=FIXED_NOW,
    )
    second_batch = second.batch_dir("cross-root")
    second_batch.mkdir(parents=True)
    (second_batch / "browser-capture.json").write_bytes(
        (first.batch_dir("cross-root") / "browser-capture.json").read_bytes()
    )

    with pytest.raises(source.SourceIntakeError, match="另一工作根"):
        source.add_page(
            second,
            "cross-root",
            1,
            [_record("fixture-cross-root")],
            1,
            1,
        )


def test_completed_case_is_queued_until_detail_project_number_is_verified(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    workspace.upsert_case(
        layout,
        PROJECT_A,
        state="COMPLETED",
        completedAt="2099-08-20T00:00:00+08:00",
        source={
            "status": "COMPLETED",
        },
        local={"status": "ARCHIVED", "workspacePath": "fixture-history"},
        upload={"status": "VERIFIED", "caseId": "fixture-case"},
        nasVerification={"status": "VERIFIED", "filesVerified": 1},
        archive={"verificationRecord": "fixture-record"},
        history={"previousCompletion": {"completedAt": "2099-01-01T00:00:00+08:00"}},
    )

    listed = _begin_and_stabilize(
        layout,
        [_record("fixture-incremental", status="复查文书")],
        batch_id="incremental-detail-gated",
    )
    assert listed["status"] == "COLLECTING_DETAILS"
    assert listed["actionRwids"] == ["fixture-incremental"]
    assert workspace.load_waterline(layout)["cases"][PROJECT_A]["state"] == "COMPLETED"

    screenshot = tmp_path / "completed-detail.png"
    screenshot.write_bytes(b"completed detail screenshot")
    resolved = source.add_detail(
        layout,
        "incremental-detail-gated",
        "fixture-incremental",
        {
            "项目编号": PROJECT_A,
            "单位名称": "测试单位",
            "大队代码": "FIXTURE",
            "文书目录": ["初查", "复查"],
        },
        "https://source.example/#/detail?RWID=fixture-incremental",
        screenshot,
        captured_at=FIXED_NOW,
    )

    assert resolved["status"] == "COMPLETED"
    assert resolved["actionRwids"] == []
    resolved_record = resolved["records"]["fixture-incremental"]
    assert resolved_record["skippedAsCompletedProject"] is True
    waterline = workspace.load_waterline(layout)["cases"][PROJECT_A]
    assert waterline["state"] == "COMPLETED"
    assert waterline["completedAt"] == "2099-08-20T00:00:00+08:00"
    assert waterline["local"]["status"] == "ARCHIVED"
    assert waterline["upload"]["status"] == "VERIFIED"
    assert waterline["nasVerification"]["status"] == "VERIFIED"
    assert waterline["source"]["batchId"] == "incremental-detail-gated"
    assert waterline["source"]["projectIdentitySource"] == "DETAIL"


def test_capture_marks_continuously_changing_list_without_false_completion(
    layout: workspace.BusinessLayout,
) -> None:
    source.begin_capture(
        layout,
        {"year": 2099},
        batch_id="changing-batch",
        origin=SOURCE_URL,
        now=FIXED_NOW,
    )
    final: dict[str, object] = {}
    for round_no, rwid in enumerate(("fixture-1", "fixture-2", "fixture-3"), 1):
        source.add_page(
            layout,
            "changing-batch",
            1,
            [_record(rwid)],
            1,
            1,
            round_no=round_no,
        )
        final = source.finalize_capture(layout, "changing-batch", now=FIXED_NOW)

    assert final["status"] == "LIST_CHANGING"
    assert final["listResult"] == "CHANGING"
    assert final["records"] == {}


@pytest.mark.parametrize(
    "items, conflict_type",
    [
        (
            [
                _record("same-rwid", unit="甲"),
                _record("same-rwid", unit="乙"),
            ],
            "RWID_CONFLICT",
        ),
        (
            [
                _record("fixture-rwid-a", unit="甲"),
                _record("fixture-rwid-b", unit="乙"),
            ],
            "PROJECT_CONFLICT",
        ),
    ],
)
def test_capture_conflicts_stop_automatic_progress(
    layout: workspace.BusinessLayout,
    items: list[dict[str, object]],
    conflict_type: str,
) -> None:
    state = _begin_and_stabilize(layout, items, batch_id=f"conflict-{conflict_type.lower()}")
    assert state["status"] == "NEEDS_MANUAL_REVIEW"
    assert any(item["type"] == conflict_type for item in state["conflicts"])


def test_detail_stage_merges_consistent_rwids_that_lacked_list_project_number(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    items = [
        {"RWID": "detail-alias-a", "单位名称": "测试单位"},
        {"RWID": "detail-alias-b", "单位名称": "测试单位"},
    ]
    _begin_and_stabilize(layout, items, batch_id="detail-alias")
    screenshot = tmp_path / "canonical.png"
    screenshot.write_bytes(b"canonical screenshot")
    detail = {
        "项目编号": PROJECT_A,
        "单位名称": "测试单位",
        "单位地址": "测试地址1号",
        "检查结论": "合格",
    }
    source.add_detail(
        layout,
        "detail-alias",
        "detail-alias-a",
        detail,
        "https://source.example/#/detail?RWID=detail-alias-a",
        screenshot,
        captured_at=FIXED_NOW,
    )
    alias_screenshot = tmp_path / "alias.png"
    alias_screenshot.write_bytes(b"alias screenshot")
    merged = source.add_detail(
        layout,
        "detail-alias",
        "detail-alias-b",
        {
            **detail,
            "文书目录": ["消防产品监督检查记录（复查）"],
            "检查结论": "复查合格",
        },
        "https://source.example/#/detail?RWID=detail-alias-b",
        alias_screenshot,
        captured_at=FIXED_NOW,
    )
    assert merged["records"]["detail-alias-b"]["aliasOf"] == "detail-alias-a"
    assert merged["records"]["detail-alias-b"]["projectNo"] == PROJECT_A
    assert merged["records"]["detail-alias-b"]["detail"]["projectInspectionStages"] == ["RECHECK"]
    assert "复查" in merged["records"]["detail-alias-b"]["detail"]["tags"]
    assert merged["actionRwids"] == ["detail-alias-a"]
    assert merged["conflicts"] == []


@pytest.mark.parametrize(
    "detail_url, detail_unit, reason",
    [
        (
            "https://source.example/#/detail?RWID=identity-b",
            "甲公司",
            "详情 URL 的 RWID",
        ),
        (
            "https://source.example/#/detail?RWID=identity-a",
            "丙公司",
            "详情单位名称",
        ),
    ],
)
def test_detail_identity_chain_mismatch_blocks_only_current_case(
    layout: workspace.BusinessLayout,
    tmp_path: Path,
    detail_url: str,
    detail_unit: str,
    reason: str,
) -> None:
    items = [
        {
            "RWID": "identity-a",
            "caseName": "甲公司",
            "documentName": "消防产品监督检查记录(编号：〔2099〕第0001号)",
            "createdAt": "2099-08-21 09:00:00",
        },
        {
            "RWID": "identity-b",
            "caseName": "乙公司",
            "documentName": "消防产品监督检查记录(编号：〔2099〕第0002号)",
            "createdAt": "2099-08-21 08:00:00",
        },
    ]
    _begin_and_stabilize(layout, items, batch_id="identity-chain")
    screenshot = tmp_path / "identity.png"
    screenshot.write_bytes(b"identity screenshot")

    with pytest.raises(source.SourceIntakeError, match="CASE_IDENTITY_CHAIN_MISMATCH"):
        source.add_detail(
            layout,
            "identity-chain",
            "identity-a",
            {"项目编号": PROJECT_A, "单位名称": detail_unit},
            detail_url,
            screenshot,
            captured_at=FIXED_NOW,
        )

    persisted = _read_json(layout.batch_dir("identity-chain") / "browser-capture.json")
    assert persisted["status"] == "COLLECTING_DETAILS"
    assert persisted["records"]["identity-a"]["caseIdentityStatus"] == "MISMATCH"
    assert "caseIdentityStatus" not in persisted["records"]["identity-b"]
    conflict = persisted["records"]["identity-a"]["identityMismatch"]
    assert conflict["type"] == "CASE_IDENTITY_CHAIN_MISMATCH"
    assert conflict["blockingScope"] == "CURRENT_CASE"
    assert reason in conflict["reasons"][0]
    assert not layout.pending_case_dir(PROJECT_A).exists()

    continued = source.add_detail(
        layout,
        "identity-chain",
        "identity-b",
        {"项目编号": PROJECT_B, "单位名称": "乙公司"},
        "https://source.example/#/detail?RWID=identity-b",
        screenshot,
        captured_at=FIXED_NOW,
    )
    assert continued["records"]["identity-b"]["projectNo"] == PROJECT_B
    assert layout.pending_case_dir(PROJECT_B).is_dir()


def test_detail_identity_chain_accepts_individual_business_suffix_normalization(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [{"RWID": "identity-suffix", "caseName": "甲公司（个体工商户）"}],
        batch_id="identity-suffix",
    )
    screenshot = tmp_path / "identity-suffix.png"
    screenshot.write_bytes(b"identity suffix screenshot")

    result = source.add_detail(
        layout,
        "identity-suffix",
        "identity-suffix",
        {"项目编号": PROJECT_A, "单位名称": "甲公司"},
        "https://source.example/#/detail?RWID=identity-suffix",
        screenshot,
        captured_at=FIXED_NOW,
    )

    assert result["records"]["identity-suffix"]["projectNo"] == PROJECT_A
    assert result["conflicts"] == []


def test_detail_stage_blocks_conflicting_rwids_for_same_project(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    items = [
        {"RWID": "detail-conflict-a", "单位名称": "测试单位"},
        {"RWID": "detail-conflict-b", "单位名称": "另一测试单位"},
    ]
    _begin_and_stabilize(layout, items, batch_id="detail-conflict")
    screenshot = tmp_path / "canonical.png"
    screenshot.write_bytes(b"canonical screenshot")
    source.add_detail(
        layout,
        "detail-conflict",
        "detail-conflict-a",
        {
            "项目编号": PROJECT_A,
            "单位名称": "测试单位",
            "单位地址": "测试地址1号",
            "检查结论": "合格",
        },
        "https://source.example/#/detail?RWID=detail-conflict-a",
        screenshot,
        captured_at=FIXED_NOW,
    )
    with pytest.raises(source.SourceIntakeError, match="身份字段冲突"):
        source.add_detail(
            layout,
            "detail-conflict",
            "detail-conflict-b",
            {
                "项目编号": PROJECT_A,
                "单位名称": "另一测试单位",
                "单位地址": "另一测试地址",
                "检查结论": "合格",
            },
            "https://source.example/#/detail?RWID=detail-conflict-b",
            captured_at=FIXED_NOW,
        )
    persisted = _read_json(layout.batch_dir("detail-conflict") / "browser-capture.json")
    assert persisted["status"] == "NEEDS_MANUAL_REVIEW"
    assert any(item["type"] == "PROJECT_DETAIL_CONFLICT" for item in persisted["conflicts"])


def test_detail_evidence_strips_session_data_and_derives_tags(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(layout, [_record("fixture-rwid-detail")])
    screenshot = tmp_path / "detail.png"
    screenshot.write_bytes(b"fixture-image-content")
    detail = {
        "项目编号": PROJECT_A,
        "单位名称": "测试单位",
        "执法单位": "测试大队",
        "大队代码": "FIXTURE",
        "单位地址": "测试地址1号",
        "检查方式": "现场检查并抽样送检",
        "检查结果": "不合格",
        "后续处理": "复查",
        "pageUrl": "https://source.example/detail?runId=discard&RWID=fixture-rwid-detail&token=discard",
        "cookie": "must-not-persist",
    }
    state = source.add_detail(
        layout,
        "fixture-batch",
        "fixture-rwid-detail",
        detail,
        "https://source.example/#/detail?runId=discard&RWID=fixture-rwid-detail",
        screenshot,
        captured_at=FIXED_NOW,
    )

    record = state["records"]["fixture-rwid-detail"]["detail"]
    assert set(record["tags"]) == {"现场检查", "抽样送检", "不合格", "复查", "初查"}
    assert record["sourceUrl"] == "https://source.example/#/detail?RWID=fixture-rwid-detail"
    evidence = _read_json(layout.pending_case_dir(PROJECT_A) / "source-evidence.json")
    serialized = json.dumps(evidence, ensure_ascii=False)
    assert evidence["schemaVersion"] == "SourceEvidenceV1"
    assert "runId" not in serialized
    assert "must-not-persist" not in serialized
    assert "discard" not in serialized
    assert evidence["records"]["fixture-rwid-detail"]["screenshot"]["sha256"].startswith("sha256:")
    waterline = workspace.load_waterline(layout)
    waterline_case = waterline["cases"][PROJECT_A]
    assert waterline_case["state"] == "DETAIL_CAPTURED"
    assert waterline_case["unitName"] == "测试单位"
    assert waterline_case["brigadeName"] == "测试大队"
    assert waterline_case["brigadeCode"] == "FIXTURE"
    assert set(waterline_case["tags"]) == {
        "现场检查",
        "抽样送检",
        "不合格",
        "复查",
        "初查",
    }
    assert waterline_case["source"]["address"] == "测试地址1号"


def test_tags_do_not_treat_negative_free_language_as_noncompliance() -> None:
    negative_free = source._derive_tags(
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结论": "未发现不合格现象"}
    )
    explicitly_qualified = source._derive_tags(
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结论": "[合格]"}
    )
    assert "不合格" not in negative_free
    assert "合格" not in negative_free
    assert "待确认" in negative_free
    assert "合格" in explicitly_qualified
    assert "不合格" not in explicitly_qualified


def test_new_detail_requires_local_screenshot_evidence(layout: workspace.BusinessLayout) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-no-screenshot")],
        batch_id="missing-screenshot",
    )
    with pytest.raises(source.SourceIntakeError, match="必须提供一张完整详情截图"):
        source.add_detail(
            layout,
            "missing-screenshot",
            "fixture-rwid-no-screenshot",
            {
                "项目编号": PROJECT_A,
                "单位名称": "测试单位",
                "检查结果": "合格",
            },
            "https://source.example/#/detail?RWID=fixture-rwid-no-screenshot",
            captured_at=FIXED_NOW,
        )
    persisted = _read_json(layout.batch_dir("missing-screenshot") / "browser-capture.json")
    assert "detail" not in persisted["records"]["fixture-rwid-no-screenshot"]
    assert not (layout.pending_case_dir(PROJECT_A) / "source-evidence.json").exists()


def test_tail_first_cursor_uses_visible_rows_and_crosses_page_boundaries() -> None:
    tail = source.plan_tail_first_cursor(521, 20, 27, 1)
    assert tail["totalPages"] == 27
    assert tail["current"] == {
        "pageNumber": 27,
        "visibleRowCount": 1,
        "rowNumber": 1,
        "sourceIndex": 521,
        "tailOrdinal": 1,
    }
    assert tail["next"] == {
        "pageNumber": 26,
        "rowNumber": None,
        "rowStrategy": "LAST_VISIBLE_ROW_AFTER_REFRESH",
        "requiresVisibleRowReadback": True,
    }

    previous_page = source.plan_tail_first_cursor(521, 20, 26, 20, row_number=20)
    assert previous_page["current"]["sourceIndex"] == 520
    assert previous_page["current"]["tailOrdinal"] == 2
    assert previous_page["next"] == {
        "pageNumber": 26,
        "rowNumber": 19,
        "rowStrategy": "EXACT_ROW",
        "requiresVisibleRowReadback": False,
    }

    with pytest.raises(source.SourceIntakeError, match="页面水位不一致"):
        source.plan_tail_first_cursor(521, 20, 27, 20)


def test_wait_for_download_candidate_distinguishes_ready_stalled_and_ambiguous(
    tmp_path: Path,
) -> None:
    ready_dir = tmp_path / "ready-downloads"
    ready_dir.mkdir()
    ready_baseline = source.capture_download_baseline(ready_dir, observed_at=FIXED_NOW)
    _write_zip(ready_dir / "source-package.zip", {"document.txt": b"ready"})
    ready = source.wait_for_download_candidate(
        ready_dir,
        download_baseline=ready_baseline,
        timeout_seconds=0,
        poll_seconds=0,
        stalled_after_seconds=30,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
        monotonic_fn=lambda: 0.0,
    )
    assert ready["status"] == "READY"
    assert ready["originalSuggestedName"] == "source-package.zip"
    assert ready["zipInspection"]["fileCount"] == 1

    stalled_dir = tmp_path / "stalled-downloads"
    stalled_dir.mkdir()
    stalled_baseline = source.capture_download_baseline(stalled_dir, observed_at=FIXED_NOW)
    (stalled_dir / "source-package.crdownload").write_bytes(b"partial")
    moments = iter((0.0, 0.0, 4.0))
    stalled = source.wait_for_download_candidate(
        stalled_dir,
        download_baseline=stalled_baseline,
        timeout_seconds=10,
        poll_seconds=0,
        stalled_after_seconds=3,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
        monotonic_fn=lambda: next(moments),
    )
    assert stalled["status"] == "STALLED"
    assert stalled["partialCandidates"] == [
        {
            "name": "source-package.crdownload",
            "sizeBytes": len(b"partial"),
            "mtimeNs": (stalled_dir / "source-package.crdownload").stat().st_mtime_ns,
        }
    ]

    ambiguous_dir = tmp_path / "ambiguous-downloads"
    ambiguous_dir.mkdir()
    ambiguous_baseline = source.capture_download_baseline(ambiguous_dir, observed_at=FIXED_NOW)
    _write_zip(ambiguous_dir / "first.zip", {"first.txt": b"first"})
    _write_zip(ambiguous_dir / "second.zip", {"second.txt": b"second"})
    ambiguous = source.wait_for_download_candidate(
        ambiguous_dir,
        download_baseline=ambiguous_baseline,
        timeout_seconds=0,
        poll_seconds=0,
        stalled_after_seconds=30,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
        monotonic_fn=lambda: 0.0,
    )
    assert ambiguous["status"] == "AMBIGUOUS"
    assert {item["name"] for item in ambiguous["zipCandidates"]} == {"first.zip", "second.zip"}


def test_await_download_records_stall_and_auto_attaches_a_complete_zip(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(layout, [_record("fixture-rwid-await")])
    screenshot = tmp_path / "await-detail.png"
    screenshot.write_bytes(b"detail screenshot")
    source.add_detail(
        layout,
        "fixture-batch",
        "fixture-rwid-await",
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结果": "合格"},
        "https://source.example/#/detail?RWID=fixture-rwid-await",
        screenshot,
        captured_at=FIXED_NOW,
    )
    downloads = tmp_path / "downloads"
    downloads.mkdir()
    baseline = _bound_download_baseline(layout, "fixture-batch", "fixture-rwid-await", downloads)
    (downloads / "pending.crdownload").write_bytes(b"partial")
    moments = iter((0.0, 0.0, 5.0))
    stalled = source.await_download(
        layout,
        "fixture-batch",
        "fixture-rwid-await",
        download_baseline=baseline,
        download_dir=downloads,
        allowed_download_dir=downloads,
        timeout_seconds=10,
        poll_seconds=0,
        stalled_after_seconds=3,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
        monotonic_fn=lambda: next(moments),
        attach=True,
    )
    assert stalled["status"] == "STALLED"
    captured_stall = _read_json(layout.batch_dir("fixture-batch") / "browser-capture.json")
    delivery = captured_stall["records"]["fixture-rwid-await"]["downloadDelivery"]
    assert delivery["status"] == "STALLED"
    assert "package" not in captured_stall["records"]["fixture-rwid-await"]
    waterline_case = workspace.load_waterline(layout)["cases"][PROJECT_A]
    assert waterline_case["state"] == "DETAIL_CAPTURED"
    assert waterline_case["source"]["status"] == "PACKAGE_STALLED"

    (downloads / "pending.crdownload").unlink()
    recovery_baseline = _bound_download_baseline(
        layout,
        "fixture-batch",
        "fixture-rwid-await",
        downloads,
        observed_at="2099-08-21T10:30:00+08:00",
    )
    completed = downloads / "completed.zip"
    _write_zip(completed, {"document.txt": b"completed"})
    attached = source.await_download(
        layout,
        "fixture-batch",
        "fixture-rwid-await",
        download_baseline=recovery_baseline,
        download_dir=downloads,
        allowed_download_dir=downloads,
        timeout_seconds=0,
        poll_seconds=0,
        stalled_after_seconds=30,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
        monotonic_fn=lambda: 0.0,
        attach=True,
    )
    assert attached["status"] == "ATTACHED"
    captured_attached = attached["capture"]
    assert captured_attached["records"]["fixture-rwid-await"]["package"]["storedName"].startswith(
        f"{PROJECT_A}_案卷包_"
    )
    assert not completed.exists()
    assert captured_attached["status"] == "READY_FOR_ORGANIZATION"


def test_attach_package_renames_mojibake_zip_and_is_idempotent(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(layout, [_record("fixture-rwid-package")])
    baseline = _bound_download_baseline(layout, "fixture-batch", "fixture-rwid-package", tmp_path)
    downloaded = tmp_path / "downloaded.zip"
    _write_zip(downloaded, {"文书/fixture.txt": b"fixture package"})
    downloaded_bytes = downloaded.read_bytes()

    first = source.attach_package(
        layout,
        "fixture-batch",
        "fixture-rwid-package",
        downloaded,
        original_name="æ¡ˆå·ä¸‹è½½.zip",
        download_baseline=baseline,
        allowed_download_dir=tmp_path,
    )
    assert not downloaded.exists()
    downloaded.write_bytes(downloaded_bytes)
    second = source.attach_package(
        layout,
        "fixture-batch",
        "fixture-rwid-package",
        downloaded,
        original_name="æ¡ˆå·ä¸‹è½½.zip",
        download_baseline=baseline,
        allowed_download_dir=tmp_path,
    )
    package = second["records"]["fixture-rwid-package"]["package"]
    first_package = first["records"]["fixture-rwid-package"]["package"]
    assert {
        **package,
        "downloadDisposition": {
            key: value
            for key, value in package["downloadDisposition"].items()
            if key != "removedAt"
        },
    } == {
        **first_package,
        "downloadDisposition": {
            key: value
            for key, value in first_package["downloadDisposition"].items()
            if key != "removedAt"
        },
    }
    assert package["storedName"].startswith(f"{PROJECT_A}_案卷包_")
    assert package["storedName"].endswith(".zip")
    assert package["originalSuggestedName"] == "æ¡ˆå·ä¸‹è½½.zip"
    assert package["downloadDisposition"]["status"] == "MOVED_TO_WORKSPACE"
    assert not downloaded.exists()
    assert len(list(layout.pending_case_dir(PROJECT_A).glob("*_案卷包_*.zip"))) == 1
    assert package["zipInspection"] == {
        "entryCount": 1,
        "fileCount": 1,
        "totalSizeBytes": len(b"fixture package"),
    }
    assert "extractedRelativePath" not in package
    assert second["status"] == "READY_FOR_ORGANIZATION"
    waterline_case = workspace.load_waterline(layout)["cases"][PROJECT_A]
    assert waterline_case["state"] == "PENDING_ORGANIZATION"
    assert waterline_case["source"]["status"] == "PACKAGE_READY"
    assert waterline_case["local"] == {"status": "PENDING_ORGANIZATION"}


def test_download_directory_requires_private_baseline_and_unique_new_zip(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(layout, [_record("fixture-rwid-baseline")])
    downloads = tmp_path / "downloads"
    downloads.mkdir()
    _write_zip(downloads / "already-there.zip", {"old.txt": b"old"})
    (downloads / "ignored.txt").write_text("not part of baseline", encoding="utf-8")
    (downloads / "unfinished.crdownload").write_bytes(b"partial")
    baseline = _bound_download_baseline(layout, "fixture-batch", "fixture-rwid-baseline", downloads)
    baseline_text = json.dumps(baseline, ensure_ascii=False)
    assert baseline["batchId"] == "fixture-batch"
    assert baseline["rwid"] == "fixture-rwid-baseline"
    assert baseline["projectNo"] == PROJECT_A
    assert baseline["consumedAt"] is None
    assert [item["name"] for item in baseline["files"]] == [
        "already-there.zip",
        "unfinished.crdownload",
    ]
    assert "ignored.txt" not in baseline_text
    assert all("sha256" not in item for item in baseline["files"])
    assert str(downloads) not in baseline_text

    _write_zip(downloads / "new-case.zip", {"new.txt": b"new"})
    state = source.attach_package(
        layout,
        "fixture-batch",
        "fixture-rwid-baseline",
        downloads,
        download_baseline=baseline,
        allowed_download_dir=downloads,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
    )
    package = state["records"]["fixture-rwid-baseline"]["package"]
    assert package["downloadSelection"]["baselineFingerprint"] == baseline["fingerprint"]
    assert package["downloadSelection"]["candidate"]["name"] == "new-case.zip"
    selection_text = json.dumps(package["downloadSelection"], ensure_ascii=False)
    assert str(downloads) not in selection_text
    receipt = layout.root / (str(baseline["relativePath"]) + ".consumed.json")
    consumed = _read_json(receipt)
    assert consumed["rwid"] == "fixture-rwid-baseline"
    assert consumed["projectNo"] == PROJECT_A
    assert consumed["packageSha256"] == package["sha256"]
    assert consumed["consumedAt"]


def test_download_baseline_is_bound_to_one_batch_rwid_and_project(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-one"), _record("fixture-rwid-two", PROJECT_B)],
        batch_id="bound-baseline",
    )
    baseline = _bound_download_baseline(layout, "bound-baseline", "fixture-rwid-one", tmp_path)
    archive = tmp_path / "bound.zip"
    _write_zip(archive, {"document.txt": b"bound"})

    with pytest.raises(source.SourceIntakeError, match="不属于当前 RWID"):
        source.attach_package(
            layout,
            "bound-baseline",
            "fixture-rwid-two",
            archive,
            download_baseline=baseline,
            allowed_download_dir=tmp_path,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )

    forged = dict(baseline)
    forged["projectNo"] = PROJECT_B
    with pytest.raises(source.SourceIntakeError, match="不属于当前项目编号"):
        source.attach_package(
            layout,
            "bound-baseline",
            "fixture-rwid-one",
            archive,
            download_baseline=forged,
            allowed_download_dir=tmp_path,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )

    raw = source.capture_download_baseline(tmp_path, observed_at=FIXED_NOW)
    with pytest.raises(source.SourceIntakeError, match="不属于当前采集批次"):
        source.attach_package(
            layout,
            "bound-baseline",
            "fixture-rwid-one",
            archive,
            download_baseline=raw,
            allowed_download_dir=tmp_path,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )


def test_attach_package_rejects_paths_outside_configured_download_directory(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-download-root")],
        batch_id="download-root",
    )
    configured = tmp_path / "configured-downloads"
    external = tmp_path / "external-downloads"
    configured.mkdir()
    external.mkdir()
    baseline = _bound_download_baseline(
        layout, "download-root", "fixture-rwid-download-root", configured
    )
    _write_zip(external / "same-name.zip", {"document.txt": b"outside"})

    with pytest.raises(source.SourceIntakeError, match="下载目录必须与工作配置完全一致"):
        source.attach_package(
            layout,
            "download-root",
            "fixture-rwid-download-root",
            external,
            download_baseline=baseline,
            allowed_download_dir=configured,
        )
    with pytest.raises(source.SourceIntakeError, match="下载文件必须是已配置下载目录"):
        source.attach_package(
            layout,
            "download-root",
            "fixture-rwid-download-root",
            external / "same-name.zip",
            download_baseline=baseline,
            allowed_download_dir=configured,
        )


def test_download_directory_without_baseline_is_rejected(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-no-baseline")],
        batch_id="no-baseline",
    )
    downloads = tmp_path / "downloads"
    downloads.mkdir()
    _write_zip(downloads / "new-case.zip", {"new.txt": b"new"})
    with pytest.raises(source.SourceIntakeError, match="必须提供本案点击打包前生成的下载基线"):
        source.attach_package(
            layout,
            "no-baseline",
            "fixture-rwid-no-baseline",
            downloads,
            allowed_download_dir=downloads,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )


def test_explicit_unchanged_zip_from_baseline_is_rejected(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-old-zip")],
        batch_id="old-zip",
    )
    old_zip = tmp_path / "already-downloaded.zip"
    _write_zip(old_zip, {"old.txt": b"old"})
    baseline = _bound_download_baseline(layout, "old-zip", "fixture-rwid-old-zip", tmp_path)
    with pytest.raises(source.SourceIntakeError, match="基线中已存在且未变化"):
        source.attach_package(
            layout,
            "old-zip",
            "fixture-rwid-old-zip",
            old_zip,
            download_baseline=baseline,
            allowed_download_dir=tmp_path,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )


def test_rejected_package_can_recover_to_ready_with_audit_trail(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-recovery")],
        batch_id="package-recovery",
    )
    screenshot = tmp_path / "recovery.png"
    screenshot.write_bytes(b"recovery screenshot")
    source.add_detail(
        layout,
        "package-recovery",
        "fixture-rwid-recovery",
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结论": "合格"},
        "https://source.example/#/detail?RWID=fixture-rwid-recovery",
        screenshot,
        captured_at=FIXED_NOW,
    )
    invalid_baseline = _bound_download_baseline(
        layout, "package-recovery", "fixture-rwid-recovery", tmp_path
    )
    invalid = tmp_path / "invalid.zip"
    invalid.write_bytes(b"not a zip")
    with pytest.raises(source.SourceIntakeError, match="不是完整 ZIP"):
        source.attach_package(
            layout,
            "package-recovery",
            "fixture-rwid-recovery",
            invalid,
            download_baseline=invalid_baseline,
            allowed_download_dir=tmp_path,
            stability_interval=0,
            sleep_fn=lambda _seconds: None,
        )
    invalid_receipt = layout.root / (str(invalid_baseline["relativePath"]) + ".consumed.json")
    assert not invalid_receipt.exists()
    rejected = _read_json(layout.batch_dir("package-recovery") / "browser-capture.json")
    assert rejected["status"] == "NEEDS_MANUAL_REVIEW"
    assert any(item["type"] == "PACKAGE_REJECTED" for item in rejected["conflicts"])

    recovery_baseline = _bound_download_baseline(
        layout,
        "package-recovery",
        "fixture-rwid-recovery",
        tmp_path,
        observed_at="2099-08-21T10:20:00+08:00",
    )
    valid = tmp_path / "valid.zip"
    _write_zip(valid, {"document.txt": b"valid"})
    recovered = source.attach_package(
        layout,
        "package-recovery",
        "fixture-rwid-recovery",
        valid,
        download_baseline=recovery_baseline,
        allowed_download_dir=tmp_path,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
    )
    assert recovered["status"] == "READY_FOR_ORGANIZATION"
    assert recovered["conflicts"] == []
    assert any(
        item["type"] == "PACKAGE_REJECTED" and item.get("resolvedAt")
        for item in recovered["resolvedConflicts"]
    )


def test_download_requires_detail_then_preserves_package_evidence(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    _begin_and_stabilize(
        layout,
        [_record("fixture-rwid-package-first")],
        batch_id="package-first",
    )
    with pytest.raises(source.SourceIntakeError, match="必须先进入详情"):
        source.record_download_baseline(
            layout,
            "package-first",
            "fixture-rwid-package-first",
            tmp_path,
            observed_at=FIXED_NOW,
        )
    screenshot = tmp_path / "package-first.png"
    screenshot.write_bytes(b"fixture screenshot")
    source.add_detail(
        layout,
        "package-first",
        "fixture-rwid-package-first",
        {"项目编号": PROJECT_A, "单位名称": "测试单位", "检查结果": "合格"},
        "https://source.example/#/detail?RWID=fixture-rwid-package-first",
        screenshot,
        captured_at=FIXED_NOW,
    )
    baseline = source.record_download_baseline(
        layout,
        "package-first",
        "fixture-rwid-package-first",
        tmp_path,
        observed_at=FIXED_NOW,
    )
    downloaded = tmp_path / "package-first.zip"
    _write_zip(downloaded, {"document.txt": b"fixture"})
    attached = source.attach_package(
        layout,
        "package-first",
        "fixture-rwid-package-first",
        downloaded,
        download_baseline=baseline,
        allowed_download_dir=tmp_path,
        stability_interval=0,
        sleep_fn=lambda _seconds: None,
    )
    package_before = attached["records"]["fixture-rwid-package-first"]["package"]
    evidence = _read_json(layout.pending_case_dir(PROJECT_A) / "source-evidence.json")
    evidence_record = evidence["records"]["fixture-rwid-package-first"]
    assert evidence_record["package"] == package_before
    assert evidence_record["fingerprint"].startswith("sha256:")
    assert evidence_record["screenshot"]["sha256"].startswith("sha256:")


@pytest.mark.parametrize(
    "name,content,match",
    [
        ("unfinished.part", b"not complete", "下载尚未完成"),
        ("not-a-zip.zip", b"plain text", "不是完整 ZIP"),
    ],
)
def test_attach_package_rejects_partial_and_non_zip_downloads(
    layout: workspace.BusinessLayout,
    tmp_path: Path,
    name: str,
    content: bytes,
    match: str,
) -> None:
    _begin_and_stabilize(layout, [_record("fixture-rwid-package")])
    baseline = _bound_download_baseline(layout, "fixture-batch", "fixture-rwid-package", tmp_path)
    candidate = tmp_path / name
    candidate.write_bytes(content)
    with pytest.raises(source.SourceIntakeError, match=match):
        source.attach_package(
            layout,
            "fixture-batch",
            "fixture-rwid-package",
            candidate,
            download_baseline=baseline,
            allowed_download_dir=tmp_path,
        )
    assert candidate.is_file()
    assert not list(layout.pending_case_dir(PROJECT_A).glob("*_案卷包_*.zip"))


@pytest.mark.parametrize(
    "members",
    [
        {"../escape.txt": b"escape"},
        {"Folder/file.txt": b"one", "folder/FILE.TXT": b"two"},
        {"C:/absolute.txt": b"absolute"},
    ],
)
def test_safe_extract_rejects_malicious_zip_without_partial_output(
    layout: workspace.BusinessLayout,
    tmp_path: Path,
    members: dict[str, bytes],
) -> None:
    archive = tmp_path / "malicious.zip"
    _write_zip(archive, members)
    target = layout.work_root / PROJECT_A / "01_解压结果" / "malicious"
    with pytest.raises(source.SourceIntakeError):
        source.safe_extract_package(archive, target)
    assert not target.exists()
    assert not list(target.parent.glob("malicious.extract-*")) if target.parent.exists() else True


@pytest.mark.parametrize(
    "member_name",
    [
        "folder/file.txt:stream",
        "folder/trailing.",
        "folder/trailing ",
        "CON",
        "con.txt",
        "folder/PrN.log",
        "AUX/data.txt",
        "folder/nul.bin",
        "COM1",
        "folder/com9.txt",
        "LPT1/output.txt",
        "folder/lpt9.log",
    ],
)
def test_zip_rejects_windows_unsafe_path_segments_before_extraction(
    layout: workspace.BusinessLayout,
    tmp_path: Path,
    member_name: str,
) -> None:
    archive = tmp_path / "windows-unsafe.zip"
    _write_zip(archive, {member_name: b"fixture"})
    target = layout.work_root / PROJECT_A / "01_解压结果" / "windows-unsafe"
    with pytest.raises(source.SourceIntakeError, match="Windows"):
        source.safe_extract_package(archive, target)
    assert not target.exists()


def test_safe_extract_enforces_resource_limits(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    archive = tmp_path / "oversized.zip"
    _write_zip(archive, {"large.bin": b"12345"})
    target = layout.work_root / PROJECT_A / "01_解压结果" / "oversized"
    with pytest.raises(source.SourceIntakeError, match="总量"):
        source.safe_extract_package(archive, target, max_total_bytes=4)
    assert not target.exists()


def test_safe_extract_rejects_nested_zip_for_manual_review(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    archive = tmp_path / "nested-package.zip"
    _write_zip(archive, {"附件/nested.zip": b"fixture nested archive"})
    target = layout.work_root / PROJECT_A / "01_解压结果" / "nested"
    with pytest.raises(source.SourceIntakeError, match="嵌套 ZIP"):
        source.safe_extract_package(archive, target)
    assert not target.exists()


def test_nested_self_extracting_zip_is_detected_by_content(
    layout: workspace.BusinessLayout, tmp_path: Path
) -> None:
    nested_buffer = io.BytesIO()
    with zipfile.ZipFile(nested_buffer, "w") as nested:
        nested.writestr("inside.txt", b"nested")
    archive = tmp_path / "outer.zip"
    _write_zip(archive, {"attachment.bin": b"MZ" + nested_buffer.getvalue()})
    target = layout.work_root / PROJECT_A / "01_解压结果" / "self-extracting"
    with pytest.raises(source.SourceIntakeError, match="嵌套 ZIP"):
        source.safe_extract_package(archive, target)
    assert not target.exists()


@pytest.mark.parametrize("missing", ["pending", "work"])
def test_archive_requires_both_pending_and_work_sources(tmp_path: Path, missing: str) -> None:
    case_layout = workspace.ensure_workspace_layout(
        workspace.BusinessLayout.from_root(tmp_path / f"missing-{missing}")
    )
    pending = case_layout.pending_case_dir(PROJECT_A)
    active = case_layout.work_case_dir(PROJECT_A)
    if missing != "pending":
        pending.mkdir(parents=True)
        (pending / "original.zip").write_bytes(b"fixture")
    if missing != "work":
        active.mkdir(parents=True)
        (active / "manifest.json").write_text("{}", encoding="utf-8")

    with pytest.raises(workspace.WorkspaceStateError, match="不存在"):
        workspace.archive_verified_case(
            case_layout,
            PROJECT_A,
            upload_status="VERIFIED",
            verification=_verified_summary(),
            manifest_sha256="sha256:" + "a" * 64,
            package_sha256="sha256:" + "b" * 64,
            verified_at=FIXED_NOW,
        )
    assert pending.exists() is (missing != "pending")
    assert active.exists() is (missing != "work")
    assert not case_layout.completed_case_dir(PROJECT_A).exists()
    assert not list(case_layout.history_workspaces.glob(f"{PROJECT_A}-*"))


def test_archive_requires_complete_verified_summary_without_rewriting_existing_status(
    layout: workspace.BusinessLayout,
) -> None:
    pending = layout.pending_case_dir(PROJECT_A)
    active = layout.work_case_dir(PROJECT_A)
    pending.mkdir(parents=True)
    active.mkdir(parents=True)
    workspace.upsert_case(
        layout,
        PROJECT_A,
        state="UPLOADED_PENDING_NAS",
        nasVerification=_verified_summary(),
    )
    invalid_summaries = [
        _verified_summary(status=None),
        {"status": "VERIFIED", "filesVerified": 1},
        {"status": "VERIFIED", "caseId": "", "filesVerified": 1},
        {"status": "VERIFIED", "caseId": "fixture", "filesVerified": True},
        {"status": "VERIFIED", "caseId": "fixture", "filesVerified": -1},
        {"status": "VERIFIED", "caseId": "fixture", "filesVerified": "1"},
    ]
    for summary in invalid_summaries:
        with pytest.raises(workspace.WorkspaceStateError, match="核验摘要完整"):
            workspace.archive_verified_case(
                layout,
                PROJECT_A,
                upload_status="VERIFIED",
                verification=summary,
                manifest_sha256="sha256:" + "a" * 64,
                package_sha256="sha256:" + "b" * 64,
                verified_at=FIXED_NOW,
            )
    assert pending.is_dir() and active.is_dir()
    saved = workspace.load_waterline(layout)["cases"][PROJECT_A]["nasVerification"]
    assert saved["status"] == "VERIFIED"
    assert saved["caseId"] == "fixture-case-id"


def test_archive_rejects_out_of_boundary_and_reparse_targets(
    layout: workspace.BusinessLayout,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pending = layout.pending_case_dir(PROJECT_A)
    active = layout.work_case_dir(PROJECT_A)
    pending.mkdir(parents=True)
    active.mkdir(parents=True)
    outside = tmp_path / "outside-completed"
    outside.mkdir()
    unsafe_layout = replace(layout, completed_cases=outside)
    arguments = {
        "upload_status": "VERIFIED",
        "verification": _verified_summary(),
        "manifest_sha256": "sha256:" + "a" * 64,
        "package_sha256": "sha256:" + "b" * 64,
        "verified_at": FIXED_NOW,
    }

    with pytest.raises(workspace.WorkspaceStateError, match="超出工作根"):
        workspace.archive_verified_case(unsafe_layout, PROJECT_A, **arguments)
    assert pending.is_dir() and active.is_dir()

    original = workspace._is_reparse
    monkeypatch.setattr(
        workspace,
        "_is_reparse",
        lambda path: Path(path) == layout.completed_cases or original(Path(path)),
    )
    with pytest.raises(workspace.WorkspaceStateError, match="重解析点"):
        workspace.archive_verified_case(layout, PROJECT_A, **arguments)
    assert pending.is_dir() and active.is_dir()


def test_archive_preserves_two_generations_for_same_project(
    layout: workspace.BusinessLayout,
) -> None:
    results: list[dict[str, object]] = []
    generations = [
        ("2099-08-20T08:00:00+08:00", "a", "b"),
        ("2099-08-21T09:30:00+08:00", "c", "d"),
    ]
    for index, (verified_at, manifest_char, package_char) in enumerate(generations, 1):
        pending = layout.pending_case_dir(PROJECT_A)
        active = layout.work_case_dir(PROJECT_A)
        pending.mkdir(parents=True)
        active.mkdir(parents=True)
        (pending / f"original-{index}.zip").write_bytes(f"original-{index}".encode())
        (active / f"manifest-{index}.json").write_text("{}", encoding="utf-8")
        results.append(
            workspace.archive_verified_case(
                layout,
                PROJECT_A,
                upload_status="VERIFIED",
                verification=_verified_summary(caseId=f"fixture-case-{index}"),
                manifest_sha256="sha256:" + manifest_char * 64,
                package_sha256="sha256:" + package_char * 64,
                verified_at=verified_at,
            )
        )

    original_paths = [Path(result["archivedOriginal"]) for result in results]
    workspace_paths = [Path(result["archivedWorkspace"]) for result in results]
    assert len(set(original_paths)) == len(set(workspace_paths)) == 2
    assert all(path.is_dir() for path in original_paths + workspace_paths)
    assert {path.parent for path in original_paths} == {layout.completed_case_dir(PROJECT_A)}
    assert (
        len([path for path in layout.completed_case_dir(PROJECT_A).iterdir() if path.is_dir()]) == 2
    )
    history = workspace.load_waterline(layout)["cases"][PROJECT_A]["history"]["completions"]
    assert len(history) == 2
    assert [item["originalPath"] for item in history] == [str(path) for path in original_paths]
    assert [item["workspacePath"] for item in history] == [str(path) for path in workspace_paths]
    assert history[0]["generation"] != history[1]["generation"]


def test_archive_rejects_legacy_flat_completed_case_content(
    layout: workspace.BusinessLayout,
) -> None:
    pending = layout.pending_case_dir(PROJECT_A)
    active = layout.work_case_dir(PROJECT_A)
    completed_parent = layout.completed_case_dir(PROJECT_A)
    pending.mkdir(parents=True)
    active.mkdir(parents=True)
    completed_parent.mkdir(parents=True)
    (pending / "new.zip").write_bytes(b"new")
    (active / "manifest.json").write_text("{}", encoding="utf-8")
    (completed_parent / "legacy-flat.zip").write_bytes(b"legacy")

    with pytest.raises(workspace.WorkspaceStateError, match="旧版平铺内容.*人工"):
        workspace.archive_verified_case(
            layout,
            PROJECT_A,
            upload_status="VERIFIED",
            verification=_verified_summary(),
            manifest_sha256="sha256:" + "a" * 64,
            package_sha256="sha256:" + "b" * 64,
            verified_at=FIXED_NOW,
        )
    assert (pending / "new.zip").is_file()
    assert (active / "manifest.json").is_file()
    assert (completed_parent / "legacy-flat.zip").read_bytes() == b"legacy"


def test_archive_requires_verified_fn_os_and_never_overwrites(
    layout: workspace.BusinessLayout,
) -> None:
    pending = layout.pending_case_dir(PROJECT_A)
    active = layout.work_case_dir(PROJECT_A)
    pending.mkdir(parents=True)
    active.mkdir(parents=True)
    (pending / "original.zip").write_bytes(b"original")
    (active / "manifest.json").write_text("{}", encoding="utf-8")
    workspace.upsert_case(layout, PROJECT_A, state="UPLOADED_PENDING_NAS")
    manifest_sha = "sha256:" + "a" * 64
    package_sha = "sha256:" + "b" * 64

    with pytest.raises(workspace.WorkspaceStateError, match="均为 VERIFIED"):
        workspace.archive_verified_case(
            layout,
            PROJECT_A,
            upload_status="FINALIZED_UNVERIFIED",
            verification=_verified_summary(),
            manifest_sha256=manifest_sha,
            package_sha256=package_sha,
            verified_at=FIXED_NOW,
        )
    with pytest.raises(workspace.WorkspaceStateError, match="均为 VERIFIED"):
        workspace.archive_verified_case(
            layout,
            PROJECT_A,
            upload_status="VERIFIED",
            verification=_verified_summary(status="PENDING"),
            manifest_sha256=manifest_sha,
            package_sha256=package_sha,
            verified_at=FIXED_NOW,
        )
    assert pending.is_dir() and active.is_dir()

    result = workspace.archive_verified_case(
        layout,
        PROJECT_A,
        upload_status="VERIFIED",
        verification=_verified_summary(),
        manifest_sha256=manifest_sha,
        package_sha256=package_sha,
        verified_at=FIXED_NOW,
    )
    assert result["status"] == "COMPLETED"
    assert not pending.exists() and not active.exists()
    archived_original = Path(result["archivedOriginal"])
    assert archived_original.parent == layout.completed_case_dir(PROJECT_A)
    assert (archived_original / "original.zip").read_bytes() == b"original"
    history = Path(result["archivedWorkspace"])
    evidence_path = Path(result["verificationRecord"])
    assert (history / "manifest.json").is_file()
    assert _read_json(evidence_path)["verification"]["status"] == "VERIFIED"
    assert workspace.load_waterline(layout)["cases"][PROJECT_A]["state"] == "COMPLETED"

    pending.mkdir()
    active.mkdir()
    (pending / "new.zip").write_bytes(b"new")
    (active / "new.json").write_text("{}", encoding="utf-8")
    with pytest.raises(workspace.WorkspaceStateError, match="拒绝覆盖"):
        workspace.archive_verified_case(
            layout,
            PROJECT_A,
            upload_status="VERIFIED",
            verification=_verified_summary(),
            manifest_sha256=manifest_sha,
            package_sha256=package_sha,
            verified_at=FIXED_NOW,
        )
    assert (pending / "new.zip").read_bytes() == b"new"
    assert (active / "new.json").is_file()
